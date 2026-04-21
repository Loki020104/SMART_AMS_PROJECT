"""
SmartAMS – Python Flask Backend
Face Recognition API using exact provided code
"""
from dotenv import load_dotenv
load_dotenv()  # Load .env file so SUPABASE_URL, SUPABASE_KEY etc. are available
import os, base64, csv, pickle, json, math
from io import BytesIO
from pathlib import Path
from datetime import datetime
from collections import defaultdict
import logging

# ── Try to import bcrypt (required for password hashing) ──────────
try:
    import bcrypt
except ImportError:
    print("[WARNING] bcrypt not installed. Install with: pip install bcrypt")
    bcrypt = None

# ── Configure logging ──────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format='[%(asctime)s] %(levelname)s: %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)
import uuid
import numpy as np
from PIL import Image

# Initialize dlib for face recognition. Try local model files first,
# then fall back to models bundled with the `face_recognition_models`
# package if available. If neither is present, fall back to detector-only
# mode (no encodings).
FACE_RECOGNITION_AVAILABLE = False
detector = None
sp = None
facerec = None

try:
    import dlib
    from pathlib import Path
    import sys
    import site

    # Try multiple paths for model files
    # Resolve paths relative to this file so backend/ folder works correctly
    _here = Path(__file__).parent.parent  # project root
    candidates = [
        _here / "models" / "shape_predictor_68_face_landmarks.dat",
        _here / "models" / "dlib_face_recognition_resnet_model_v1.dat",
    ]
    
    # Add venv site-packages paths
    for site_pkg in site.getsitepackages() + [site.getusersitepackages()]:
        if site_pkg:
            candidates.append(Path(site_pkg) / "face_recognition_models" / "models" / "shape_predictor_68_face_landmarks.dat")
            candidates.append(Path(site_pkg) / "face_recognition_models" / "models" / "dlib_face_recognition_resnet_model_v1.dat")

    sp_path = None
    rec_path = None
    
    for cand in candidates:
        if cand.exists() and "shape_predictor" in str(cand):
            sp_path = cand
            break
    
    for cand in candidates:
        if cand.exists() and "dlib_face_recognition_resnet_model_v1" in str(cand):
            rec_path = cand
            break

    detector = dlib.get_frontal_face_detector()

    if sp_path and sp_path.exists():
        try:
            sp = dlib.shape_predictor(str(sp_path))
            print(f"[FACE] Loaded shape predictor from {sp_path}")
        except Exception as e:
            print(f"[WARNING] Failed to load shape predictor {sp_path}: {e}")
            sp = None
    else:
        print("[WARNING] shape_predictor_68_face_landmarks.dat not found")
        sp = None

    if rec_path and rec_path.exists():
        try:
            facerec = dlib.face_recognition_model_v1(str(rec_path))
            print(f"[FACE] Loaded face recognition model from {rec_path}")
        except Exception as e:
            print(f"[WARNING] Failed to load face recognition model {rec_path}: {e}")
            facerec = None
    else:
        print("[WARNING] dlib_face_recognition_resnet_model_v1.dat not found")
        facerec = None

    # Only mark full face-recognition available when detector + models exist
    FACE_RECOGNITION_AVAILABLE = detector is not None and sp is not None and facerec is not None
    if FACE_RECOGNITION_AVAILABLE:
        print("[FACE] ✓ dlib module loaded successfully with face detector and models")
    elif detector is not None:
        print("[FACE] ✓ dlib frontal face detector loaded (encodings disabled)")
    else:
        print("[WARNING] dlib not available")
except Exception as e:
    print(f"[WARNING] Error loading dlib: {e}")
    import traceback
    traceback.print_exc()
    FACE_RECOGNITION_AVAILABLE = False

from flask import Flask, request, jsonify, Response
from flask_cors import CORS

# ── Enhanced bulk import system ────────────────────────────────────
try:
    from bulk_routes_enhanced import register_bulk_routes
    BULK_ROUTES_AVAILABLE = True
except ImportError:
    print("[WARNING] bulk_routes_enhanced module not found. Bulk import endpoints disabled.")
    BULK_ROUTES_AVAILABLE = False

# ── Analytics system (Linways-like) ────────────────────────────────
try:
    from analytics_linways import register_analytics_endpoints
    ANALYTICS_AVAILABLE = True
except ImportError:
    print("[WARNING] analytics_linways module not found. Analytics endpoints disabled.")
    ANALYTICS_AVAILABLE = False

# ── Security hardening imports ────────────────────────────────────
try:
    from backend.security_hardening import (
        rate_limit_login, rate_limit_api, apply_security_headers,
        InputValidator, RateLimiter, log_auth_event
    )
except ImportError:
    print("[WARNING] security_hardening module not found. Some security features disabled.")
    rate_limit_login = lambda f: f  # No-op decorator
    rate_limit_api = lambda **kwargs: lambda f: f  # No-op decorator
    InputValidator = None
    log_auth_event = lambda *args, **kwargs: None
    def apply_security_headers(app): pass

app = Flask(__name__)

# ── Configure request size limits for large bulk operations ──────────
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # 50MB max upload
app.config['JSON_MAX_DEPTH'] = 2000  # Allow deep nested JSON

# ── Configure CORS for production (Firebase Hosting frontend) ──────────
cors_origins = [
    "https://smart-ams-project-faa5f.web.app",  # Firebase Hosting
    "http://localhost:3000",  # Local development
    "http://localhost:4200",  # Angular dev
]
CORS(app, resources={r"/api/*": {"origins": cors_origins}}, supports_credentials=True, 
     methods=["GET", "POST", "PUT", "DELETE", "OPTIONS", "PATCH"],
     allow_headers=["Content-Type", "Authorization", "X-Requested-With", "X-Batch-Info"])

# Apply security headers to all responses
apply_security_headers(app)

# ── CORS Preflight Handler ─────────────────────────────────────
@app.before_request
def before_request():
    """Handle CORS preflight requests and set headers early."""
    # Set CORS headers for OPTIONS (preflight) requests
    if request.method == 'OPTIONS':
        response = Response()
        response.headers['Access-Control-Allow-Origin'] = request.headers.get('Origin', '*')
        response.headers['Access-Control-Allow-Methods'] = 'GET, POST, PUT, DELETE, OPTIONS, PATCH, HEAD'
        response.headers['Access-Control-Allow-Headers'] = 'Content-Type, Authorization, X-Requested-With, Accept'
        response.headers['Access-Control-Allow-Credentials'] = 'true'
        response.headers['Access-Control-Max-Age'] = '86400'
        return response

# ── CORS Error Handler & After-Request Hook ────────────────
@app.after_request
def after_request(response):
    """Ensure CORS headers are added to ALL responses, including error responses."""
    origin = request.headers.get('Origin')
    allowed_origins = [
        "https://smart-ams-project-faa5f.web.app",
        "http://localhost:3000",
        "http://localhost:4200",
        "http://localhost:5173",
    ]
    
    # Only set Allow-Origin if origin is in allowed list, otherwise use wildcard
    if origin in allowed_origins:
        response.headers['Access-Control-Allow-Origin'] = origin
    else:
        response.headers['Access-Control-Allow-Origin'] = 'https://smart-ams-project-faa5f.web.app'
    
    response.headers['Access-Control-Allow-Credentials'] = 'true'
    response.headers['Access-Control-Allow-Methods'] = 'GET, POST, PUT, DELETE, OPTIONS, PATCH, HEAD'
    response.headers['Access-Control-Allow-Headers'] = 'Content-Type, Authorization, X-Requested-With, Accept'
    response.headers['Access-Control-Expose-Headers'] = 'Content-Type, X-Total-Count'
    response.headers['Vary'] = 'Origin'
    
    # Prevent caching of API responses
    response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '0'
    return response

@app.errorhandler(500)
def handle_500(e):
    """Handle 500 errors with proper error detail and logging."""
    print(f"[ERROR] 500 Internal Server Error: {str(e)}")
    import traceback
    traceback.print_exc()
    return jsonify(success=False, error="Internal server error", details=str(e)), 500

@app.errorhandler(404)
def handle_404(e):
    """Handle 404 errors."""
    return jsonify(success=False, error="Endpoint not found"), 404

@app.errorhandler(403)
def handle_403(e):
    """Handle 403 errors."""
    return jsonify(success=False, error="Forbidden"), 403

# ── Firebase Admin SDK ─────────────────────────────────────
# On Cloud Run the default service account provides credentials automatically.
# Locally, set GOOGLE_APPLICATION_CREDENTIALS or just provide the project ID.
import firebase_admin
from firebase_admin import credentials, auth as firebase_auth, db as firebase_db, firestore as firebase_firestore
from functools import wraps

FIREBASE_PROJECT_ID = os.getenv("FIREBASE_PROJECT_ID", "smart-ams-project-faa5f")
FIREBASE_DB_URL = os.getenv("FIREBASE_DB_URL", "https://smart-ams-project-faa5f-default-rtdb.firebaseio.com")

_firebase_db_enabled = False
_fstore = None  # Firestore client

_SA_KEY_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'serviceAccountKey.json')
try:
    if os.path.exists(_SA_KEY_PATH):
        cred = credentials.Certificate(_SA_KEY_PATH)
        print(f"[FIREBASE] Using service account key: {_SA_KEY_PATH}")
    else:
        cred = credentials.ApplicationDefault()
        print("[FIREBASE] Using Application Default Credentials")
    firebase_admin.initialize_app(cred, {
        "projectId": FIREBASE_PROJECT_ID,
        "databaseURL": FIREBASE_DB_URL
    })
    _firebase_db_enabled = True
    _fstore = firebase_firestore.client()
    print(f"[FIREBASE] ✓ Admin SDK + RTDB + Firestore initialized (project: {FIREBASE_PROJECT_ID})")
except Exception as e:
    print(f"[FIREBASE] ⚠ Admin SDK not initialized: {e}")

# ── RBAC & Analytics System ──────────────────────────────────────
try:
    from role_based_access_control import (
        apply_rbac_middleware, 
        require_role, 
        require_minimum_role,
        get_user_scope,
        get_accessible_students,
        get_accessible_classes,
        ROLE_HIERARCHY,
    )
    from analytics_rbac import (
        get_class_wise_analytics,
        get_subject_wise_analytics,
        get_student_wise_analytics,
        get_faculty_performance_analytics,
        get_at_risk_students,
        get_daily_report,
        get_weekly_report,
        get_monthly_report,
        get_compliance_report,
    )
    from analytics_rbac_routes import register_rbac_analytics_routes
    print("[RBAC] ✓ Successfully imported RBAC and Analytics modules")
    RBAC_AVAILABLE = True
except ImportError as e:
    print(f"[RBAC] ⚠ Warning: Could not import RBAC modules: {e}")
    print("[RBAC] Analytics and role-based access control will be unavailable")
    RBAC_AVAILABLE = False

# ── Initialize RBAC Middleware (after RBAC imports) ──────────────
if RBAC_AVAILABLE:
    try:
        apply_rbac_middleware(app)
        logger.info("[APP] ✓ RBAC middleware initialized")
    except Exception as e:
        logger.warning(f"[APP] Could not initialize RBAC middleware: {e}")


def rtdb_set(path: str, data: dict):
    """Write (overwrite) a path in Firebase Realtime Database. Silent on failure."""
    if not _firebase_db_enabled:
        return
    try:
        firebase_db.reference(path).set(_clean_for_rtdb(data))
    except Exception as e:
        print(f"[RTDB-SET] Warning: {e}")


def rtdb_update(path: str, data: dict):
    """Update specific fields at an RTDB path. Silent on failure."""
    if not _firebase_db_enabled:
        return
    try:
        firebase_db.reference(path).update(_clean_for_rtdb(data))
    except Exception as e:
        print(f"[RTDB-UPDATE] Warning: {e}")


def rtdb_delete(path: str):
    """Delete an RTDB path. Silent on failure."""
    if not _firebase_db_enabled:
        return
    try:
        firebase_db.reference(path).delete()
    except Exception as e:
        print(f"[RTDB-DELETE] Warning: {e}")


def _clean_for_rtdb(data: dict) -> dict:
    """RTDB cannot store None — replace with empty string."""
    if not isinstance(data, dict):
        return data
    return {k: ("" if v is None else v) for k, v in data.items()}


# ── Firestore helpers ──────────────────────────────────────
def fstore_set(collection: str, doc_id: str, data: dict):
    """Create or overwrite a Firestore document. Silent on failure."""
    if not _fstore:
        return
    try:
        _fstore.collection(collection).document(str(doc_id)).set(data)
    except Exception as e:
        print(f"[FSTORE-SET] Warning: {e}")

def fstore_update(collection: str, doc_id: str, data: dict):
    """Merge-update a Firestore document. Silent on failure."""
    if not _fstore:
        return
    try:
        _fstore.collection(collection).document(str(doc_id)).set(data, merge=True)
    except Exception as e:
        print(f"[FSTORE-UPDATE] Warning: {e}")

def fstore_delete(collection: str, doc_id: str):
    """Delete a Firestore document. Silent on failure."""
    if not _fstore:
        return
    try:
        _fstore.collection(collection).document(str(doc_id)).delete()
    except Exception as e:
        print(f"[FSTORE-DELETE] Warning: {e}")


# ── RTDB Sync Decorator ────────────────────────────────────
import functools

def rtdb_sync(table: str, id_field: str = "id", path_prefix: str = None):
    """
    Decorator for Flask route functions that return a JSON response with
    a top-level 'data' dict (the created/updated record).

    After a successful response it automatically writes the record to RTDB.

    Usage:
        @rtdb_sync("users", id_field="id")
        @app.route("/api/users/register", methods=["POST"])
        def register_user(): ...
    """
    prefix = path_prefix or f"/{table}"

    def decorator(func):
        @functools.wraps(func)
        def wrapper(*args, **kwargs):
            response = func(*args, **kwargs)
            # Only sync on success
            try:
                # Flask may return (response_obj, status_code) or just response_obj
                resp_obj = response[0] if isinstance(response, tuple) else response
                status = response[1] if isinstance(response, tuple) and len(response) > 1 else 200
                if status and int(status) >= 400:
                    return response
                body = resp_obj.get_json(silent=True) if hasattr(resp_obj, 'get_json') else None
                if body and body.get("success") and body.get("data"):
                    record = body["data"]
                    rec_id = record.get(id_field)
                    if rec_id:
                        rtdb_set(f"{prefix}/{rec_id}", record)
            except Exception as e:
                print(f"[RTDB-SYNC] Warning syncing {table}: {e}")
            return response
        return wrapper
    return decorator


def verify_firebase_token(f):
    """Decorator: verify Firebase ID token from Authorization header.

    Sets ``request.firebase_user`` with the decoded token claims (uid, email, etc.).
    If the header is missing or invalid the request is rejected with 401.
    """
    @wraps(f)
    def decorated(*args, **kwargs):
        auth_header = request.headers.get("Authorization", "")
        if not auth_header.startswith("Bearer "):
            return jsonify(success=False, error="Missing or invalid Authorization header"), 401
        id_token = auth_header.split("Bearer ", 1)[1]
        try:
            decoded = firebase_auth.verify_id_token(id_token)
            request.firebase_user = decoded
        except Exception as e:
            print(f"[FIREBASE] Token verification failed: {e}")
            return jsonify(success=False, error="Invalid or expired Firebase token"), 401
        return f(*args, **kwargs)
    return decorated


def sync_firebase_user_to_supabase(firebase_uid, email, display_name, role, extra=None):
    """Create or update a user row in the Supabase ``users`` table.

    This keeps Supabase in sync with Firebase Auth as the primary identity
    provider.  The ``firebase_uid`` column acts as the link between the two.
    """
    if not sb:
        return None
    import hashlib
    try:
        # Check if user already exists by firebase_uid or email
        existing = sb.table("users").select("*").eq("email", email).execute()
        if existing.data:
            user = existing.data[0]
            # Update firebase_uid if not set
            if not user.get("firebase_uid"):
                try:
                    sb_update("users", {"firebase_uid": firebase_uid}, f"id=eq.{user['id']}")
                except Exception:
                    pass
            return user

        # Create new user linked to Firebase
        new_user = {
            "email": email,
            "username": email.split("@")[0],
            "full_name": display_name or email.split("@")[0],
            "role": role or "student",
            "firebase_uid": firebase_uid,
            "password_hash": _hash_password_secure(firebase_uid),
            "is_active": True,
        }
        if extra:
            new_user.update(extra)
        sb.table("users").insert(new_user).execute()
        # Re-fetch
        result = sb.table("users").select("*").eq("email", email).execute()
        return result.data[0] if result.data else new_user
    except Exception as e:
        print(f"[SYNC] Error syncing Firebase user to Supabase: {e}")
        return None


def sb_update(table, data, filter_str):
    """PATCH a row via Supabase REST API.  ``filter_str`` e.g. 'id=eq.abc'."""
    if not SUPABASE_URL or not SUPABASE_KEY:
        return
    url = f"{SUPABASE_URL}/rest/v1/{table}?{filter_str}"
    hdrs = {
        "apikey": SUPABASE_KEY,
        "Authorization": f"Bearer {SUPABASE_KEY}",
        "Content-Type": "application/json",
        "Prefer": "return=minimal",
    }
    resp = requests.patch(url, json=data, headers=hdrs)
    resp.raise_for_status()

def _hash_password_secure(password: str) -> str:
    """
    Hash password using bcrypt (SECURE)
    
    SECURITY: Never use SHA256 or other weak hashes for passwords!
    """
    if not password:
        return None
    if not bcrypt:
        logger.error("[SECURITY] bcrypt not available for password hashing")
        return None
    try:
        salt = bcrypt.gensalt(rounds=12)
        hashed = bcrypt.hashpw(password.encode('utf-8'), salt)
        return hashed.decode('utf-8')
    except Exception as e:
        logger.error(f"[SECURITY] Password hashing error: {e}")
        return None


def _verify_password_secure(password: str, password_hash: str) -> bool:
    """
    Verify password against bcrypt hash (SECURE)
    """
    if not password_hash:
        return False
    if not bcrypt:
        logger.error("[SECURITY] bcrypt not available for password verification")
        return False
    try:
        return bcrypt.checkpw(password.encode('utf-8'), password_hash.encode('utf-8'))
    except Exception as e:
        logger.error(f"[SECURITY] Password verification error: {e}")
        return False


# Paths resolved relative to project root (parent of backend/)
_PROJECT_ROOT = Path(__file__).parent.parent
# Use /tmp for encodings (writable in Cloud Run)
ENC_PATH = "/tmp/smart-ams-encodings.pkl"
ATT_CSV  = str(_PROJECT_ROOT / "attendance.csv")

SUPABASE_URL = os.getenv("SUPABASE_URL", "").rstrip('/')
SUPABASE_KEY = os.getenv("SUPABASE_KEY", "")

# Use plain HTTP requests instead of the official supabase client to avoid
# compatibility problems with httpx/httpcore on Python 3.14.  The original
# implementation imported ``from supabase import create_client`` which pulled
# in httpx and crashed at import time.  We just need a tiny subset of the
# REST API (select/insert/upsert) so a lightweight wrapper around ``requests``
# is sufficient.

import requests


def sb_select(table, filters=None, select="*"):
    """Perform a filtered SELECT against a Supabase table.

    ``filters`` should be a dict mapping column names to values; they are
    translated into ``col=eq.value`` query parameters.  ``select`` may be a
    comma-separated string of columns or ``"*"`` for all.
    """
    if not SUPABASE_URL or not SUPABASE_KEY:
        return []
    qs = []
    if filters:
        for col, val in filters.items():
            qs.append(f"{col}=eq.{val}")
    if select:
        qs.insert(0, f"select={select}")
    query = "?" + "&".join(qs) if qs else ""
    url = f"{SUPABASE_URL}/rest/v1/{table}{query}"
    hdrs = {
        "apikey": SUPABASE_KEY,
        "Authorization": f"Bearer {SUPABASE_KEY}",
    }
    try:
        resp = requests.get(url, headers=hdrs, timeout=5)
        resp.raise_for_status()
    except requests.Timeout:
        print(f"[SUPABASE] SELECT timeout on {table}")
        return []
    except Exception as e:
        print(f"[SUPABASE] SELECT error on {table}: {str(e)}")
        return []
    try:
        return resp.json()
    except ValueError:
        return []


def sb_delete(table, filters):
    """Delete rows from a Supabase table matching filters."""
    if not SUPABASE_URL or not SUPABASE_KEY:
        return []
    qs = [f"{c}=eq.{v}" for c, v in (filters or {}).items()]
    query = "?" + "&".join(qs) if qs else ""
    url = f"{SUPABASE_URL}/rest/v1/{table}{query}"
    hdrs = {
        "apikey": SUPABASE_KEY,
        "Authorization": f"Bearer {SUPABASE_KEY}",
        "Prefer": "return=representation",
    }
    try:
        resp = requests.delete(url, headers=hdrs, timeout=5)
        resp.raise_for_status()
    except requests.Timeout:
        print(f"[SUPABASE] DELETE timeout on {table}")
        return []
    except Exception as e:
        print(f"[SUPABASE] DELETE error on {table}: {str(e)}")
        return []
    try:
        return resp.json()
    except ValueError:
        return []


def sb_update(table, data, filters):
    """PATCH rows in a Supabase table matching filters."""
    if not SUPABASE_URL or not SUPABASE_KEY:
        return []
    qs = [f"{c}=eq.{v}" for c, v in (filters or {}).items()]
    query = "?" + "&".join(qs) if qs else ""
    url = f"{SUPABASE_URL}/rest/v1/{table}{query}"
    hdrs = {
        "apikey": SUPABASE_KEY,
        "Authorization": f"Bearer {SUPABASE_KEY}",
        "Content-Type": "application/json",
        "Prefer": "return=representation",
    }
    try:
        resp = requests.patch(url, json=data, headers=hdrs, timeout=5)
        resp.raise_for_status()
    except requests.Timeout:
        print(f"[SUPABASE] UPDATE timeout on {table}")
        return []
    except Exception as e:
        print(f"[SUPABASE] UPDATE error on {table}: {str(e)}")
        return []
    try:
        return resp.json()
    except ValueError:
        return []


def sb_insert(table, data, upsert=False, on_conflict=None):
    """Insert (or upsert) a row into a Supabase table.

    If ``upsert`` is True and ``on_conflict`` is provided we add the proper
    ``Prefer`` header and a query parameter to instruct the REST API to merge
    duplicates.
    """
    if not SUPABASE_URL or not SUPABASE_KEY:
        return []
    hdrs = {
        "apikey": SUPABASE_KEY,
        "Authorization": f"Bearer {SUPABASE_KEY}",
        "Content-Type": "application/json",
    }
    url = f"{SUPABASE_URL}/rest/v1/{table}"
    if upsert and on_conflict:
        hdrs["Prefer"] = "resolution=merge-duplicates"
        url = url + f"?on_conflict={on_conflict}"
    try:
        resp = requests.post(url, json=data, headers=hdrs, timeout=5)
        resp.raise_for_status()
    except requests.Timeout:
        print(f"[SUPABASE] INSERT/UPSERT timeout on {table}")
        return []
    except Exception as e:
        print(f"[SUPABASE] INSERT/UPSERT error on {table}: {str(e)}")
        return []
    try:
        return resp.json()
    except ValueError:
        return []


# ``sb`` placeholder: construct a minimal client that mimics the
# subset of the official supabase client used elsewhere in this file.
# It delegates to the helper functions above which perform REST calls
# using ``requests``.


class SimpleResult:
    def __init__(self, data):
        self.data = data


class SBTable:
    def __init__(self, table):
        self.table = table
        self._select = "*"
        self._filters = []
        self._order = None
        self._limit = None
        self._insert_data = None
        self._upsert = False
        self._on_conflict = None
        self._single = False
        self._delete = False
        self._update_data = None

    def select(self, cols="*"):
        self._select = cols
        return self

    def eq(self, col, val):
        self._filters.append((col, val))
        return self

    def order(self, column, desc=False):
        self._order = (column, desc)
        return self

    def limit(self, n):
        self._limit = n
        return self

    def single(self):
        self._single = True
        return self

    def update(self, data):
        self._update_data = data
        return self

    def delete(self):
        self._delete = True
        return self

    def insert(self, data):
        self._insert_data = data
        return self

    def upsert(self, data, on_conflict=None):
        self._insert_data = data
        self._upsert = True
        self._on_conflict = on_conflict
        return self

    def execute(self):
        filters = {c: v for c, v in self._filters} if self._filters else {}
        if self._delete:
            res = sb_delete(self.table, filters)
            return SimpleResult(res if isinstance(res, list) else [])
        elif self._update_data is not None:
            res = sb_update(self.table, self._update_data, filters)
            return SimpleResult(res if isinstance(res, list) else [])
        elif self._insert_data is not None:
            res = sb_insert(
                self.table,
                self._insert_data,
                upsert=self._upsert,
                on_conflict=self._on_conflict,
            )
            return SimpleResult(res)
        else:
            rows = sb_select(self.table, filters or None, select=self._select)
            if self._order:
                col, desc = self._order
                rows = sorted(rows, key=lambda r: r.get(col), reverse=desc)
            if self._limit is not None:
                rows = rows[: self._limit]
            if self._single:
                return SimpleResult(rows[0] if rows else None)
            return SimpleResult(rows)


class SimpleSupabaseClient:
    def table(self, name):
        return SBTable(name)


sb = SimpleSupabaseClient() if SUPABASE_URL and SUPABASE_KEY else None

# ── Firestore Helper Functions ────────────────────────────────────
# NOTE: Firebase Admin SDK is initialized earlier in the file (lines ~200-226)
# The _fstore variable is set there and used by these helper functions
def write_to_firestore(collection, doc_id, data, merge=True):
    """Write data to Firestore with error handling."""
    if not _fstore:
        return False
    try:
        _fstore.collection(collection).document(doc_id).set(data, merge=merge)
        return True
    except Exception as e:
        print(f"[FIRESTORE] Error writing to {collection}/{doc_id}: {e}")
        return False

def add_to_firestore_collection(collection, data):
    """Add data to Firestore collection (auto-generate doc ID)."""
    if not _fstore:
        return None
    try:
        ref = _fstore.collection(collection).document()
        ref.set(data)
        return ref.id
    except Exception as e:
        print(f"[FIRESTORE] Error adding to {collection}: {e}")
        return None

def delete_from_firestore(collection, doc_id):
    """Delete document from Firestore."""
    if not _fstore:
        return False
    try:
        _fstore.collection(collection).document(doc_id).delete()
        return True
    except Exception as e:
        print(f"[FIRESTORE] Error deleting {collection}/{doc_id}: {e}")
        return False

def sync_batch_to_firestore(operations):
    """Execute batch write operations to Firestore.
    
    Args:
        operations: List of tuples (action, collection, doc_id, data)
        where action is 'set', 'update', or 'delete'
    """
    if not _fstore:
        return False
    try:
        batch = _fstore.batch()
        for action, collection, doc_id, data in operations:
            ref = _fstore.collection(collection).document(doc_id)
            if action == 'set':
                batch.set(ref, data, merge=True)
            elif action == 'update':
                batch.update(ref, data)
            elif action == 'delete':
                batch.delete(ref)
        batch.commit()
        return True
    except Exception as e:
        print(f"[FIRESTORE] Error in batch operation: {e}")
        return False

# ── Original functions (unchanged) ──
def load_encodings():
    if Path(ENC_PATH).exists():
        with open(ENC_PATH,"rb") as f:
            d=pickle.load(f)
        return [np.array(e) for e in d["encodings"]], d["names"]
    return [],[]

def save_encodings(encs,names):
    with open(ENC_PATH,"wb") as f:
        pickle.dump({"encodings":[e.tolist() for e in encs],"names":names},f)

def load_encodings_supabase():
    """Return all face encodings from Supabase with names and roll_nos.

    Returns list of (encoding_array, name, roll_no) tuples.
    Falls back to local file if Supabase is unavailable or empty.
    """
    if not sb:
        # For local, return as tuples for consistency
        local_encs, local_names = load_encodings()
        return [(e, n, n) for e, n in zip(local_encs, local_names)]
    try:
        result = sb.table("face_encodings").select("encoding,name,roll_no").execute()
        data = []
        if result.data:
            for row in result.data:
                enc_text = row.get("encoding")
                name = row.get("name", "")
                roll_no = row.get("roll_no")
                try:
                    arr = np.array(json.loads(enc_text))
                    data.append((arr, name, roll_no))
                except Exception:
                    pass
        if data:
            return data
    except Exception as e:
        print(f"[FACE] Error loading encodings from Supabase: {e}")
    # Fallback to local
    local_encs, local_names = load_encodings()
    return [(e, n, n) for e, n in zip(local_encs, local_names)]

def get_system_config(key, default=None):
    """Fetch a value from system_config table."""
    if not sb:
        return default
    try:
        r = sb.table("system_config").select("value").eq("key", key).single().execute()
        if r.data:
            return r.data.get("value")
    except Exception as e:
        print(f"[CONFIG] Error reading {key}: {e}")
    return default

def is_face_enabled():
    val = get_system_config("face_recognition_enabled", "false")
    return str(val).lower() == "true"

def haversine(lat1, lon1, lat2, lon2):
    """Return distance in kilometers between two lat/lon points."""
    from math import radians, sin, cos, asin, sqrt
    lat1, lon1, lat2, lon2 = map(radians, [lat1, lon1, lat2, lon2])
    dlat = lat2 - lat1
    dlon = lon2 - lon1
    a = sin(dlat/2)**2 + cos(lat1) * cos(lat2) * sin(dlon/2)**2
    return 6371 * 2 * asin(sqrt(a))


def verify_face_for_user(user_id=None, image_b64=None, roll_no=None):
    """Verify a base64 face image against a stored encoding.

    The function prefers to look up by roll_no, but will fall back to
    user_id if the table has that column.  Returns (verified, confidence).
    """
    if not is_face_enabled() or not sb or not image_b64:
        return False, 0.0
    try:
        import face_recognition
        header, b64data = image_b64.split(",",1) if "," in image_b64 else ("", image_b64)
        bts = base64.b64decode(b64data)
        im = Image.open(BytesIO(bts)).convert("RGB")
        arr = np.array(im)
        locs = face_recognition.face_locations(arr)
        if not locs:
            return False,0.0
        encs = face_recognition.face_encodings(arr, locs)
        if not encs:
            return False,0.0
        enc = encs[0]

        # lookup stored encoding
        result = None
        try:
            if roll_no:
                result = sb.table("face_encodings").select("encoding").eq("roll_no", roll_no).execute()
        except Exception:
            result = None
        if not result or not result.data:
            try:
                result = sb.table("face_encodings").select("encoding").eq("user_id", user_id).execute()
            except Exception:
                result = None
        if not result or not result.data:
            return False,0.0
        raw_enc = result.data[0]["encoding"]
        # encoding may be stored as a JSON string or already a list
        if isinstance(raw_enc, str):
            raw_enc = json.loads(raw_enc)
        db_enc = np.array(raw_enc)
        dist = np.linalg.norm(enc - db_enc)
        verified = dist <= 0.6
        confidence = float(max(0,1 - (dist / 2.0)))
        return verified, confidence
    except Exception as e:
        print(f"[FACE VERIFY] error {e}")
        return False,0.0

def encode_image(path, model='hog'):
    """Use face_recognition library to detect faces and return encodings."""
    try:
        import face_recognition as fr
        img = np.array(Image.open(path).convert("RGB"))
        face_encs = fr.face_encodings(img)
        print(f"[FACE] Detected {len(face_encs)} face(s)")
        return face_encs
    except Exception as e:
        print(f"[FACE] Error encoding image: {e}")
        import traceback
        traceback.print_exc()
        return []

def calculate_eye_aspect_ratio(eye_points):
    """Calculate the eye aspect ratio using dlib landmarks.
    
    EAR = (||p2 - p6|| + ||p3 - p5||) / (2 * ||p1 - p4||)
    where p1...p6 are the eye landmark points.
    High EAR = eye open, Low EAR = eye closed
    """
    try:
        pts = np.array(eye_points)
        if len(pts) < 6:
            return 0
        
        # Vertical distances
        dist_top_bottom_1 = np.linalg.norm(pts[1] - pts[5])
        dist_top_bottom_2 = np.linalg.norm(pts[2] - pts[4])
        
        # Horizontal distance
        dist_left_right = np.linalg.norm(pts[0] - pts[3])
        
        # Calculate EAR
        ear = (dist_top_bottom_1 + dist_top_bottom_2) / (2.0 * dist_left_right)
        return ear
    except:
        return 0

def detect_liveness(image_path_or_array, threshold_eye_height=3, eye_aspect_ratio_threshold=0.1):
    """Detect eye movement/blinks to prevent fake image verification.
    
    Uses dlib-based eye aspect ratio (EAR) for robust liveness detection.
    Analyzes eye landmarks to ensure eyes are open and visible (not winking/closed).
    Returns True if eyes are open (live), False if eyes closed or face not detected.
    """
    try:
        import face_recognition as fr
        import cv2
        
        # Load image
        if isinstance(image_path_or_array, str):
            img = cv2.imread(image_path_or_array)
            if img is None:
                img = np.array(Image.open(image_path_or_array).convert("RGB"))
                img = cv2.cvtColor(img, cv2.COLOR_RGB2BGR)
        else:
            img = image_path_or_array
        
        if img is None:
            print("[LIVENESS] Invalid image")
            return False
        
        # Convert to RGB for face_recognition
        img_rgb = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
        
        # Try dlib-based detection first if available
        if detector is not None and sp is not None:
            print("[LIVENESS] Using dlib for eye detection...")
            dlib_img = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
            faces = detector(dlib_img, 0)
            
            if len(faces) == 0:
                print("[LIVENESS] No face detected via dlib")
                return False
            
            face = faces[0]
            landmarks = sp(dlib_img, face)
            
            # dlib landmark indices for eyes: 36-41 (left), 42-47 (right)
            left_eye_pts = [(landmarks.part(i).x, landmarks.part(i).y) for i in range(36, 42)]
            right_eye_pts = [(landmarks.part(i).x, landmarks.part(i).y) for i in range(42, 48)]
            
            # Calculate eye aspect ratios
            left_ear = calculate_eye_aspect_ratio(left_eye_pts)
            right_ear = calculate_eye_aspect_ratio(right_eye_pts)
            avg_ear = (left_ear + right_ear) / 2.0
            
            print(f"[LIVENESS] Left EAR: {left_ear:.3f}, Right EAR: {right_ear:.3f}, Avg: {avg_ear:.3f}")
            
            # Both eyes should have EAR > threshold (eyes open, not closed/winking)
            if avg_ear < eye_aspect_ratio_threshold:
                print(f"[LIVENESS] Eyes appear closed/winking (EAR: {avg_ear:.3f} < {eye_aspect_ratio_threshold})")
                return False
            
            print(f"[LIVENESS] ✓ Eyes open (EAR: {avg_ear:.3f}px)")
            return True
        
        # Fallback to face_recognition if dlib not available
        print("[LIVENESS] Using face_recognition for eye detection (fallback)...")
        face_landmarks_list = fr.face_landmarks_in_image(img_rgb)
        
        if not face_landmarks_list:
            print("[LIVENESS] No face detected")
            return False
        
        # Get eye regions from first face
        landmarks = face_landmarks_list[0]
        left_eye = landmarks.get('left_eye', [])
        right_eye = landmarks.get('right_eye', [])
        
        if not left_eye or not right_eye:
            print("[LIVENESS] Eyes not detected")
            return False
        
        # Calculate eye openness (vertical distance between eyelids)
        left_eye_pts = np.array(left_eye)
        right_eye_pts = np.array(right_eye)
        
        left_eye_height = abs(np.max(left_eye_pts[:, 1]) - np.min(left_eye_pts[:, 1]))
        right_eye_height = abs(np.max(right_eye_pts[:, 1]) - np.min(right_eye_pts[:, 1]))
        
        # Eyes should have significant vertical opening (not closed/winking)
        if left_eye_height < threshold_eye_height or right_eye_height < threshold_eye_height:
            print(f"[LIVENESS] Eyes appear closed/winking (L:{left_eye_height:.1f}, R:{right_eye_height:.1f})")
            return False
        
        print(f"[LIVENESS] ✓ Eyes open (L:{left_eye_height:.1f}px, R:{right_eye_height:.1f}px)")
        return True
    
    except Exception as e:
        print(f"[LIVENESS] Error in liveness check: {e}")
        import traceback
        traceback.print_exc()
        # If liveness check fails, allow but log warning
        return True

def mark_attendance(name):
    ts=datetime.utcnow().isoformat()
    write_header=not Path(ATT_CSV).exists()
    with open(ATT_CSV,"a",newline="") as f:
        w=csv.writer(f)
        if write_header: w.writerow(["name","timestamp"])
        w.writerow([name,ts])

# ── API ──
def decode_b64_image(b64_str):
    if "," in b64_str:
        b64_str=b64_str.split(",")[1]
    return Image.open(BytesIO(base64.b64decode(b64_str))).convert("RGB")

def upsert_face_encoding_local(enc_data, roll_no):
    """Save face encoding locally, replacing existing entry for same roll_no."""
    encs_list, names = load_encodings()
    if roll_no in names:
        idx = names.index(roll_no)
        encs_list[idx] = enc_data
        print(f"[FACE] Updated existing local encoding for {roll_no}")
    else:
        encs_list.append(enc_data)
        names.append(roll_no)
    save_encodings(encs_list, names)

def generate_roll_number(department: str, semester: int = 1) -> str:
    """Create a new roll number. Format: <year><semester><dept_3lc><seq_4digits>
    e.g. 20261cse0001.  Scans BOTH roll_no and username columns so a legacy user
    whose username equals a roll number is never re-issued.
    Continuous numbering — no duplicates within the year+sem+dept prefix.
    """
    if not sb:
        return f"{datetime.utcnow().year}{semester}{department.lower()[:3]}{str(uuid.uuid4())[:4]}"
    dept_lc = ''.join(ch for ch in department.lower() if ch.isalnum())[:3]
    prefix = f"{datetime.utcnow().year}{semester}{dept_lc}"
    try:
        import re as _re
        max_seq = 0
        for row in (sb.table("users").select("roll_no").like("roll_no", f"{prefix}%").execute().data or []):
            rn = (row.get("roll_no") or "").lower()
            if rn.startswith(prefix.lower()):
                m = _re.match(r'^(\d+)', rn[len(prefix):])
                if m:
                    max_seq = max(max_seq, int(m.group(1)))
        for row in (sb.table("users").select("username").like("username", f"{prefix}%").execute().data or []):
            un = (row.get("username") or "").lower()
            if un.startswith(prefix.lower()):
                m = _re.match(r'^(\d+)', un[len(prefix):])
                if m:
                    max_seq = max(max_seq, int(m.group(1)))
    except Exception:
        max_seq = 0
    return f"{prefix}{max_seq+1:04d}"


def upsert_face_encoding_supabase(metadata, enc_data):
    """Upsert face encoding to Supabase and sync to Firestore."""
    if not sb:
        return
    try:
        payload = {
            "name": metadata.get('name'),
            "roll_no": metadata.get('roll_no'),
            "admission_no": metadata.get('admission_no'),
            "section": metadata.get('section'),
            "academic_year": metadata.get('academic_year'),
            "encoding": json.dumps(enc_data.tolist()),
            "created_at": datetime.utcnow().isoformat()
        }
        try:
            print(f"[SUPABASE] Upserting face_encodings payload: {json.dumps(payload)[:1000]}")
            sb.table("face_encodings").upsert(payload, on_conflict="roll_no").execute()
            print(f"[FACE] Encoding upserted to Supabase for {metadata.get('roll_no')}")
        except Exception as e:
            print(f"[SUPABASE] Upsert face_encodings error: {e}")
            try:
                import traceback
                traceback.print_exc()
            except Exception:
                pass
        # Sync to Firestore (store summary — no raw encoding for read perf)
        roll = metadata.get('roll_no') or metadata.get('admission_no') or 'unknown'
        fstore_set("face_registrations", roll, {
            "roll_no":       roll,
            "name":          metadata.get('name'),
            "section":       metadata.get('section'),
            "academic_year": metadata.get('academic_year'),
            "registered_at": datetime.utcnow().isoformat(),
            "source":        "mass_upload",
        })
    except Exception as e:
        print(f"[FACE] Warning: Failed to upsert to Supabase: {e}")

@app.route("/health")
def health():
    return jsonify(status="ok",supabase=sb is not None,time=datetime.utcnow().isoformat())

@app.route("/api/test-create-student", methods=["POST"])
def test_create_student():
    """Debug endpoint: Create a test student account"""
    try:
        import hashlib
        
        test_data = {
            "username": "student001",
            "password_hash": _hash_password_secure("password123"),
            "role": "student",
            "full_name": "Test Student",
            "email": "student@test.com",
            "roll_no": "student001",
            "department": "CSE",
            "section": "A",
            "created_at": datetime.utcnow().isoformat()
        }
        
        if sb:
            result = sb.table("users").insert(test_data).execute()
            return jsonify(success=True, message="Test student created", user=test_data)
        else:
            return jsonify(success=False, error="Supabase not configured"), 500
            
    except Exception as e:
        error_msg = str(e)
        if "duplicate" in error_msg.lower():
            return jsonify(success=True, message="Test student already exists", credentials={
                "username": "student001",
                "password": "password123"
            })
        return jsonify(success=False, error=error_msg), 500

@app.route("/api/register", methods=["POST"])
def register():
    """Register face only (standalone face registration). Handles both multipart and JSON."""
    try:
        img = None
        metadata = {}
        
        # Check if it's FormData with file (multipart/form-data)
        if request.files.get('image'):
            image_file = request.files['image']
            print(f"[FACE] Received file: {image_file.filename}")
            img = Image.open(image_file.stream).convert("RGB")
            metadata = {
                'name': request.form.get('name', 'unknown'),
                'roll_no': request.form.get('roll_no', '').strip() or None,
                'admission_no': request.form.get('admission_no', '').strip() or str(uuid.uuid4()),
                'section': request.form.get('section', '–'),
                'academic_year': request.form.get('academic_year', '2024-25')
            }
            # if roll is empty and department info present, generate one
            if not metadata['roll_no'] and request.form.get('department'):
                metadata['roll_no'] = generate_roll_number(request.form.get('department'))
        
        # Otherwise, handle JSON with base64 image
        elif request.is_json:
            d = request.json or {}
            if not d.get("image"):
                print("[FACE] No image in JSON request")
                return jsonify(error="No image provided"), 400
            img = decode_b64_image(d["image"])
            metadata = {
                'name': d.get("name", 'unknown'),
                'roll_no': d.get("roll_no", '').strip() or None,
                'admission_no': d.get("admission_no", '').strip() or str(uuid.uuid4()),
                'section': d.get("section", '–'),
                'academic_year': d.get("academic_year", '2024-25')
            }
            if not metadata['roll_no'] and d.get('department'):
                metadata['roll_no'] = generate_roll_number(d.get('department'))
        else:
            print("[FACE] No image provided (neither multipart nor JSON)")
            return jsonify(error="No image provided. Send multipart form with 'image' file or JSON with 'image' (base64)"), 400
        
        # Process image
        tmp = "tmp_reg.jpg"
        img.save(tmp)
        print(f"[FACE] Image saved, detecting faces...")
        encs = encode_image(tmp)
        print(f"[FACE] Faces detected: {len(encs)}")
        os.remove(tmp)
        
        if len(encs) != 1:
            error_msg = f"Found {len(encs)} faces. Ensure exactly ONE face is visible"
            print(f"[FACE] Error: {error_msg}")
            return jsonify(error=error_msg), 400
        
        enc_data = encs[0]
        roll_no = metadata.get('roll_no', 'unknown')

        # Save locally (upsert — replace if same roll_no exists)
        upsert_face_encoding_local(enc_data, roll_no)
        print(f"[FACE] Encoding saved locally")

        # Upsert to Supabase
        upsert_face_encoding_supabase(metadata, enc_data)
        
        print(f"[FACE] Registration successful for {metadata.get('name')}")
        return jsonify(success=True, message=f"Registered {metadata.get('name')}", encoding=enc_data.tolist())
    
    except Exception as e:
        print(f"[FACE] Exception: {e}")
        import traceback
        traceback.print_exc()
        return jsonify(error=f"Server error: {str(e)}"), 500


# ── BULK / MASS FACE REGISTRATION ────────────────────────────────────────────
@app.route("/api/face/bulk-register", methods=["POST"])
def bulk_register_faces():
    """Mass face registration: upload multiple images at once.

    Rules:
    - If a filename (without extension) matches the roll-number pattern
      (digits + 2-4 letters + digits, e.g. 20261CSE0001) it is used as-is.
    - Otherwise a roll number is auto-generated sequentially for the batch.
    - If an existing student in the `users` table has that roll_no the face
      encoding is linked to their account (synced for login verification).
    """
    images = request.files.getlist("images")
    if not images:
        return jsonify(success=False, error="No images provided"), 400

    department  = request.form.get("department", "").strip()
    year        = request.form.get("year", str(datetime.utcnow().year)).strip()
    semester    = request.form.get("semester", "1").strip()
    academic_year = request.form.get("academic_year", f"{datetime.utcnow().year}-{str(datetime.utcnow().year+1)[2:]}").strip()

    dept_clean  = ''.join(ch for ch in department.lower() if ch.isalnum())
    roll_prefix = f"{year}{semester}{dept_clean}"

    # Seed the auto-sequence counter from the HIGHEST existing sequence for this prefix
    # Using max (not count) so deleted rows never cause duplicates.
    auto_seq = 0
    if sb and dept_clean:
        try:
            existing = sb.table("users").select("roll_no").like("roll_no", f"{roll_prefix}%").execute()
            import re as _re
            for row in (existing.data or []):
                rn = (row.get("roll_no") or "").lower()
                if rn.startswith(roll_prefix.lower()):
                    tail = rn[len(roll_prefix):]
                    m = _re.match(r'^(\d+)', tail)
                    if m:
                        seq = int(m.group(1))
                        if seq > auto_seq:
                            auto_seq = seq
        except Exception:
            pass

    import re
    results = []

    for img_file in images:
        filename  = img_file.filename or "unknown"
        basename  = os.path.splitext(filename)[0].strip()

        # Determine roll_no
        if re.match(r'^[0-9]{4,}[a-zA-Z]{2,4}[0-9]+$', basename):
            roll_no     = basename.upper()
            roll_source = "filename"
        else:
            auto_seq += 1
            roll_no     = f"{roll_prefix}{auto_seq:04d}".upper()
            roll_source = "auto"

        try:
            img = Image.open(img_file.stream).convert("RGB")
            tmp = f"tmp_bulk_{roll_no}.jpg"
            img.save(tmp)
            encs = encode_image(tmp)
            os.remove(tmp)

            if len(encs) != 1:
                results.append({"filename": filename, "roll_no": roll_no,
                                 "status": "failed",
                                 "error": f"{len(encs)} faces detected – need exactly 1"})
                continue

            enc_data = encs[0]

            # Check if a student with this roll_no exists
            matched_user = None
            if sb:
                try:
                    uq = sb.table("users").select("id,full_name,roll_no").eq("roll_no", roll_no).execute()
                    if uq.data:
                        matched_user = uq.data[0]
                except Exception:
                    pass

            student_name = matched_user["full_name"] if matched_user else basename

            # Save locally
            upsert_face_encoding_local(enc_data, roll_no)

            # Save to Supabase face_encodings (upsert on roll_no conflict)
            metadata = {
                "name":          student_name,
                "roll_no":       roll_no,
                "admission_no":  str(uuid.uuid4()),
                "section":       "–",
                "academic_year": academic_year,
            }
            upsert_face_encoding_supabase(metadata, enc_data)

            # Mark the user as having a registered face
            if sb and matched_user:
                try:
                    sb.table("users").update({"face_registered": True}).eq("id", matched_user["id"]).execute()
                except Exception:
                    pass

            results.append({
                "filename":       filename,
                "roll_no":        roll_no,
                "roll_source":    roll_source,
                "status":         "success",
                "matched":        matched_user is not None,
                "matched_student": matched_user["full_name"] if matched_user else None,
            })

        except Exception as e:
            import traceback; traceback.print_exc()
            results.append({"filename": filename, "roll_no": roll_no,
                             "status": "failed", "error": str(e)})

    registered = sum(1 for r in results if r["status"] == "success")
    matched    = sum(1 for r in results if r.get("matched"))
    failed     = sum(1 for r in results if r["status"] == "failed")

    return jsonify(success=True, total=len(results),
                   registered=registered, matched=matched, failed=failed,
                   results=results)


@app.route("/api/register-and-add-user", methods=["POST"])
def register_and_add_user():
    """Combined endpoint: register face + create user account in one call."""
    try:
        if not request.files.get('image'):
            return jsonify(success=False, error="No face image provided"), 400

        image_file = request.files['image']
        role        = request.form.get('role', 'student')
        full_name   = request.form.get('full_name', '').strip()
        username    = request.form.get('username', '').strip().upper()  # Convert to uppercase
        email       = request.form.get('email', '').strip()
        password    = request.form.get('password', '')
        department  = request.form.get('department', '').strip()
        roll_no     = request.form.get('roll_no', '').strip().upper()  # Convert to uppercase
        section     = request.form.get('section', '').strip()
        academic_year = request.form.get('academic_year', '2024-25')

        if role == 'student' and not roll_no:
            # generate roll for students automatically if omitted
            roll_no = generate_roll_number(department).upper()  # Convert to uppercase
            username = roll_no
        # admission number always a uuid
        admission_no = str(uuid.uuid4())

        if not all([full_name, username, email, password, department]):
            return jsonify(success=False, error="Missing required fields"), 400

        # ── Step 1: Process face image ──
        img = Image.open(image_file.stream).convert("RGB")
        tmp = "tmp_combined_reg.jpg"
        img.save(tmp)
        print(f"[COMBINED] Detecting faces for {full_name}...")
        encs = encode_image(tmp)
        os.remove(tmp)
        print(f"[COMBINED] Faces detected: {len(encs)}")

        if len(encs) != 1:
            return jsonify(success=False, error=f"Found {len(encs)} faces. Need exactly 1 clear face."), 400

        enc_data = encs[0]
        effective_roll = roll_no or username

        # ── Step 2: Save face locally (upsert) ──
        upsert_face_encoding_local(enc_data, effective_roll)
        print(f"[COMBINED] Face encoding saved locally for {effective_roll}")

        # ── Step 3: Upsert face to Supabase ──
        upsert_face_encoding_supabase({
            'name': full_name,
            'roll_no': effective_roll,
            'admission_no': effective_roll,
            'section': section or '–',
            'academic_year': academic_year
        }, enc_data)

        # ── Step 4: Create user account in Supabase ──
        if not sb:
            return jsonify(success=False, error="Supabase not configured"), 500

        # SECURITY: Use bcrypt for password hashing
        pwd_hash = _hash_password_secure(password)

        # Check duplicate username
        existing_user = sb.table("users").select("id").eq("username", username).execute()
        if existing_user.data:
            return jsonify(success=False, error="Username already taken"), 400

        # Check duplicate roll_no for students
        if roll_no and role == "student":
            existing_roll = sb.table("users").select("id").eq("roll_no", roll_no).execute()
            if existing_roll.data:
                return jsonify(success=False, error=f"Roll number {roll_no} is already registered"), 400

        user_payload = {
            "username":      username,
            "password_hash": pwd_hash,
            "role":          role,
            "full_name":     full_name,
            "email":         email,
            "roll_no":       roll_no if role == "student" else None,
            "department":    department,
            "section":       section if role == "student" else None,
            "created_at":    datetime.utcnow().isoformat()
        }
        try:
            print(f"[SUPABASE] Inserting user payload: {json.dumps(user_payload)[:1000]}")
            result = sb.table("users").insert(user_payload).execute()
            user_id = result.data[0]["id"] if result.data else None
            print(f"[COMBINED] User account created: {username} (id={user_id})")
            
            # Sync user to Firestore with matching structure
            if _fstore and user_id:
                firestore_data = {
                    "id": user_id,
                    "username": username,
                    "role": role,
                    "full_name": full_name,
                    "email": email,
                    "roll_no": roll_no if role == "student" else None,
                    "department": department,
                    "section": section if role == "student" else None,
                    "created_at": datetime.utcnow().isoformat(),
                    "synced_at": datetime.utcnow().isoformat()
                }
                try:
                    write_to_firestore("users", str(user_id), firestore_data)
                    print(f"[FIRESTORE] User synced: {username} (id={user_id})")
                except Exception as fs_err:
                    print(f"[FIRESTORE] Warning: Could not sync user: {fs_err}")
            
        except Exception as e:
            print(f"[SUPABASE] Error inserting user: {e}")
            try:
                import traceback
                traceback.print_exc()
            except Exception:
                pass
            return jsonify(success=False, error=f"Supabase insert error: {str(e)}"), 500

        return jsonify(success=True, message=f"User {full_name} registered with face!", user_id=user_id)

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify(success=False, error=str(e)), 500


@app.route("/api/verify", methods=["POST"])
def verify():
    try:
        import face_recognition as fr
        
        d = request.json or {}
        if not d or not d.get("image"):
            return jsonify(verified=False, error="No image provided"), 400
        
        # Get logged-in student's roll number
        logged_in_roll_no = d.get("roll_no")
        session_id = d.get("session_id")
        
        if not logged_in_roll_no:
            return jsonify(verified=False, error="Student identity not provided"), 400
        
        # Get max attempts from config
        max_attempts = 2
        if sb:
            try:
                r = sb.table("system_config").select("value").eq("key","max_face_attempts").execute()
                if r.data and len(r.data) > 0:
                    max_attempts = int(r.data[0].get("value", "2"))
            except:
                max_attempts = 2
        
        # Check current attempt count
        current_attempt = 1
        if sb:
            try:
                attempts_result = sb.table("verification_attempts").select("COUNT(*)").eq("roll_no", logged_in_roll_no).eq("session_id", session_id).execute()
                # This is a workaround since COUNT isn't standard in our client
                attempt_records = sb.table("verification_attempts").select("*").eq("roll_no", logged_in_roll_no).eq("session_id", session_id).execute()
                if attempt_records.data:
                    current_attempt = len(attempt_records.data) + 1
            except:
                current_attempt = 1
        
        # Check if attempts exceeded
        if current_attempt > max_attempts:
            print(f"[VERIFY] ❌ ATTEMPTS EXCEEDED: Student {logged_in_roll_no} has exceeded {max_attempts} attempts")
            return jsonify(
                verified=False, 
                error=f"Maximum verification attempts ({max_attempts}) exceeded. Please contact SmartAMS Admin for attendance.", 
                attempts_exhausted=True,
                current_attempt=current_attempt,
                max_attempts=max_attempts
            ), 403
        
        # check feature toggle
        def is_face_enabled():
            if not sb: return False
            try:
                r = sb.table("system_config").select("value").eq("key","face_recognition_enabled").execute()
                if r.data and len(r.data) > 0:
                    return r.data[0].get("value","false") == "true"
                return False
            except Exception as e:
                print(f"[VERIFY] Error checking face_rec_enabled: {e}")
                return False
        
        if not is_face_enabled():
            print(f"[VERIFY] Face recognition disabled, rejecting verification")
            return jsonify(verified=False, error="Face recognition disabled"), 403
        
        # Use stricter tolerance (0.45) for better accuracy - rejects similar but different faces
        tol = float(d.get("tolerance",0.45))
        latitude = d.get("latitude")
        longitude = d.get("longitude")
        
        # location verification if coords provided
        in_campus = None
        if latitude is not None and longitude is not None:
            def haversine(lat1, lon1, lat2, lon2):
                # return kilometers
                from math import radians, sin, cos, asin, sqrt
                lat1, lon1, lat2, lon2 = map(radians, [lat1, lon1, lat2, lon2])
                dlat = lat2 - lat1
                dlon = lon2 - lon1
                a = sin(dlat/2)**2 + cos(lat1) * cos(lat2) * sin(dlon/2)**2
                return 6371 * 2 * asin(sqrt(a))
            # fetch college coords
            try:
                clat = float(sb.table("system_config").select("value").eq("key","college_lat").single().execute().data.get("value",0))
                clng = float(sb.table("system_config").select("value").eq("key","college_lng").single().execute().data.get("value",0))
                crad = float(sb.table("system_config").select("value").eq("key","college_radius_km").single().execute().data.get("value",0.0))
                in_campus = haversine(latitude, longitude, clat, clng) <= crad
            except Exception:
                in_campus = False
        
        img = decode_b64_image(d.get("image"))
        tmp = "tmp_ver.jpg"
        img.save(tmp)
        
        # Check liveness (eye movement/detection)
        is_live = detect_liveness(tmp)
        if not is_live:
            os.remove(tmp)
            return jsonify(verified=False, error="Liveness check failed - eyes must be open. Possible fake/static image detected.", face_count=0, location_verified=in_campus)
        
        face_encs = encode_image(tmp)
        os.remove(tmp)
        
        # Check if face was detected
        if not face_encs:
            return jsonify(verified=False, error="No face detected. Face is not visible.", face_count=0, location_verified=in_campus)
        
        # Check if multiple faces were detected
        if len(face_encs) > 1:
            return jsonify(verified=False, error=f"More than one person detected ({len(face_encs)} faces found). Please ensure only one person is visible.", face_count=len(face_encs), location_verified=in_campus)
        
        # Load registered encodings (prefer Supabase)
        # Returns list of (encoding_array, name, roll_no) tuples
        encoding_data = load_encodings_supabase() if sb else load_encodings()
        if not encoding_data:
            return jsonify(verified=False, error="No registered users", face_count=1, location_verified=in_campus)
        
        # Extract encodings for distance calculation
        encs_array = np.array([e[0] for e in encoding_data])
        
        # Compute distances using face_recognition library
        current_encoding = face_encs[0]
        distances = fr.face_distance(encs_array, current_encoding)
        
        idx = np.argmin(distances)
        min_distance = distances[idx]
        
        # Get matched student info
        matched_encoding, student_name, student_roll = encoding_data[idx]
        
        print(f"[VERIFY] Min distance: {min_distance:.4f}, Tolerance: {tol}, Face count: {len(face_encs)}")
        
        # common attendance payload
        attendance_record = {
            "name": None,
            "roll_no": None,
            "date": datetime.utcnow().date().isoformat(),
            "timestamp": datetime.utcnow().isoformat(),
            "method": "face",
            "latitude": latitude,
            "longitude": longitude,
            "in_campus": in_campus
        }
        
        confidence = max(0, 1 - (min_distance / 2.0))
        
        # Record attempt in database
        if sb:
            try:
                sb.table("verification_attempts").insert({
                    "roll_no": logged_in_roll_no,
                    "attempt_number": current_attempt,
                    "session_id": session_id,
                    "matched_roll_no": student_roll,
                    "distance": float(min_distance),
                    "verified": False
                }).execute()
            except Exception as e:
                print(f"[VERIFY] Failed to record attempt: {e}")
        
        # Check TWO conditions for successful verification:
        # 1. Face distance is within tolerance
        # 2. Logged-in student's roll_no matches the matched face's roll_no
        identity_match = (logged_in_roll_no == student_roll)
        face_match = (min_distance <= tol)
        
        if face_match and identity_match:
            # SUCCESS: Face matches AND student identity matches
            attendance_record.update({"name": student_name, "roll_no": student_roll, "verified": True})
            print(f"[VERIFY] ✅ VERIFIED: {student_name} ({student_roll}) - Face match + Identity match - distance={min_distance:.4f}, confidence={confidence:.2%}, attempt {current_attempt}/{max_attempts}")
            
            # Update attempt record as verified
            if sb:
                try:
                    sb.table("verification_attempts").update({"verified": True}).eq("roll_no", logged_in_roll_no).eq("attempt_number", current_attempt).eq("session_id", session_id).execute()
                except:
                    pass
            
            # Mark attendance locally
            mark_attendance(student_name)
            
            # Also save to Supabase if available
            if sb:
                try:
                    sb.table("attendance").insert(attendance_record).execute()
                    print(f"[VERIFY] ✅ Attendance saved to Supabase for {student_name}")
                except Exception as e:
                    print(f"[VERIFY] ❌ Supabase insert error: {e}")
            
            return jsonify(
                verified=True, 
                name=student_name, 
                roll_no=student_roll, 
                confidence=float(confidence), 
                face_count=1, 
                location_verified=in_campus,
                current_attempt=current_attempt,
                max_attempts=max_attempts
            )
        else:
            # FAILED: Either face doesn't match OR student identity doesn't match
            failures = []
            if not face_match:
                failures.append("face does not match registered users")
            if not identity_match:
                failures.append(f"matched student ({student_roll}) is not the logged-in student ({logged_in_roll_no})")
            
            error_msg = "Verification failed: " + " AND ".join(failures)
            attendance_record.update({"name": None, "roll_no": logged_in_roll_no, "verified": False})
            print(f"[VERIFY] ❌ FAILED {'(Attempt ' + str(current_attempt) + '/' + str(max_attempts) + ')'}: {error_msg}")
            
            # Calculate remaining attempts
            remaining_attempts = max_attempts - current_attempt
            
            if remaining_attempts > 0:
                error_display = error_msg + f". You have {remaining_attempts} attempt{'s' if remaining_attempts != 1 else ''} remaining."
            else:
                error_display = f"All {max_attempts} verification attempts used. Please contact SmartAMS Admin for attendance."
            
            # Only record absence if this is the final failed attempt
            if current_attempt >= max_attempts and sb:
                try:
                    sb.table("attendance").insert(attendance_record).execute()
                    print(f"[VERIFY] ✅ Absence recorded to Supabase (attempts exhausted)")
                except Exception as e:
                    print(f"[VERIFY] ❌ Supabase insert error: {e}")
            
            return jsonify(
                verified=False, 
                error=error_display,
                confidence=float(confidence), 
                face_count=1, 
                location_verified=in_campus,
                current_attempt=current_attempt,
                max_attempts=max_attempts,
                attempts_remaining=remaining_attempts
            )
    
    except Exception as e:
        print(f"[VERIFY] Error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify(verified=False, error=f"Server error: {str(e)}"), 500
@app.route("/api/attendance", methods=["GET"])
def get_attendance():
    if sb:
        q=sb.table("attendance").select("*").order("timestamp",desc=True).limit(500)
        if request.args.get("date"):       q=q.eq("date", request.args["date"])
        if request.args.get("roll_no"):    q=q.eq("roll_no", request.args["roll_no"])
        if request.args.get("batch"):      q=q.eq("batch", request.args["batch"])
        if request.args.get("subject"):    q=q.ilike("subject_name", f"%{request.args['subject']}%")
        if request.args.get("session_id"): q=q.eq("session_id", request.args["session_id"])
        if request.args.get("faculty_id"): q=q.eq("faculty_id", request.args["faculty_id"])
        return jsonify(records=q.execute().data)
    rows=[]
    if Path(ATT_CSV).exists():
        with open(ATT_CSV) as f:
            rows=list(csv.DictReader(f))
    return jsonify(records=rows)

@app.route("/api/config/face-recognition", methods=["GET","POST"])
def face_rec_config():
    if request.method=="POST":
        enabled=request.json.get("enabled",False)
        if sb:
            try:
                # Use upsert with on_conflict on "key" column to handle both insert and update
                result = sb.table("system_config").upsert(
                    {"key":"face_recognition_enabled","value":str(enabled).lower()},
                    on_conflict="key"
                ).execute()
                print(f"[FACE REC] Saved face_recognition_enabled to {enabled}")
            except Exception as e:
                print(f"[FACE REC] Error setting config: {e}")
                return jsonify(success=False, error=str(e)), 500
        return jsonify(success=True, enabled=enabled)
    
    # GET request - retrieve current value
    if sb:
        try:
            result = sb.table("system_config").select("value").eq("key","face_recognition_enabled").execute()
            if result.data and len(result.data) > 0:
                enabled = result.data[0].get("value","false") == "true"
                return jsonify(enabled=enabled)
            else:
                # Row doesn't exist, default to false
                return jsonify(enabled=False)
        except Exception as e:
            print(f"[FACE REC] Error getting config: {e}")
            return jsonify(enabled=False)
    
    return jsonify(enabled=False)


@app.route("/api/system-config", methods=["GET","POST"])
def system_config_api():
    """Get or set system-wide configuration values used by the frontend.

    GET returns a JSON object with keys: college_lat, college_lng, college_radius_km,
    tolerance, qr_expiry_minutes, attendance_window_end (values may be strings).

    POST accepts a JSON payload with any of those keys and upserts them into
    `system_config` table as key/value strings.
    """
    if not sb:
        # Return sensible defaults when Supabase isn't configured
        if request.method == 'GET':
            return jsonify({
                "college_lat": 13.145615,
                "college_lng": 77.574597,
                "college_radius_km": 0.2,
                "tolerance": "0.5",
                "qr_expiry_minutes": 5,
                "attendance_window_end": "18:00"
            })
        return jsonify(success=False, error="Supabase not configured"), 500

    if request.method == 'GET':
        keys = ["college_lat", "college_lng", "college_radius_km", "tolerance", "qr_expiry_minutes", "attendance_window_end"]
        out = {}
        try:
            for k in keys:
                try:
                    r = sb.table("system_config").select("value").eq("key", k).single().execute()
                    if r.data and r.data.get("value") is not None:
                        v = r.data.get("value")
                        # Attempt numeric conversion where appropriate
                        if k in ("college_lat", "college_lng", "college_radius_km"):
                            try:
                                out[k] = float(v)
                            except Exception:
                                out[k] = v
                        elif k == "qr_expiry_minutes":
                            try:
                                out[k] = int(v)
                            except Exception:
                                out[k] = v
                        else:
                            out[k] = v
                except Exception:
                    continue
        except Exception as e:
            print(f"[CONFIG-GET] Error: {e}")
        return jsonify(out)

    # POST: upsert provided keys
    try:
        d = request.json or {}
        for k, v in d.items():
            try:
                sb.table("system_config").upsert({"key": k, "value": str(v)} , on_conflict="key").execute()
            except Exception as e:
                print(f"[CONFIG-POST] failed to upsert {k}: {e}")
        return jsonify(success=True)
    except Exception as e:
        print(f"[CONFIG-POST] Error: {e}")
        return jsonify(success=False, error=str(e)), 500

@app.route("/api/registered-students", methods=["GET"])
def get_registered_students():
    """Get active registered students from users table.
    Supports ?batch=, ?department=, ?section= filters.
    Deduplicates by roll_no (keeps freshest record).
    """
    try:
        print("[API] GET /api/registered-students")
        if sb:
            try:
                q = sb.table("users") \
                    .select("id,full_name,username,roll_no,email,department,section,program,last_login,is_active,created_at") \
                    .eq("role", "student") \
                    .eq("is_active", True)

                if request.args.get("batch"):
                    q = q.eq("section", request.args["batch"])
                if request.args.get("department"):
                    q = q.eq("department", request.args["department"])
                if request.args.get("section"):
                    q = q.eq("section", request.args["section"])

                rows = q.order("created_at", desc=True).execute().data or []

                # Deduplicate by roll_no (keep first = most recent created_at)
                seen_rolls = set()
                students = []
                for r in rows:
                    rn = (r.get("roll_no") or "").strip()
                    if not rn:
                        continue  # skip students with no roll number
                    name = (r.get("full_name") or r.get("username") or "").strip()
                    if not name or name.lower() == "unknown" or name.lower() == rn.lower():
                        # Use username as fallback but don't skip
                        name = r.get("username") or rn
                    if rn in seen_rolls:
                        continue
                    seen_rolls.add(rn)
                    students.append({
                        "id":         r.get("id"),
                        "name":       name,
                        "roll_no":    rn,
                        "email":      r.get("email") or "",
                        "section":    r.get("section") or "",
                        "department": r.get("department") or "",
                        "program":    r.get("program") or "",
                        "last_login": r.get("last_login") or "",
                        "created_at": r.get("created_at") or "",
                    })

                students.sort(key=lambda x: x["roll_no"])
                print(f"[STUDENTS] Returning {len(students)} unique active students")
                return jsonify(success=True, students=students)
            except Exception as e:
                print(f"[STUDENTS] Supabase error: {e}")
                raise
        else:
            encs, names = load_encodings()
            students = [{"name": n, "roll_no": n, "section": "A", "status": "registered"} for n in names]
            return jsonify(success=True, students=students)
    except Exception as e:
        print(f"[STUDENTS] Error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify(success=False, error=str(e), students=[]), 500

@app.route("/api/mark-qr-attendance", methods=["POST"])
def mark_qr_attendance():
    """Mark attendance via QR code scan."""
    d=request.json
    roll_no=d.get("roll_no")
    name=d.get("name")
    course=d.get("course")
    session_id=d.get("session_id")
    
    if not all([roll_no,name,course]):
        return jsonify(verified=False,error="Missing required fields"),400
    
    face_img_b64=d.get("face_image","")
    if not face_img_b64:
        if sb:
            sb.table("attendance").insert({
                "name":name,"roll_no":roll_no,"course":course,
                "date":datetime.utcnow().date().isoformat(),
                "timestamp":datetime.utcnow().isoformat(),
                "verified":False,"method":"qr","session_id":session_id
            }).execute()
        return jsonify(verified=False,message="Marked without face verification")
    
    img=decode_b64_image(face_img_b64)
    tmp="tmp_qr_ver.jpg"; img.save(tmp)
    face_encs=encode_image(tmp); os.remove(tmp)
    
    if not face_encs:
        return jsonify(verified=False,error="No face detected")
    
    encs,names=load_encodings()
    if not encs:
        return jsonify(verified=False,error="No registered faces")
    
    current_encoding = face_encs[0]
    distances = [np.linalg.norm(current_encoding - e) for e in encs]
    distances = np.array(distances)
    idx = np.argmin(distances)
    min_distance = distances[idx]
    verified = min_distance <= 0.6
    confidence = float(max(0, 1 - (min_distance / 2.0)))
    
    if sb:
        sb.table("attendance").insert({
            "name":name,"roll_no":roll_no,"course":course,
            "date":datetime.utcnow().date().isoformat(),
            "timestamp":datetime.utcnow().isoformat(),
            "verified":verified,"method":"qr","session_id":session_id,
            "confidence":confidence
        }).execute()
    
    return jsonify(verified=verified,name=names[idx] if verified else name,confidence=confidence)

# ── MOBILE QR ATTENDANCE ENDPOINTS ──
# Simple URL-based QR attendance for phone scanning

@app.route("/api/qr/mobile-session", methods=["POST"])
def create_mobile_qr_session():
    """Faculty: Create a mobile-friendly QR attendance session with URL"""
    try:
        d = request.json
        faculty_id = d.get("faculty_id")
        course_id = d.get("course_id", "")
        subject = d.get("subject", "Class")
        validity_minutes = int(d.get("validity_minutes", 5))
        
        if not faculty_id:
            return jsonify(success=False, error="Missing faculty_id"), 400
        
        # Generate unique session ID
        session_id = str(uuid.uuid4())[:8].upper()
        expires_at = (datetime.utcnow() + timedelta(minutes=validity_minutes)).isoformat()
        
        # Store session
        if sb:
            try:
                sb.table("qr_sessions").insert({
                    "session_id": session_id,
                    "faculty_id": faculty_id,
                    "course_id": course_id,
                    "subject": subject,
                    "expires_at": expires_at,
                    "created_at": datetime.utcnow().isoformat(),
                    "active": True,
                    "attendance_count": 0
                }).execute()
            except Exception as e:
                print(f"[Mobile-QR] DB Error: {e}")
        
        # Generate QR code with URL
        qr_url = f"{request.host_url.rstrip('/')}/attendance/mark?session={session_id}"
        qr = qrcode.QRCode(version=1, box_size=10, border=5)
        qr.add_data(qr_url)
        qr.make(fit=True)
        
        img = qr.make_image(fill_color="black", back_color="white")
        img_byte = BytesIO()
        img.save(img_byte, format='PNG')
        qr_base64 = base64.b64encode(img_byte.getvalue()).decode()
        
        return jsonify(
            success=True,
            session_id=session_id,
            qr_code_base64=qr_base64,
            qr_url=qr_url,
            expires_at=expires_at,
            validity_minutes=validity_minutes
        )
    except Exception as e:
        print(f"[Mobile-QR] Error: {e}")
        return jsonify(success=False, error=str(e)), 500

@app.route("/api/attendance/validate-session", methods=["POST"])
def validate_attendance_session():
    """Validate QR session from mobile attendance page"""
    try:
        d = request.json
        session_id = d.get("session_id")
        
        if not session_id or not sb:
            return jsonify(success=False, error="Invalid session"), 400
        
        # Check session exists and is active
        result = sb.table("qr_sessions").select("*").eq("session_id", session_id).execute()
        
        if not result.data:
            return jsonify(success=False, error="Session not found or expired"), 404
        
        session = result.data[0]
        
        # Check if expired
        expires = datetime.fromisoformat(session["expires_at"])
        if datetime.utcnow() > expires:
            return jsonify(success=False, error="QR session expired"), 410
        
        return jsonify(
            success=True,
            subject=session.get("subject", "Class"),
            faculty_id=session.get("faculty_id"),
            validity_minutes=session.get("validity_minutes", 5)
        )
    except Exception as e:
        print(f"[Validate-Session] Error: {e}")
        return jsonify(success=False, error=str(e)), 500

@app.route("/api/attendance/mark-qr", methods=["POST"])
def mark_qr_attendance_mobile():
    """Mark attendance from mobile attendance page with location and face"""
    try:
        d = request.json
        session_id = d.get("session_id")
        roll_no = d.get("roll_no", "").strip().upper()
        name = d.get("name", "").strip()
        department = d.get("department", "")
        latitude = d.get("latitude")
        longitude = d.get("longitude")
        face_image_b64 = d.get("face_image", "")
        
        if not all([session_id, roll_no, name, latitude, longitude, face_image_b64]):
            return jsonify(success=False, error="Missing required fields"), 400
        
        if not sb:
            return jsonify(success=False, error="Database not configured"), 500
        
        # Validate session
        session_result = sb.table("qr_sessions").select("*").eq("session_id", session_id).execute()
        if not session_result.data:
            return jsonify(success=False, error="Session not found"), 404
        
        session_data = session_result.data[0]
        expires = datetime.fromisoformat(session_data["expires_at"])
        if datetime.utcnow() > expires:
            return jsonify(success=False, error="Session expired"), 410
        
        # Process face recognition
        verified = False
        confidence = 0
        matched_name = name
        
        try:
            img = decode_b64_image(face_image_b64)
            tmp_file = f"tmp_face_{uuid.uuid4().hex[:8]}.jpg"
            img.save(tmp_file)
            
            face_encs = encode_image(tmp_file)
            os.remove(tmp_file)
            
            if face_encs:
                encs, names_list = load_encodings()
                if encs:
                    current_encoding = face_encs[0]
                    distances = [np.linalg.norm(current_encoding - e) for e in encs]
                    distances = np.array(distances)
                    idx = np.argmin(distances)
                    min_distance = distances[idx]
                    verified = min_distance <= 0.6
                    confidence = float(max(0, 1 - (min_distance / 2.0)))
                    if verified:
                        matched_name = names_list[idx]
        except Exception as e:
            print(f"[Face-Recognition] Error: {e}")
        
        # Check location (simple distance check from college)
        college_lat, college_lng = 13.145615, 77.574597  # Example coordinates
        location_verified = False
        if latitude is not None and longitude is not None:
            # Calculate distance
            distance_km = calculate_distance(latitude, longitude, college_lat, college_lng)
            location_verified = distance_km <= 1.0  # Within 1km
        
        # Determine final verified status (all conditions must pass)
        final_verified = verified and location_verified
        
        # Store attendance record
        attendance_record = {
            "session_id": session_id,
            "roll_no": roll_no,
            "name": matched_name if verified else name,
            "department": department,
            "latitude": latitude,
            "longitude": longitude,
            "face_verified": verified,
            "location_verified": location_verified,
            "confidence": confidence,
            "verified": final_verified,
            "timestamp": datetime.utcnow().isoformat(),
            "date": datetime.utcnow().date().isoformat(),
            "method": "qr-mobile"
        }
        
        try:
            sb.table("attendance").insert(attendance_record).execute()
            
            # Update session attendance count
            sb.table("qr_sessions").update({
                "attendance_count": session_data.get("attendance_count", 0) + 1
            }).eq("session_id", session_id).execute()
        except Exception as e:
            print(f"[Attendance-DB] Error: {e}")
        
        return jsonify(
            success=True,
            verified=final_verified,
            name=matched_name if verified else name,
            face_verified=verified,
            location_verified=location_verified,
            confidence=confidence,
            message="Attendance recorded" + (" - All verifications passed!" if final_verified else " - Pending faculty review")
        )
    except Exception as e:
        print(f"[Mark-QR-Mobile] Error: {e}")
        return jsonify(success=False, error=str(e)), 500

def calculate_distance(lat1, lon1, lat2, lon2):
    """Calculate distance between two coordinates in km (Haversine formula)"""
    from math import radians, sin, cos, asin, sqrt
    R = 6371  # Earth's radius in km
    lat1, lon1, lat2, lon2 = map(radians, [lat1, lon1, lat2, lon2])
    dlat = lat2 - lat1
    dlon = lon2 - lon1
    a = sin(dlat/2)**2 + cos(lat1) * cos(lat2) * sin(dlon/2)**2
    c = 2 * asin(sqrt(a))
    return R * c

# ── USER MANAGEMENT API ──
@app.route("/api/users/register", methods=["POST"])
def user_register():
    """Register a new user (student/faculty/admin)."""
    d=request.json
    username=d.get("username","").strip().upper()  # Convert to uppercase
    password=d.get("password","")
    role=d.get("role","student")
    full_name=d.get("full_name","")
    email=d.get("email","")
    
    if not username or not password or role not in ["student","faculty","admin"]:
        return jsonify(success=False,error="Invalid registration data"),400
    
    if not sb:
        return jsonify(success=False,error="Supabase not configured"),500
    
    try:
        existing=sb.table("users").select("id").eq("username",username).execute()
        if existing.data:
            return jsonify(success=False,error="Username already taken"),400
        
        # SECURITY: Use bcrypt for password hashing
        pwd_hash=_hash_password_secure(password)
        
        result=sb.table("users").insert({
            "username":username,
            "password_hash":pwd_hash,
            "role":role,
            "full_name":full_name,
            "email":email,
            "created_at":datetime.utcnow().isoformat()
        }).execute()
        
        new_user = result.data[0] if result.data else {}
        user_id = new_user.get("id")

        # Mirror to RTDB (no password hash sent)
        rtdb_set(f"/users/{user_id}", {
            "id": user_id,
            "username": username,
            "role": role,
            "full_name": full_name,
            "email": email,
            "created_at": new_user.get("created_at", datetime.utcnow().isoformat()),
        })

        return jsonify(success=True, user_id=user_id, message="Registration successful")
    except Exception as e:
        return jsonify(success=False,error=str(e)),500

@app.route("/api/users/register-face", methods=["POST"])
def register_face():
    """Register face for an existing user. Accepts multipart (file) or JSON (base64)."""
    if not sb:
        return jsonify(success=False,error="Supabase not configured"),500
    
    try:
        img = None
        face_image_b64 = None
        user_id = None
        roll_no = request.json.get("roll_no", "").strip().upper() if request.is_json else request.form.get("roll_no", "").strip().upper()  # Convert to uppercase
        
        # Check if multipart form with image file
        if request.files.get('image'):
            image_file = request.files['image']
            print(f"[FACE] Received face file: {image_file.filename}")
            img = Image.open(image_file.stream).convert("RGB")
            # Get user identifier from form
            user_id = request.form.get("user_id") or request.form.get("id")
            roll_no = request.form.get("roll_no", "").strip().upper()  # Convert to uppercase
        
        # Otherwise handle JSON with base64 image
        elif request.is_json:
            d = request.json or {}
            face_image_b64 = d.get("face_image")
            user_id = d.get("user_id") or d.get("id")
            roll_no = d.get("roll_no", "").strip()
            
            if not face_image_b64:
                return jsonify(success=False,error="Missing face_image (base64)"),400
            
            img = decode_b64_image(face_image_b64)
        else:
            return jsonify(success=False,error="Send multipart form with 'image' file or JSON with 'face_image' (base64)"),400
        
        # Identify user: prefer roll_no, then user_id
        identified_user = None
        identified_user_id = None
        identified_roll_no = None
        
        # Validate that we have at least one identifier
        if not roll_no and not user_id:
            error_msg = "Missing user identifier. Provide either 'roll_no' or 'user_id'."
            print(f"[FACE] Error: {error_msg}")
            return jsonify(success=False, error=error_msg), 400
        
        if roll_no:
            print(f"[FACE] Looking up user by roll_no: {roll_no}")
            # Case-insensitive lookup: fetch all and check uppercase
            all_users = sb.table("users").select("id,roll_no,username").execute().data
            for user_rec in all_users:
                if (user_rec.get("roll_no") or "").upper() == roll_no:
                    identified_user = user_rec
                    identified_user_id = user_rec.get("id")
                    identified_roll_no = roll_no
                    break
            if not identified_user:
                # Try fallback: username might equal roll_no for students
                user_res = sb.table("users").select("id,roll_no,username").execute().data
                for user_rec in user_res:
                    if (user_rec.get("username") or "").upper() == roll_no:
                        identified_user = user_rec
                        identified_user_id = user_rec.get("id")
                        identified_roll_no = roll_no
                        break
        
        if not identified_user and user_id:
            print(f"[FACE] Looking up user by user_id: {user_id}")
            try:
                user_res = sb.table("users").select("id,roll_no,username").eq("id", user_id).execute()
                if user_res.data:
                    identified_user = user_res.data[0]
                    identified_user_id = identified_user.get("id")
                    identified_roll_no = identified_user.get("roll_no")
            except Exception as e:
                print(f"[FACE] Error looking up by user_id: {e}")
                # Continue to error handling below
        
        if not identified_user:
            error_msg = f"User not found. Provide valid roll_no or user_id."
            print(f"[FACE] Error: {error_msg}")
            return jsonify(success=False,error=error_msg),404
        
        # Process face image
        tmp = "tmp_face.jpg"
        img.save(tmp)
        print(f"[FACE] Image saved, detecting faces...")
        face_encs = encode_image(tmp)
        os.remove(tmp)
        
        if len(face_encs) != 1:
            error_msg = f"Found {len(face_encs)} faces. Ensure exactly ONE face is visible."
            print(f"[FACE] Error: {error_msg}")
            return jsonify(success=False,error=error_msg),400
        
        enc = face_encs[0].tolist()
        
        # Build payload for face_encodings table
        payload = {
            "user_id": identified_user_id,
            "roll_no": identified_roll_no,
            "encoding": enc,
            "created_at": datetime.utcnow().isoformat()
        }
        
        # If we have base64 from JSON, also store it
        if face_image_b64:
            payload["image"] = face_image_b64
        
        print(f"[FACE] Storing face encoding for user_id={identified_user_id}, roll_no={identified_roll_no}")

        try:
            upsert_payload = {
                "user_id":    identified_user_id,
                "roll_no":    identified_roll_no,
                "encoding":   json.dumps(enc) if isinstance(enc, list) else enc,
                "created_at": datetime.utcnow().isoformat()
            }
            if face_image_b64:
                upsert_payload["image"] = face_image_b64

            # Upsert so re-registering the same user replaces the old encoding
            sb.table("face_encodings").upsert(upsert_payload, on_conflict="roll_no").execute()
            print(f"[FACE] Face registered successfully")

            # Sync registration record to Firestore
            fstore_set("face_registrations", identified_roll_no or identified_user_id, {
                "user_id":       identified_user_id,
                "roll_no":       identified_roll_no,
                "registered_at": datetime.utcnow().isoformat(),
                "source":        "single_capture",
            })

            return jsonify(success=True, message="Face registered successfully",
                           user_id=identified_user_id, roll_no=identified_roll_no)
        except Exception as e:
            print(f"[SUPABASE] Error inserting face_encodings: {e}")
            import traceback
            traceback.print_exc()
            return jsonify(success=False, error=f"Supabase insert error: {str(e)}"), 500
            
    except Exception as e:
        print(f"[FACE] Exception in register_face: {e}")
        import traceback
        traceback.print_exc()
        return jsonify(success=False,error=str(e)),500

@app.route("/api/users/firebase-login", methods=["POST"])
@verify_firebase_token
def firebase_login():
    """Login via Firebase Auth.  The frontend signs in with Firebase, gets an ID
    token, sends it here.  We verify it, sync the user to Supabase, and return
    the user profile exactly like /api/users/login does."""
    fb = request.firebase_user  # set by @verify_firebase_token
    firebase_uid = fb["uid"]
    email = fb.get("email", "")
    display_name = fb.get("name", "") or email.split("@")[0]

    d = request.json or {}
    role = d.get("role", "student")
    latitude = d.get("latitude")
    longitude = d.get("longitude")
    face_image = d.get("face_image")

    # Sync to Supabase (create if not exists)
    user = sync_firebase_user_to_supabase(firebase_uid, email, display_name, role)
    if not user:
        return jsonify(success=False, error="Failed to sync user to database"), 500

    # ── CHECK 1: User must be ACTIVE ────────────────────────────
    is_active = user.get("is_active", False)
    if not is_active:
        print(f"[AUTH-FB] Firebase login rejected: {email} is inactive/archived")
        return jsonify(success=False, error="Your account is inactive. Contact admin."), 403

    # Validate selected role matches actual user role
    actual_role = user.get("role", role)
    if role and actual_role and role != actual_role:
        return jsonify(success=False, error=f"Invalid credentials. This account is not a {role} account."), 403

    resp_user = {
        "id": user.get("id"),
        "username": user.get("username"),
        "role": user.get("role", role),
        "full_name": user.get("full_name", display_name),
        "email": user.get("email", email),
        "roll_no": user.get("roll_no"),
        "employee_id": user.get("employee_id"),
        "firebase_uid": firebase_uid,
    }

    # student-specific: face registration check & attendance
    if resp_user["role"] == "student" and sb:
        face_registered = False
        try:
            if user.get("roll_no"):
                enc_res = sb.table("face_encodings").select("id").eq("roll_no", user["roll_no"]).execute()
                face_registered = bool(enc_res.data)
            if not face_registered:
                enc_res = sb.table("face_encodings").select("id").eq("user_id", user["id"]).execute()
                face_registered = bool(enc_res.data)
        except Exception as e:
            print(f"[FIREBASE-LOGIN] face_registered lookup error: {e}")

        resp_user["face_registered"] = face_registered
        if not face_registered and is_face_enabled():
            resp_user["needs_face_registration"] = True

        if is_face_enabled() and face_registered and face_image:
            verified, confidence = verify_face_for_user(
                user_id=user.get("id"), roll_no=user.get("roll_no"), image_b64=face_image)
            in_campus = None
            if latitude is not None and longitude is not None:
                try:
                    dist = haversine(float(latitude), float(longitude), 12.981139, 80.249593)
                    in_campus = dist <= 0.5
                except Exception:
                    in_campus = None
            try:
                sb.table("attendance").insert({
                    "name": user.get("full_name"),
                    "roll_no": user.get("roll_no"),
                    "course": user.get("department"),
                    "date": datetime.utcnow().date().isoformat(),
                    "timestamp": datetime.utcnow().isoformat(),
                    "verified": verified,
                    "method": "firebase-login",
                    "confidence": confidence,
                    "in_campus": in_campus,
                    "latitude": latitude,
                    "longitude": longitude,
                }).execute()
            except Exception as e:
                print(f"[ATTENDANCE] error recording firebase login attendance: {e}")
            resp_user["face_verified"] = verified
            resp_user["confidence"] = confidence
            resp_user["in_campus"] = in_campus

    # Sync user profile to RTDB on every successful Firebase login
    rtdb_set(f"/users/{resp_user['id']}", {k: v for k, v in resp_user.items() if v is not None})

    return jsonify(success=True, user=resp_user)


@app.route("/api/users/login", methods=["POST"])
@rate_limit_login
def user_login():
    """Verify user credentials and optionally handle face/location attendance. Rate-limited to 5 attempts per 15 minutes."""
    d=request.json
    username=d.get("username","" ).strip()
    password=d.get("password","" )
    selected_role=d.get("role","").strip()
    latitude=d.get("latitude")
    longitude=d.get("longitude")
    face_image=d.get("face_image")

    if not username or not password:
        return jsonify(success=False,error="Missing credentials"),400

    if not sb:
        return jsonify(success=False, error="Supabase not configured. Login disabled."), 500

    try:
        # SECURITY: Verify password using bcrypt checkpw
        result=sb.table("users").select("*").eq("username",username).execute()

        # Student: also try roll_no
        if not result.data:
            result=sb.table("users").select("*").eq("roll_no",username).execute()

        # Faculty/Admin: also try email
        if not result.data and '@' in username:
            result=sb.table("users").select("*").eq("email",username).execute()

        if not result.data:
            return jsonify(success=False,error="User not found"),404

        user=result.data[0]
        if not _verify_password_secure(password, user["password_hash"]):
            return jsonify(success=False,error="Invalid password"),401

        # ── CHECK 1: User must be ACTIVE ────────────────────────────
        is_active = user.get("is_active", False)
        if not is_active:
            print(f"[AUTH] Login rejected: {username} is inactive/archived")
            return jsonify(success=False, error="Your account is inactive. Contact admin."), 403

        # Check if user is suspended (RTDB flag set by admin)
        if _firebase_db_enabled:
            try:
                rtdb_u = firebase_db.reference(f"/users/{user['id']}").get()
                if rtdb_u and rtdb_u.get("is_active") is False:
                    return jsonify(success=False, error="Your account is inactive. Contact admin."), 403
            except Exception:
                pass

        # Validate that the selected role matches the user's actual role
        if selected_role and user.get("role") and selected_role != user["role"]:
            return jsonify(success=False, error=f"Invalid credentials. This account is not a {selected_role} account."), 403

        resp_user={
            "id":user["id"],
            "username":user["username"],
            "role":user["role"],
            "full_name":user["full_name"],
            "email":user["email"],
            "roll_no":user.get("roll_no"),
            "employee_id":user.get("employee_id")
        }

        # student-specific logic
        if user["role"] == "student":
            face_registered = False
            try:
                # determine whether this student already has a face encoding
                # prefer lookup by roll number (schema may not contain user_id)
                if user.get("roll_no"):
                    enc_res = sb.table("face_encodings").select("id").eq("roll_no", user.get("roll_no")).execute()
                    face_registered = bool(enc_res.data)
                if not face_registered:
                    enc_res = sb.table("face_encodings").select("id").eq("user_id", user["id"]).execute()
                    face_registered = bool(enc_res.data)
            except Exception as e:
                print(f"[LOGIN] face_registered lookup error: {e}")
                # leave face_registered default value (False)

            resp_user["face_registered"] = face_registered

            if not face_registered and is_face_enabled():
                resp_user["needs_face_registration"] = True

            # if face provided and registered, verify and record attendance
            if is_face_enabled() and face_registered and face_image:
                verified, confidence = verify_face_for_user(user_id=user.get("id"), roll_no=user.get("roll_no"), image_b64=face_image)
                in_campus=None
                if latitude is not None and longitude is not None:
                    try:
                        dist = haversine(float(latitude), float(longitude), 12.981139, 80.249593)
                        in_campus = dist <= 0.5
                    except Exception:
                        in_campus=None
                if sb:
                    try:
                        sb.table("attendance").insert({
                            "name": user.get("full_name"),
                            "roll_no": user.get("roll_no"),
                            "course": user.get("department"),
                            "date": datetime.utcnow().date().isoformat(),
                            "timestamp": datetime.utcnow().isoformat(),
                            "verified": verified,
                            "method": "login",
                            "confidence": confidence,
                            "in_campus": in_campus,
                            "latitude": latitude,
                            "longitude": longitude
                        }).execute()
                    except Exception as e:
                        print(f"[ATTENDANCE] error recording login attendance: {e}")
                resp_user["face_verified"] = verified
                resp_user["confidence"] = confidence
                resp_user["in_campus"] = in_campus

        # Sync user profile to RTDB on every successful login
        rtdb_set(f"/users/{resp_user['id']}", {k: v for k, v in resp_user.items() if v is not None})

        return jsonify(success=True,user=resp_user)
    except Exception as e:
        return jsonify(success=False,error=str(e)),500

@app.route("/api/init-admin-demo", methods=["POST"])
def init_admin_demo():
    """Initialize admin_demo user for first-time setup. One-time use."""
    try:
        if not sb:
            return jsonify(success=False, error="Database not configured"), 500
        
        password = "Admin@123"
        username = "admin_demo"
        
        # Delete existing admin_demo if present
        try:
            sb.table("users").delete().eq("username", username).execute()
        except:
            pass  # Ignore if user doesn't exist
        
        # Create new admin_demo with correct hash
        pwd_hash = _hash_password_secure(password)
        
        user_data = {
            "id": str(__import__('uuid').uuid4()),
            "username": username,
            "email": "admin@smartams.demo",
            "full_name": "Demo Admin",
            "role": "admin",
            "password_hash": pwd_hash,
            "is_active": True,
            "department": "Administration",
        }
        
        sb.table("users").insert(user_data).execute()
        
        return jsonify(
            success=True, 
            message="Admin demo user created",
            username=username,
            password=password
        ), 200
    
    except Exception as e:
        logger.error(f"[INIT] Admin demo init error: {e}")
        return jsonify(success=False, error=str(e)), 500

def init_admin_demo2():
    """Initialize second demo admin user for testing. One-time use."""
    try:
        if not sb:
            return jsonify(success=False, error="Database not configured"), 500
        
        password = "SuperAdmin@456"
        username = "superadmin"
        
        # Delete existing superadmin if present
        try:
            sb.table("users").delete().eq("username", username).execute()
        except:
            pass  # Ignore if user doesn't exist
        
        # Create new superadmin with correct hash
        pwd_hash = _hash_password_secure(password)
        
        user_data = {
            "id": str(__import__('uuid').uuid4()),
            "username": username,
            "email": "superadmin@smartams.demo",
            "full_name": "Super Admin",
            "role": "admin",
            "password_hash": pwd_hash,
            "is_active": True,
            "department": "Administration",
        }
        
        sb.table("users").insert(user_data).execute()
        
        return jsonify(
            success=True, 
            message="Second demo admin user created",
            username=username,
            password=password
        ), 200
    
    except Exception as e:
        logger.error(f"[INIT] Admin demo2 init error: {e}")
        return jsonify(success=False, error=str(e)), 500

@app.route("/api/admin/reset-password", methods=["POST"])
def admin_reset_password():
    """Admin endpoint to reset a user's password. Can be called with username and new password."""
    try:
        d = request.json or {}
        username = d.get("username", "").strip()
        new_password = d.get("new_password", "").strip()
        admin_key = d.get("admin_key", "")
        
        # Basic security check
        if admin_key != os.getenv("ADMIN_RESET_KEY", "smartams-reset-admin-key"):
            return jsonify(success=False, error="Unauthorized"), 401
        
        if not username or not new_password:
            return jsonify(success=False, error="Missing username or password"), 400
        
        if not sb:
            return jsonify(success=False, error="Database not configured"), 500
        
        # Hash the new password
        pwd_hash = _hash_password_secure(new_password)
        
        # Update user
        result = sb.table("users").update({
            "password_hash": pwd_hash
        }).eq("username", username).execute()
        
        if result.data:
            return jsonify(success=True, message=f"Password reset for {username}"), 200
        else:
            return jsonify(success=False, error="User not found"), 404
    
    except Exception as e:
        logger.error(f"[ADMIN] Password reset error: {e}")
        return jsonify(success=False, error=str(e)), 500

@app.route("/api/users/list", methods=["GET"])
def list_users():
    """Get list of users filtered by role, department, and/or semester."""
    if not sb:
        return jsonify(success=True, users=[])
    
    role       = request.args.get("role")
    department = request.args.get("department")
    semester   = request.args.get("semester")
    is_admin_view = request.args.get("is_admin_view", "0") == "1"
    section = request.args.get("section")
    year    = request.args.get("year")
    
    try:
        q = sb.table("users").select(
            "id,username,role,full_name,department,roll_no,employee_id,"
            "program,section,designation,subjects,is_active,created_at,year,email"
        )
        
        if is_admin_view:
            pass  # Show all users
        else:
            q = q.eq("is_active", True)  # Show only active users
        
        if role:
            q = q.eq("role", role)
        if department:
            q = q.eq("department", department)
        if semester:
            q = q.eq("semester", int(semester))
        if section:
            q = q.eq("section", section)
        if year:
            try:
                q = q.eq("year", int(year))
            except ValueError:
                pass
        result = q.execute()
        return jsonify(success=True, users=result.data)
    except Exception as e:
        return jsonify(success=False, error=str(e)), 500

def _generate_faculty_emp_id(department, sb_client, extra_offset=0):
    """Generate faculty employee ID: UNIYYDEPT_SEQ (e.g. PUC26CSE001)."""
    uni_abbr = os.environ.get('UNI_ABBR', 'PUC')
    year2 = str(datetime.utcnow().year)[-2:]   # last 2 digits e.g. '26'
    dept = (department or 'GEN').upper()
    prefix = f"{uni_abbr}{year2}{dept}"
    max_seq = 0
    if sb_client:
        try:
            rows = sb_client.table("users").select("employee_id,username").eq("role", "faculty").execute().data or []
            for row in rows:
                for val in [row.get("employee_id",""), row.get("username","")]:
                    if val and val.upper().startswith(prefix.upper()):
                        tail = val[len(prefix):]
                        if tail.isdigit():
                            max_seq = max(max_seq, int(tail))
        except Exception:
            pass
    seq = max_seq + extra_offset + 1
    return f"{prefix}{seq:03d}"

@app.route("/api/users/add", methods=["POST"])
def add_user():
    """Add user by admin (JSON only, no face). Use /api/register-and-add-user for face+user together."""
    d=request.json
    username=d.get("username","").strip().upper()  # Convert to uppercase
    password=d.get("password","")
    role=d.get("role","student")
    full_name=d.get("full_name","")
    email=d.get("email","")
    roll_no=d.get("roll_no","").upper() if d.get("roll_no","") else ""  # Convert to uppercase
    department=d.get("department","")
    program=d.get("program","")
    section=d.get("section","")
    semester=d.get("semester", 1)
    employee_id=d.get("employee_id","").upper() if d.get("employee_id","") else ""  # Convert to uppercase
    designation=d.get("designation","")
    subjects=d.get("subjects","")
    # auto-generate admission id
    admission_no = str(uuid.uuid4())
    if role == "student":
        # ensure username matches roll
        if not roll_no:
            roll_no = generate_roll_number(department, int(semester or 1)).upper()
        username = roll_no

    if role == "faculty":
        # auto-generate employee_id if not provided
        if not employee_id:
            employee_id = _generate_faculty_emp_id(department, sb).upper()
        # username must equal employee_id for faculty
        username = employee_id
    
    if not username or not password or role not in ["student","faculty","admin"]:
        return jsonify(success=False,error="Invalid user data"),400
    
    if not sb:
        return jsonify(success=False,error="Supabase not configured"),500
    
    try:
        existing=sb.table("users").select("id").eq("username",username).execute()
        if existing.data:
            return jsonify(success=False,error="Username already taken"),400

        if roll_no and role == "student":
            # Check roll_no column (case-insensitive)
            existing_roll = sb.table("users").select("id").execute().data
            for row in existing_roll:
                if row.get("roll_no","").upper() == roll_no:
                    return jsonify(success=False, error=f"Roll number {roll_no} is already registered"), 400
            # Also check username column (catches legacy rows with no roll_no value set)
            if sb.table("users").select("id").eq("username", roll_no).execute().data:
                return jsonify(success=False, error=f"Roll number {roll_no} already exists"), 400
        
        # SECURITY: Use bcrypt for password hashing
        pwd_hash=_hash_password_secure(password)
        
        user_payload = {
            "username":    username,
            "password_hash": pwd_hash,
            "role":        role,
            "full_name":   full_name,
            "email":       email,
            "roll_no":     roll_no if role == "student" else None,
            "department":  department,
            "program":     program,
            "section":     section if role == "student" else None,
            "semester":    int(semester or 1) if role == "student" else None,
            "employee_id": employee_id if role == "faculty" else None,
            "designation": designation if role == "faculty" else None,
            "subjects":    subjects if role == "faculty" else None,
            "is_active":   True,
            "created_at":  datetime.utcnow().isoformat()
        }
        result = sb.table("users").insert(user_payload).execute()
        new_id = result.data[0]["id"] if result.data else None

        # ── Sync to RTDB + Firestore ──────────────────────────────
        if new_id:
            sync_payload = {k: (v if v is not None else "") for k, v in user_payload.items()
                            if k != "password_hash"}  # never sync password hash
            sync_payload["id"] = new_id
            rtdb_set(f"/users/{new_id}", sync_payload)
            fstore_set("users", new_id, sync_payload)

        return jsonify(success=True, user_id=new_id, message="User added successfully")
    except Exception as e:
        return jsonify(success=False, error=str(e)), 500

@app.route("/api/users/fix-duplicates", methods=["POST"])
def fix_duplicate_rolls():
    """Delete student rows that have no roll_no but whose username duplicates
    another student's roll_no — these are legacy ghost entries."""
    if not sb:
        return jsonify(success=False, error="Supabase not configured"), 500
    try:
        rows = sb.table("users").select("id,username,roll_no,full_name").eq("role","student").execute().data or []
        real_rolls = {r["roll_no"] for r in rows if r.get("roll_no")}
        deleted = []
        for r in rows:
            if not r.get("roll_no") and r.get("username") in real_rolls:
                sb.table("users").delete().eq("id", r["id"]).execute()
                deleted.append(r["username"])
        return jsonify(success=True, deleted=deleted,
                       message=f"Removed {len(deleted)} duplicate(s): {', '.join(deleted) or 'none found'}")
    except Exception as e:
        return jsonify(success=False, error=str(e)), 500

# ─────────────────────────────────────────────────────────
# ── ENHANCED QR CODE SYSTEM WITH SECURITY FEATURES ──
# ─────────────────────────────────────────────────────────

from qr_security import (
    QRSessionManager, DeviceFingerprint, FraudDetection,
    AuditTrail, OfflineQueue, QREncryption
)

# In-memory storage for active sessions (replace with Supabase in production)
active_qr_sessions = {}
qr_usage_logs = []

@app.route("/api/qr/generate", methods=["POST"])
def generate_qr_code():
    """Faculty: Generate encrypted QR code for attendance"""
    try:
        d = request.json
        faculty_id = d.get("faculty_id")
        course_id = d.get("course_id")
        subject = d.get("subject", "Class")
        validity_minutes = int(d.get("validity_minutes", 5))
        require_face = d.get("require_face", True)
        require_location = d.get("require_location", True)
        latitude = d.get("latitude")
        longitude = d.get("longitude")
        gps_radius = int(d.get("gps_radius_meters", 100))
        
        if not all([faculty_id, course_id]):
            return jsonify(success=False, error="Missing faculty_id or course_id"), 400
        
        # Generate QR session
        qr_session = QRSessionManager.generate_session_qr(
            course_id=course_id,
            faculty_id=faculty_id,
            subject=subject,
            validity_minutes=validity_minutes,
            require_face=require_face,
            require_location=require_location,
            latitude=latitude,
            longitude=longitude,
            gps_radius=gps_radius
        )
        
        # Store in database if available
        if sb:
            try:
                sb.table("qr_sessions").insert({
                    "session_id": qr_session["session_id"],
                    "course_id": course_id,
                    "faculty_id": faculty_id,
                    "expires_at": qr_session["expires_at"],
                    "encrypted_data": qr_session["encrypted_data"],
                    "qr_code_data": f"AMSQR:2.0:{qr_session['session_id']}:{qr_session['encrypted_data']}",
                    "validity_minutes": validity_minutes,
                    "latitude": latitude,
                    "longitude": longitude,
                    "gps_radius_meters": gps_radius,
                    "require_face": require_face,
                    "require_location": require_location,
                    "active": True
                }).execute()
            except Exception as e:
                print(f"[QR-DB] Error storing session: {e}")

        # Mirror QR session to RTDB for real-time live tracker
        rtdb_set(f"/qr_sessions/{qr_session['session_id']}", {
            "session_id": qr_session["session_id"],
            "course_id": course_id,
            "faculty_id": faculty_id,
            "subject": subject,
            "expires_at": qr_session["expires_at"],
            "validity_minutes": validity_minutes,
            "require_face": require_face,
            "require_location": require_location,
            "active": True,
            "created_at": datetime.utcnow().isoformat(),
        })
        
        # Log event
        audit_log = AuditTrail.log_qr_event(
            event_type="QR_GENERATED",
            user_id=faculty_id,
            session_id=qr_session["session_id"],
            details={"subject": subject, "validity_minutes": validity_minutes},
            severity="low"
        )
        
        return jsonify(
            success=True,
            session_id=qr_session["session_id"],
            qr_code_base64=qr_session["qr_code_base64"],
            expires_at=qr_session["expires_at"],
            validity_seconds=qr_session["validity_seconds"],
            subject=qr_session["subject"]
        )
    
    except Exception as e:
        print(f"[QR-GEN] Error: {e}")
        return jsonify(success=False, error=str(e)), 500


@app.route("/api/qr/validate", methods=["POST"])
def validate_qr():
    """Validate QR code format and expiry"""
    try:
        d = request.json
        qr_string = d.get("qr_data", "")
        
        valid, qr_data = QRSessionManager.validate_qr_data(qr_string)
        
        if not valid:
            return jsonify(success=False, error=qr_data.get("error")), 400
        
        # Check for duplicate use
        is_duplicate, dup_msg = FraudDetection.check_duplicate_use(
            qr_data.get("session_id"),
            d.get("student_id"),
            qr_usage_logs
        )
        
        if is_duplicate:
            AuditTrail.log_qr_event(
                event_type="DUPLICATE_QR_ATTEMPT",
                user_id=d.get("student_id"),
                session_id=qr_data.get("session_id"),
                details={"reason": dup_msg},
                severity="high"
            )
            return jsonify(success=False, error=dup_msg), 403
        
        return jsonify(
            success=True,
            session_id=qr_data.get("session_id"),
            subject=qr_data.get("subject"),
            require_face=qr_data.get("require_face"),
            require_location=qr_data.get("require_location"),
            latitude=qr_data.get("latitude"),
            longitude=qr_data.get("longitude"),
            gps_radius=qr_data.get("gps_radius")
        )
    
    except Exception as e:
        print(f"[QR-VAL] Error: {e}")
        return jsonify(success=False, error=str(e)), 500


@app.route("/api/qr/mark-attendance", methods=["POST"])
def mark_qr_attendance_enhanced():
    """Mark attendance with full security checks"""
    try:
        d = request.json
        session_id = d.get("session_id")
        student_id = d.get("student_id")
        roll_no = d.get("roll_no")
        name = d.get("name")
        face_image = d.get("face_image", "")
        latitude = d.get("latitude")
        longitude = d.get("longitude")
        device_fingerprint = d.get("device_fingerprint", "")
        user_agent = d.get("user_agent", "")
        ip_address = request.remote_addr
        
        if not all([session_id, roll_no, name]):
            return jsonify(success=False, error="Missing required fields"), 400
        
        # Get session from DB
        session_data = None
        if sb:
            result = sb.table("qr_sessions").select("*").eq("session_id", session_id).execute()
            if result.data:
                session_data = result.data[0]
        
        if not session_data:
            return jsonify(success=False, error="Session not found or expired"), 404
        
        # ── SECURITY CHECK 1: Location Verification ──
        location_verified = True
        location_error = ""
        if session_data.get("require_location") and latitude and longitude:
            location_anomaly, loc_msg = FraudDetection.check_location_anomaly(
                latitude, longitude,
                session_data.get("latitude"),
                session_data.get("longitude"),
                session_data.get("gps_radius_meters", 100)
            )
            if location_anomaly:
                location_verified = False
                location_error = loc_msg
        
        # ── SECURITY CHECK 2: Face Recognition ──
        face_verified = False
        face_confidence = 0.0
        face_name = name
        
        if session_data.get("require_face") and face_image:
            try:
                img = decode_b64_image(face_image)
                tmp = f"tmp_qr_face_{session_id}.jpg"
                img.save(tmp)
                face_encs = encode_image(tmp)
                os.remove(tmp)
                
                if face_encs:
                    encs, names = load_encodings()
                    if encs:
                        current_enc = face_encs[0]
                        distances = [np.linalg.norm(current_enc - e) for e in encs]
                        distances = np.array(distances)
                        idx = np.argmin(distances)
                        min_distance = distances[idx]
                        face_verified = min_distance <= 0.6
                        face_confidence = float(max(0, 1 - (min_distance / 2.0)))
                        face_name = names[idx] if face_verified else name
                        
                        # Check confidence threshold
                        conf_anomaly, conf_msg = FraudDetection.check_face_confidence(face_confidence)
                        if conf_anomaly:
                            face_verified = False
            except Exception as e:
                print(f"[QR-FACE] Error: {e}")
        
        # ── SECURITY CHECK 3: Rapid Reuse Detection ──
        rapid_reuse, reuse_msg = FraudDetection.check_rapid_reuse(session_id, qr_usage_logs)
        if rapid_reuse:
            AuditTrail.log_qr_event(
                event_type="RAPID_REUSE_ATTEMPT",
                user_id=student_id,
                session_id=session_id,
                details={"reason": reuse_msg},
                severity="critical"
            )
            return jsonify(success=False, error=reuse_msg), 403
        
        # ── SECURITY CHECK 4: Fraud Detection ──
        is_fraud, fraud_reason, severity = FraudDetection.detect_proxy_attempt(
            face_verified, face_confidence, location_verified, True, 0
        )
        
        if is_fraud:
            AuditTrail.log_qr_event(
                event_type="FRAUD_ATTEMPT",
                user_id=student_id,
                session_id=session_id,
                details={"reason": fraud_reason},
                severity=severity
            )
            return jsonify(success=False, error=f"Security check failed: {fraud_reason}"), 403
        
        # ── Mark Attendance ──
        final_status = "valid"
        if not face_verified or not location_verified:
            final_status = "partial"
        
        log_entry = {
            "session_id": session_id,
            "student_id": student_id,
            "roll_no": roll_no,
            "face_verified": face_verified,
            "face_confidence": face_confidence,
            "location_verified": location_verified,
            "latitude": latitude,
            "longitude": longitude,
            "device_fingerprint": device_fingerprint,
            "ip_address": ip_address,
            "device_os": DeviceFingerprint.extract_device_info(user_agent).get("os", "Unknown"),
            "device_browser": DeviceFingerprint.extract_device_info(user_agent).get("browser", "Unknown"),
            "status": final_status,
            "created_at": datetime.utcnow().isoformat()
        }
        qr_usage_logs.append(log_entry)
        
        # Store in database
        if sb:
            try:
                sb.table("qr_usage_log").insert({
                    "session_id": session_id,
                    "student_id": student_id,
                    "roll_no": roll_no,
                    "face_verified": face_verified,
                    "face_confidence": face_confidence,
                    "location_verified": location_verified,
                    "latitude": latitude,
                    "longitude": longitude,
                    "device_fingerprint": device_fingerprint,
                    "ip_address": ip_address,
                    "status": final_status,
                    "used_at": datetime.utcnow().isoformat()
                }).execute()
                
                # Also mark attendance
                attendance_record = {
                    "student_id": student_id,
                    "name": face_name,
                    "roll_no": roll_no,
                    "course_id": session_data.get("course_id"),
                    "date": datetime.utcnow().date().isoformat(),
                    "timestamp": datetime.utcnow().isoformat(),
                    "method": "qr",
                    "verified": face_verified,
                    "latitude": latitude,
                    "longitude": longitude,
                    "in_campus": location_verified,
                    "qr_session_id": session_id
                }
                att_result = sb.table("attendance").insert(attendance_record).execute()
                attendance_id = att_result.data[0]["id"] if att_result.data else None
                
                # Sync attendance to Firestore
                if _fstore and attendance_id:
                    firestore_att = {
                        "id": attendance_id,
                        "student_id": student_id,
                        "name": face_name,
                        "roll_no": roll_no,
                        "course_id": session_data.get("course_id"),
                        "date": datetime.utcnow().date().isoformat(),
                        "timestamp": datetime.utcnow().isoformat(),
                        "method": "qr",
                        "verified": face_verified,
                        "face_confidence": face_confidence,
                        "latitude": latitude,
                        "longitude": longitude,
                        "in_campus": location_verified,
                        "qr_session_id": session_id,
                        "synced_at": datetime.utcnow().isoformat()
                    }
                    try:
                        write_to_firestore("attendance", str(attendance_id), firestore_att)
                        print(f"[FIRESTORE] Attendance synced for {roll_no} (id={attendance_id})")
                    except Exception as fs_err:
                        print(f"[FIRESTORE] Warning: Could not sync attendance: {fs_err}")
                
            except Exception as e:
                print(f"[QR-DB] Error recording attendance: {e}")
        
        AuditTrail.log_qr_event(
            event_type="QR_ATTENDANCE_MARKED",
            user_id=student_id,
            session_id=session_id,
            details={"face_verified": face_verified, "location_verified": location_verified},
            severity="low"
        )

        # ── Write to Firebase RTDB for real-time live tracker ──
        today_str = datetime.utcnow().date().isoformat()
        rtdb_record = {
            "roll_no": roll_no,
            "name": face_name,
            "student_id": student_id,
            "session_id": session_id,
            "subject": session_data.get("subject", ""),
            "status": "present" if face_verified else "absent",
            "face_verified": face_verified,
            "location_verified": location_verified,
            "timestamp": int(datetime.utcnow().timestamp() * 1000),
            "method": "qr"
        }
        rtdb_set(f"/attendance/{today_str}/{session_id}/{roll_no}", rtdb_record)

        return jsonify(
            success=True,
            message="Attendance marked successfully",
            student_name=face_name,
            subject=session_data.get("subject", ""),
            face_verified=face_verified,
            face_confidence=f"{face_confidence:.2%}",
            location_verified=location_verified,
            timestamp=datetime.utcnow().isoformat()
        )
    
    except Exception as e:
        print(f"[QR-ATT] Error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify(success=False, error=str(e)), 500


@app.route("/api/qr/device-fingerprint", methods=["POST"])
def register_device():
    """Register and manage device fingerprints"""
    try:
        d = request.json
        user_id = d.get("user_id")
        user_agent = d.get("user_agent", request.headers.get('User-Agent', ''))
        ip_address = request.remote_addr
        device_name = d.get("device_name", "")
        trust_device = d.get("trust_device", False)
        
        if not user_id:
            return jsonify(success=False, error="Missing user_id"), 400
        
        # Generate fingerprint
        fingerprint = DeviceFingerprint.generate_fingerprint(user_agent, ip_address)
        device_info = DeviceFingerprint.extract_device_info(user_agent)
        
        # Store in database
        if sb:
            try:
                sb.table("device_fingerprints").upsert({
                    "user_id": user_id,
                    "fingerprint_hash": fingerprint,
                    "device_name": device_name,
                    "os": device_info.get("os"),
                    "browser": device_info.get("browser"),
                    "ip_address": ip_address,
                    "trusted": trust_device,
                    "last_seen": datetime.utcnow().isoformat()
                }, on_conflict="fingerprint_hash").execute()
            except Exception as e:
                print(f"[FINGERPRINT] Error: {e}")
        
        return jsonify(
            success=True,
            fingerprint=fingerprint,
            device_info=device_info,
            message="Device registered successfully"
        )
    
    except Exception as e:
        print(f"[DEVICE] Error: {e}")
        return jsonify(success=False, error=str(e)), 500


@app.route("/api/qr/audit-log", methods=["GET"])
def get_audit_log():
    """Retrieve audit trail for security review"""
    try:
        session_id = request.args.get("session_id")
        user_id = request.args.get("user_id")
        severity = request.args.get("severity")
        limit = int(request.args.get("limit", 100))
        
        if not sb:
            return jsonify(success=False, error="Database not available"), 500
        
        q = sb.table("audit_trail").select("*")
        
        if session_id:
            q = q.eq("session_id", session_id)
        if user_id:
            q = q.eq("user_id", user_id)
        if severity:
            q = q.eq("severity", severity)
        
        result = q.order("created_at", desc=True).limit(limit).execute()
        
        return jsonify(
            success=True,
            logs=result.data,
            count=len(result.data)
        )
    
    except Exception as e:
        print(f"[AUDIT] Error: {e}")
        return jsonify(success=False, error=str(e)), 500


@app.route("/api/qr/attendance-history", methods=["GET"])
def get_qr_attendance_history():
    """Get attendance history via QR for student"""
    try:
        roll_no = request.args.get("roll_no")
        limit = int(request.args.get("limit", 50))
        
        if not roll_no:
            return jsonify(success=False, error="Missing roll_no"), 400
        
        if not sb:
            return jsonify(success=False, error="Database not available"), 500
        
        result = sb.table("attendance").select("*").eq("roll_no", roll_no).order("date", desc=True).limit(limit).execute()
        
        return jsonify(
            success=True,
            attendance_records=result.data,
            total_records=len(result.data)
        )
    
    except Exception as e:
        print(f"[HISTORY] Error: {e}")
        return jsonify(success=False, error=str(e)), 500


@app.route("/api/qr/create-profile", methods=["POST"])
def create_qr_profile():
    """Create personal QR profile for student"""
    try:
        import hashlib, qrcode
        from datetime import timedelta
        d = request.json
        user_id = d.get("user_id")
        roll_no = d.get("roll_no")
        full_name = d.get("full_name")
        email = d.get("email")
        share_enabled = d.get("share_enabled", True)
        expires_hours = int(d.get("expires_hours", 168))  # Default 7 days
        
        if not all([user_id, roll_no]):
            return jsonify(success=False, error="Missing required fields"), 400
        
        # Create profile data
        profile_data = {
            "user_id": user_id,
            "roll_no": roll_no,
            "full_name": full_name,
            "email": email,
            "created_at": datetime.utcnow().isoformat(),
            "version": "1.0"
        }
        
        # Generate QR code
        profile_qr_string = f"AMSPROFILE:{user_id}:{roll_no}:{base64.b64encode(json.dumps(profile_data).encode()).decode()}"
        profile_hash = hashlib.sha256(profile_qr_string.encode()).hexdigest()
        
        # Create QR image
        qr = qrcode.QRCode(version=1, box_size=10, border=5)
        qr.add_data(profile_qr_string)
        qr.make(fit=True)
        qr_image = qr.make_image(fill_color="black", back_color="white")
        
        buffer = BytesIO()
        qr_image.save(buffer, format='PNG')
        qr_base64 = base64.b64encode(buffer.getvalue()).decode()
        
        expires_at = datetime.utcnow() + timedelta(hours=expires_hours)
        
        # Store in database
        if sb:
            try:
                sb.table("qr_profiles").insert({
                    "user_id": user_id,
                    "roll_no": roll_no,
                    "profile_qr_data": profile_qr_string,
                    "profile_hash": profile_hash,
                    "share_enabled": share_enabled,
                    "expires_at": expires_at.isoformat()
                }).execute()
            except Exception as e:
                print(f"[PROFILE] Error: {e}")
        
        return jsonify(
            success=True,
            profile_hash=profile_hash,
            qr_code_base64=qr_base64,
            expires_at=expires_at.isoformat(),
            share_url=f"/api/qr/profile/{profile_hash}"
        )
    
    except Exception as e:
        print(f"[PROFILE-CREATE] Error: {e}")
        return jsonify(success=False, error=str(e)), 500


@app.route("/api/qr/profile/<profile_hash>", methods=["GET"])
def get_qr_profile(profile_hash):
    """Retrieve student QR profile"""
    try:
        if not sb:
            return jsonify(success=False, error="Database not available"), 500
        
        result = sb.table("qr_profiles").select("*").eq("profile_hash", profile_hash).execute()
        
        if not result.data:
            return jsonify(success=False, error="Profile not found"), 404
        
        profile = result.data[0]
        
        # Check expiry
        if profile.get("expires_at"):
            expires_at = datetime.fromisoformat(profile.get("expires_at"))
            if expires_at < datetime.utcnow():
                return jsonify(success=False, error="Profile has expired"), 410
        
        # Check if sharing is enabled
        if not profile.get("share_enabled"):
            return jsonify(success=False, error="Profile sharing is disabled"), 403
        
        # Increment view count
        try:
            sb.table("qr_profiles").update({
                "view_count": profile.get("view_count", 0) + 1
            }).eq("profile_hash", profile_hash).execute()
        except:
            pass
        
        return jsonify(
            success=True,
            roll_no=profile.get("roll_no"),
            user_id=profile.get("user_id"),
            view_count=profile.get("view_count", 0)
        )
    
    except Exception as e:
        print(f"[PROFILE-GET] Error: {e}")
        return jsonify(success=False, error=str(e)), 500


@app.route("/api/qr/offline-sync", methods=["POST"])
def sync_offline_queue():
    """Sync offline attendance entries to server"""
    try:
        d = request.json
        queue_entries = d.get("queue", [])
        user_id = d.get("user_id")
        
        if not queue_entries or not user_id:
            return jsonify(success=False, error="Missing queue or user_id"), 400
        
        synced_count = 0
        failed_entries = []
        
        for entry in queue_entries:
            try:
                if not sb:
                    failed_entries.append(entry)
                    continue
                
                # Try to sync the offline entry
                sb.table("offline_queue").update({
                    "synced": True,
                    "synced_at": datetime.utcnow().isoformat()
                }).eq("id", entry.get("id")).execute()
                
                synced_count += 1
            except Exception as e:
                print(f"[SYNC] Error syncing entry: {e}")
                failed_entries.append(entry)
        
        AuditTrail.log_qr_event(
            event_type="OFFLINE_SYNC",
            user_id=user_id,
            session_id="offline",
            details={"synced": synced_count, "failed": len(failed_entries)},
            severity="low"
        )
        
        return jsonify(
            success=True,
            synced_count=synced_count,
            failed_count=len(failed_entries),
            failed_entries=failed_entries
        )
    
    except Exception as e:
        print(f"[OFFLINE-SYNC] Error: {e}")
        return jsonify(success=False, error=str(e)), 500


@app.route("/api/qr/session-stats", methods=["GET"])
def get_session_stats():
    """Get real-time stats for an active QR session"""
    try:
        session_id = request.args.get("session_id")
        
        if not session_id:
            return jsonify(success=False, error="Missing session_id"), 400
        
        # Calculate stats from logs
        session_logs = [log for log in qr_usage_logs if log.get("session_id") == session_id]
        
        total_checkins = len(session_logs)
        face_verified = len([l for l in session_logs if l.get("face_verified")])
        location_verified = len([l for l in session_logs if l.get("location_verified")])
        anomalies = len([l for l in session_logs if l.get("status") == "partial" or l.get("status") == "fraud_attempt"])
        
        return jsonify(
            success=True,
            session_id=session_id,
            total_checkins=total_checkins,
            face_verified_count=face_verified,
            location_verified_count=location_verified,
            anomaly_count=anomalies,
            attendees=session_logs
        )
    
    except Exception as e:
        print(f"[STATS] Error: {e}")
        return jsonify(success=False, error=str(e)), 500


@app.route("/api/qr/session-reports", methods=["GET"])
def get_session_reports():
    """Generate detailed reports for QR sessions"""
    try:
        faculty_id = request.args.get("faculty_id")
        start_date = request.args.get("start_date")
        end_date = request.args.get("end_date")
        
        if not sb or not faculty_id:
            return jsonify(success=False, error="Missing parameters"), 400
        
        q = sb.table("qr_sessions").select("*").eq("faculty_id", faculty_id)
        if start_date:
            q = q.gte("created_at", start_date)
        if end_date:
            q = q.lte("created_at", end_date)
        
        sessions = q.execute().data
        
        # Compile statistics
        total_sessions = len(sessions)
        total_attendance = 0
        avg_attendance = 0
        security_incidents = 0
        
        for session in sessions:
            session_logs = [log for log in qr_usage_logs if log.get("session_id") == session.get("session_id")]
            total_attendance += len(session_logs)
            security_incidents += len([l for l in session_logs if l.get("status") != "valid"])
        
        avg_attendance = total_attendance / total_sessions if total_sessions > 0 else 0
        
        return jsonify(
            success=True,
            report={
                "period": f"{start_date} to {end_date}",
                "total_sessions": total_sessions,
                "total_attendance_marks": total_attendance,
                "average_attendance_per_session": round(avg_attendance, 2),
                "security_incidents": security_incidents,
                "fraud_attempts": len([l for l in qr_usage_logs if l.get("status") == "fraud_attempt"]),
                "duplicate_attempts": len([l for l in qr_usage_logs if l.get("status") == "duplicate"]),
                "face_verification_rate": round((len([l for l in qr_usage_logs if l.get("face_verified")]) / max(total_attendance, 1)) * 100, 2),
                "location_verification_rate": round((len([l for l in qr_usage_logs if l.get("location_verified")]) / max(total_attendance, 1)) * 100, 2)
            },
            sessions=sessions[:10]  # Last 10 sessions
        )
    
    except Exception as e:
        print(f"[REPORTS] Error: {e}")
        return jsonify(success=False, error=str(e)), 500


@app.route("/api/qr/fraudulent-attempts", methods=["GET"])
def get_fraudulent_attempts():
    """Get flagged fraudulent attendance attempts"""
    try:
        session_id = request.args.get("session_id")
        limit = int(request.args.get("limit", 50))
        
        fraud_logs = [
            log for log in qr_usage_logs 
            if log.get("status") in ["fraud_attempt", "duplicate", "partial"]
        ]
        
        if session_id:
            fraud_logs = [l for l in fraud_logs if l.get("session_id") == session_id]
        
        fraud_logs = sorted(fraud_logs, key=lambda x: x.get("created_at", ""), reverse=True)[:limit]
        
        return jsonify(
            success=True,
            fraud_attempts=fraud_logs,
            count=len(fraud_logs)
        )
    
    except Exception as e:
        print(f"[FRAUD] Error: {e}")
        return jsonify(success=False, error=str(e)), 500


@app.route("/api/qr/student-attendance-summary", methods=["GET"])
def get_student_attendance_summary():
    """Get attendance summary for a student"""
    try:
        roll_no = request.args.get("roll_no")
        course_id = request.args.get("course_id")
        
        if not sb or not roll_no:
            return jsonify(success=False, error="Missing roll_no"), 400
        
        q = sb.table("attendance").select("*").eq("roll_no", roll_no)
        if course_id:
            q = q.eq("course_id", course_id)
        
        records = q.execute().data
        
        # Calculate stats
        total = len(records)
        verified = len([r for r in records if r.get("verified")])
        face_verified = len([r for r in records if r.get("method") == "qr"])
        manual = len([r for r in records if r.get("method") == "manual"])
        
        attendance_percentage = (verified / total * 100) if total > 0 else 0
        
        return jsonify(
            success=True,
            summary={
                "roll_no": roll_no,
                "total_classes": total,
                "present": verified,
                "attendance_percentage": round(attendance_percentage, 2),
                "qr_attendance": face_verified,
                "manual_attendance": manual,
                "last_marked": records[0].get("timestamp") if records else None
            },
            recent_records=records[:10]
        )
    
    except Exception as e:
        print(f"[SUMMARY] Error: {e}")
        return jsonify(success=False, error=str(e)), 500


# ── Min-pct detection by program ─────────────────────────
_PG_PROGRAMS = {
    "m.tech", "mtech", "m.e", "me ", "ph.d", "phd", "ph.d.",
    "mphil", "m.phil", "mba", "mca", "m.sc", "msc",
    "m.com", "mcom", "llm", "m.arch", "march",
}

def _min_pct_for_program(program: str) -> float:
    """Return 85.0 for postgraduate/doctoral programs, 75.0 for all others."""
    if not program:
        return 75.0
    lower = program.lower().strip()
    for pg in _PG_PROGRAMS:
        if lower.startswith(pg) or f" {pg}" in lower:
            return 85.0
    return 75.0


@app.route("/api/attendance/report", methods=["GET"])
def attendance_report():
    """
    Cumulative per-subject attendance report with university 75%/85% rule.
    Tracks lecture / tutorial / practical / seminar separately and in aggregate.
    Formula: (Classes Attended ÷ Classes Conducted) × 100

    Thresholds (per University norms):
      ≥ min_pct+10 → Safe
      ≥ min_pct    → Met / OK
      ≥ 65         → Warning (fine zone)
      < 65         → Critical (debarment risk)
    Programs: PG/PhD → 85%, all others → 75% (auto-detected from student record).
    """
    try:
        roll_no    = request.args.get("roll_no")
        student_id = request.args.get("student_id")
        # Caller can force a specific min_pct; otherwise auto-detect from program
        forced_min = request.args.get("min_pct")

        if not sb or not (roll_no or student_id):
            return jsonify(success=False, error="roll_no or student_id required"), 400

        # Auto-detect min_pct from student's program
        min_pct = float(forced_min) if forced_min else 75.0
        if not forced_min and sb:
            try:
                q_user = sb.table("users").select("program,section")
                if roll_no:
                    q_user = q_user.eq("roll_no", roll_no)
                else:
                    q_user = q_user.eq("id", student_id)
                user_rows = q_user.limit(1).execute().data
                if user_rows:
                    min_pct = _min_pct_for_program(user_rows[0].get("program") or "")
            except Exception:
                pass  # fall back to 75%

        # Fine / debarment thresholds (university standard)
        fine_threshold       = 65.0   # below → financial fine
        debarment_threshold  = 50.0   # below → exam debarment

        q = sb.table("attendance").select("*")
        if roll_no:
            q = q.eq("roll_no", roll_no)
        else:
            q = q.eq("student_id", student_id)

        records = q.order("date", desc=False).execute().data or []

        # Group by subject_name, then by session_type within that subject
        subjects = defaultdict(lambda: {"conducted": 0, "present": 0, "absent": 0,
                                         "late": 0, "types": defaultdict(lambda: {"c": 0, "p": 0})})

        for r in records:
            subj  = (r.get("subject_name") or r.get("course") or "General").strip()
            stype = (r.get("session_type") or "lecture").lower().strip()
            raw_status = (r.get("status") or "").lower().strip()
            verified   = bool(r.get("verified") or r.get("is_verified"))

            # Determine present/absent
            if raw_status == "present" or raw_status == "late" or (raw_status == "" and verified):
                is_present = True
            elif raw_status == "absent":
                is_present = False
            else:
                is_present = verified  # fallback on verified flag

            subjects[subj]["conducted"]       += 1
            subjects[subj]["types"][stype]["c"] += 1
            if raw_status == "late":
                subjects[subj]["late"] += 1
            if is_present:
                subjects[subj]["present"]          += 1
                subjects[subj]["types"][stype]["p"] += 1
            else:
                subjects[subj]["absent"] += 1

        subject_list   = []
        total_conducted = 0
        total_present   = 0

        for subj, data in subjects.items():
            conducted = data["conducted"]
            present   = data["present"]
            absent    = conducted - present
            pct       = round(present / conducted * 100, 1) if conducted else 0.0

            # Classes needed to reach min_pct:
            # (present + x) / (conducted + x) >= min_pct/100
            # => x >= (min_pct/100 * conducted - present) / (1 - min_pct/100)
            needed = 0
            if pct < min_pct and conducted > 0:
                denom = 1 - min_pct / 100
                needed = max(0, math.ceil((min_pct / 100 * conducted - present) / denom)) if denom > 0 else 0

            # Classes can skip while staying above min_pct:
            # present / (conducted + x) >= min_pct/100  => x <= present/(min_pct/100) - conducted
            can_miss = 0
            if pct >= min_pct and conducted > 0:
                can_miss = max(0, int(present / (min_pct / 100) - conducted))

            total_conducted += conducted
            total_present   += present

            if pct >= min_pct + 10:
                status_label = "safe"
            elif pct >= min_pct:
                status_label = "ok"
            elif pct >= fine_threshold:
                status_label = "warning"   # fine zone (65–min_pct)
            else:
                status_label = "critical"  # debarment risk (< 65%)

            # Build session-type breakdown
            type_breakdown = [
                {"type": t, "conducted": v["c"], "present": v["p"],
                 "percentage": round(v["p"] / v["c"] * 100, 1) if v["c"] else 0.0}
                for t, v in data["types"].items()
            ]

            subject_list.append({
                "subject":        subj,
                "conducted":      conducted,
                "present":        present,
                "absent":         absent,
                "late":           data["late"],
                "percentage":     pct,
                "status":         status_label,
                "classes_needed": needed,
                "can_miss":       can_miss,
                "type_breakdown": type_breakdown,
            })

        subject_list.sort(key=lambda x: x["percentage"])

        overall_pct = round(total_present / total_conducted * 100, 1) if total_conducted else 0.0
        overall_absent = total_conducted - total_present

        if overall_pct >= min_pct + 10:
            ov_status = "safe"
        elif overall_pct >= min_pct:
            ov_status = "ok"
        elif overall_pct >= fine_threshold:
            ov_status = "warning"   # fine zone
        else:
            ov_status = "critical"  # debarment zone

        # Aggregate across all session types (lecture+tutorial+practical+seminar)
        agg_by_type = {}
        for s in subject_list:
            for t in s.get("type_breakdown", []):
                k = t["type"]
                if k not in agg_by_type:
                    agg_by_type[k] = {"conducted": 0, "present": 0}
                agg_by_type[k]["conducted"] += t["conducted"]
                agg_by_type[k]["present"]   += t["present"]
        aggregate_by_type = [
            {
                "type":       k,
                "conducted":  v["conducted"],
                "present":    v["present"],
                "percentage": round(v["present"] / v["conducted"] * 100, 1) if v["conducted"] else 0.0,
            }
            for k, v in agg_by_type.items()
        ]

        return jsonify(
            success=True,
            roll_no=roll_no,
            overall={
                "conducted":   total_conducted,
                "present":     total_present,
                "absent":      overall_absent,
                "percentage":  overall_pct,
                "status":      ov_status,
            },
            subjects=subject_list,
            aggregate_by_type=aggregate_by_type,
            minimum_required=min_pct,
            fine_threshold=fine_threshold,
            debarment_threshold=debarment_threshold,
        )

    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify(success=False, error=str(e)), 500


@app.route("/api/attendance/batch-summary", methods=["GET"])
def attendance_batch_summary():
    """Cumulative attendance summary for all students in a batch/section.
    Query params: department, section, semester (optional).
    Returns per-student overall % with shortage warnings.
    """
    if not sb:
        return jsonify(success=False, error="Supabase not configured"), 500
    dept    = request.args.get("department")
    section = request.args.get("section")
    if not dept:
        return jsonify(success=False, error="department required"), 400
    try:
        q = sb.table("users").select("id,roll_no,full_name,section,year,semester").eq("role","student").eq("department", dept)
        if section:
            q = q.eq("section", section)
        students = q.execute().data or []
        if not students:
            return jsonify(success=True, students=[], count=0)

        roll_list = [s.get("roll_no","") for s in students if s.get("roll_no")]
        att_map = defaultdict(lambda: {"total": 0, "present": 0})
        for rn in roll_list:
            try:
                rows = sb.table("attendance").select("status").eq("roll_no", rn).execute().data or []
                for r in rows:
                    att_map[rn]["total"] += 1
                    if (r.get("status","")).lower() in ("present","late"):
                        att_map[rn]["present"] += 1
            except Exception:
                pass

        result = []
        shortage_count = 0
        for s in students:
            rn = s.get("roll_no","")
            total = att_map[rn]["total"]
            present = att_map[rn]["present"]
            pct = round(present / total * 100, 1) if total else 0.0
            status = "SHORTAGE" if pct < 75 else ("WARNING" if pct < 85 else "OK")
            if pct < 75:
                shortage_count += 1
            result.append({
                "roll_no": rn, "full_name": s.get("full_name",""),
                "section": s.get("section",""), "year": s.get("year"),
                "semester": s.get("semester"), "total_classes": total,
                "total_present": present, "percentage": pct, "status": status,
            })
        result.sort(key=lambda x: x["percentage"])
        return jsonify(success=True, students=result, count=len(result), shortage_count=shortage_count)
    except Exception as e:
        return jsonify(success=False, error=str(e)), 500


# ═══════════════════════════════════════════════════════════════
# ADMIN MANAGEMENT ROUTES
# ═══════════════════════════════════════════════════════════════

# ── ANNOUNCEMENTS ─────────────────────────────────────────────
@app.route("/api/announcements", methods=["GET"])
def get_announcements():
    if not sb:
        return jsonify(success=True, announcements=[])
    try:
        q = sb.table("announcements").select("*").order("created_at", desc=True)
        audience = request.args.get("audience")
        if audience:
            q = q.eq("target_audience", audience)
        result = q.execute()
        return jsonify(success=True, announcements=result.data or [])
    except Exception as e:
        return jsonify(success=False, error=str(e)), 500

@app.route("/api/announcements", methods=["POST"])
def create_announcement():
    if not sb:
        return jsonify(success=False, error="Supabase not configured"), 500
    d = request.json or {}
    if not d.get("title") or not d.get("content"):
        return jsonify(success=False, error="title and content required"), 400
    try:
        payload = {
            "title": d["title"],
            "content": d["content"],
            "posted_by": d.get("posted_by", "admin"),
            "target_audience": d.get("target_audience", "all"),
            "priority": d.get("priority", "info"),
            "valid_until": d.get("valid_until"),
            "created_at": datetime.utcnow().isoformat(),
        }
        result = sb.table("announcements").insert(payload).execute()
        record = result.data[0] if result.data else payload
        rtdb_set(f"/announcements/{record.get('id', str(uuid.uuid4()))}", record)
        return jsonify(success=True, announcement=record)
    except Exception as e:
        return jsonify(success=False, error=str(e)), 500

@app.route("/api/announcements/<ann_id>", methods=["DELETE"])
def delete_announcement(ann_id):
    if not sb:
        return jsonify(success=False, error="Supabase not configured"), 500
    try:
        sb.table("announcements").delete().eq("id", ann_id).execute()
        rtdb_delete(f"/announcements/{ann_id}")
        return jsonify(success=True)
    except Exception as e:
        return jsonify(success=False, error=str(e)), 500

# ── TIMETABLE ─────────────────────────────────────────────────
@app.route("/api/timetable", methods=["GET"])
def get_timetable():
    if not sb:
        return jsonify(success=True, entries=[], timetable=[])
    try:
        q = sb.table("timetable").select("*").order("day_of_week").order("period_number").order("start_time")
        faculty_id_param = (request.args.get("faculty_id") or "").strip()
        faculty_username_param = (request.args.get("faculty_username") or "").strip()
        if request.args.get("day"):
            q = q.eq("day_of_week", request.args["day"])
        if request.args.get("department"):
            q = q.eq("department", request.args["department"])
        if faculty_id_param:
            q = q.eq("faculty_id", faculty_id_param)
        if faculty_username_param:
            q = q.eq("faculty_username", faculty_username_param)
        if request.args.get("batch"):
            q = q.eq("batch", request.args["batch"])
        if request.args.get("section"):
            q = q.eq("section", request.args["section"])
        if request.args.get("year"):
            try:
                q = q.eq("year", int(request.args["year"]))
            except ValueError:
                pass
        if request.args.get("program"):
            q = q.eq("program", request.args["program"])
        sem_type = request.args.get("sem_type", "").strip().lower()
        if sem_type == "odd":
            q = q.in_("semester", [1, 3, 5, 7])
        elif sem_type == "even":
            q = q.in_("semester", [2, 4, 6, 8])
        elif request.args.get("semester"):
            try:
                q = q.eq("semester", int(request.args["semester"]))
            except ValueError:
                pass
        if request.args.get("block"):
            block_val = request.args["block"].strip()
            # block maps to batch prefix or room prefix (ilike search)
            q = q.ilike("batch", f"{block_val}%")
        if request.args.get("academic_year"):
            q = q.eq("academic_year", request.args["academic_year"])
        # Only active entries by default; pass active=all to bypass
        if request.args.get("active") != "all":
            try:
                q = q.eq("active", True)
            except Exception:
                pass
        result = q.execute()
        data = result.data or []

        # Fallback: if strict faculty_id+faculty_username filter returns nothing,
        # retry by username only (most timetable rows are keyed by faculty_username).
        if not data and (faculty_id_param or faculty_username_param):
            try:
                fq = sb.table("timetable").select("*").order("day_of_week").order("period_number").order("start_time")
                if request.args.get("day"):
                    fq = fq.eq("day_of_week", request.args["day"])
                if request.args.get("department"):
                    fq = fq.eq("department", request.args["department"])
                if request.args.get("batch"):
                    fq = fq.eq("batch", request.args["batch"])
                if request.args.get("section"):
                    fq = fq.eq("section", request.args["section"])
                if request.args.get("year"):
                    try:
                        fq = fq.eq("year", int(request.args["year"]))
                    except ValueError:
                        pass
                if request.args.get("program"):
                    fq = fq.eq("program", request.args["program"])
                if request.args.get("semester"):
                    try:
                        fq = fq.eq("semester", int(request.args["semester"]))
                    except ValueError:
                        pass
                if request.args.get("academic_year"):
                    fq = fq.eq("academic_year", request.args["academic_year"])
                if request.args.get("active") != "all":
                    try:
                        fq = fq.eq("active", True)
                    except Exception:
                        pass

                # Prefer explicit username; otherwise use faculty_id as username token.
                fallback_username = (faculty_username_param or faculty_id_param).strip().lower()
                if fallback_username:
                    fq = fq.eq("faculty_username", fallback_username)
                    data = fq.execute().data or []
            except Exception:
                pass

        return jsonify(success=True, entries=data, timetable=data)
    except Exception as e:
        return jsonify(success=False, error=str(e)), 500

@app.route("/api/timetable", methods=["POST"])
def create_timetable_entry():
    if not sb:
        return jsonify(success=False, error="Supabase not configured"), 500
    d = request.json or {}
    if not d.get("day_of_week") or not d.get("start_time") or not d.get("end_time"):
        return jsonify(success=False, error="day_of_week, start_time, end_time required"), 400
    try:
        payload = {
            "course_id":    d.get("course_id"),
            "subject_name": d.get("subject_name", ""),
            "faculty_id":   d.get("faculty_id", ""),
            "faculty_name": d.get("faculty_name", ""),
            "batch":        d.get("batch", ""),
            "department":   d.get("department", ""),
            "hour_number":  int(d.get("hour_number") or 1),
            "day_of_week":  d["day_of_week"],
            "start_time":   d["start_time"],
            "end_time":     d["end_time"],
            "room_number":  d.get("room_number", ""),
            "created_at":   datetime.utcnow().isoformat(),
        }
        result = sb.table("timetable").insert(payload).execute()
        record = result.data[0] if result.data else payload
        rtdb_set(f"/timetable/{record.get('id', str(uuid.uuid4()))}", record)
        return jsonify(success=True, entry=record)
    except Exception as e:
        return jsonify(success=False, error=str(e)), 500


@app.route("/api/timetable/<entry_id>", methods=["PUT"])
def update_timetable_entry(entry_id):
    if not sb:
        return jsonify(success=False, error="Supabase not configured"), 500
    d = request.json or {}
    try:
        allowed = ("day_of_week","start_time","end_time","room_number","course_id",
                   "subject_name","faculty_id","faculty_name","batch","department","hour_number")
        update_fields = {k: v for k, v in d.items() if k in allowed}
        result = sb.table("timetable").update(update_fields).eq("id", entry_id).execute()
        record = result.data[0] if result.data else update_fields
        rtdb_update(f"/timetable/{entry_id}", update_fields)
        return jsonify(success=True, entry=record)
    except Exception as e:
        return jsonify(success=False, error=str(e)), 500


@app.route("/api/timetable/<entry_id>", methods=["DELETE"])
def delete_timetable_entry(entry_id):
    """Delete timetable entry by archiving it (soft delete).
    Entry can be restored from archive later."""
    if not sb:
        return jsonify(success=False, error="Supabase not configured"), 500
    try:
        # Fetch the entry first
        entry_data = sb.table("timetable").select("*").eq("id", entry_id).execute().data
        if not entry_data:
            return jsonify(success=False, error="Entry not found"), 404
        
        entry = entry_data[0]
        
        # Archive the entry
        archive_record = {
            "original_id": entry.get("id"),
            "faculty_id": entry.get("faculty_id"),
            "faculty_name": entry.get("faculty_name"),
            "faculty_username": entry.get("faculty_username"),
            "batch": entry.get("batch"),
            "session_type": entry.get("session_type"),
            "subject": entry.get("subject"),
            "day_of_week": entry.get("day_of_week"),
            "hour_number": entry.get("hour_number"),
            "start_time": entry.get("start_time"),
            "end_time": entry.get("end_time"),
            "room_number": entry.get("room_number"),
            "academic_year": entry.get("academic_year"),
            "semester": entry.get("semester"),
            "mode": entry.get("mode"),
            "deleted_at": datetime.utcnow().isoformat()
        }
        
        sb.table("timetable_archive").insert(archive_record).execute()
        
        # Delete from timetable
        sb.table("timetable").delete().eq("id", entry_id).execute()
        rtdb_delete(f"/timetable/{entry_id}")
        fstore_delete("timetable", entry_id)
        
        return jsonify(success=True, message="Timetable entry archived (can be restored)")
    except Exception as e:
        return jsonify(success=False, error=str(e)), 500


@app.route("/api/timetable/delete-all", methods=["DELETE"])
def delete_all_timetable():
    """Archive ALL timetable entries (soft delete, can be restored).
    Synced to Supabase, RTDB, and Firestore."""
    if not sb:
        return jsonify(success=False, error="Supabase not configured"), 500
    try:
        # Fetch all timetable entries
        existing = sb.table("timetable").select("*").execute().data or []
        print(f"[DELETE-ALL] Found {len(existing)} timetable entries to archive")
        
        # Archive each entry
        archived_count = 0
        for entry in existing:
            archive_record = {
                "original_id": entry.get("id"),
                "faculty_id": entry.get("faculty_id"),
                "faculty_name": entry.get("faculty_name"),
                "faculty_username": entry.get("faculty_username"),
                "batch": entry.get("batch"),
                "session_type": entry.get("session_type"),
                "subject": entry.get("subject"),
                "day_of_week": entry.get("day_of_week"),
                "hour_number": entry.get("hour_number"),
                "start_time": entry.get("start_time"),
                "end_time": entry.get("end_time"),
                "room_number": entry.get("room_number"),
                "academic_year": entry.get("academic_year"),
                "semester": entry.get("semester"),
                "mode": entry.get("mode"),
                "deleted_at": datetime.utcnow().isoformat()
            }
            try:
                sb.table("timetable_archive").insert(archive_record).execute()
                archived_count += 1
            except Exception as e:
                print(f"[WARNING] Failed to archive entry {entry.get('id')}: {e}")
        
        print(f"[DELETE-ALL] Successfully archived {archived_count} entries")
        
        # Delete all entries from timetable in batches
        ids = [r["id"] for r in existing]
        deleted_count = 0
        BATCH = 100
        for i in range(0, max(len(ids), 1), BATCH):
            batch = ids[i:i + BATCH]
            if batch:
                try:
                    sb.table("timetable").delete().in_("id", batch).execute()
                    deleted_count += len(batch)
                except Exception as e:
                    print(f"[ERROR] Failed to delete batch: {e}")
                    raise
        
        print(f"[DELETE-ALL] Deleted {deleted_count} entries from main table")
        
        # Mirror to RTDB + Firestore (skip on error to avoid blocking)
        try:
            for rid in ids:
                try:
                    rtdb_delete(f"/timetable/{rid}")
                except Exception as e:
                    print(f"[WARNING] RTDB delete failed for {rid}: {e}")
                try:
                    fstore_delete("timetable", rid)
                except Exception as e:
                    print(f"[WARNING] Firestore delete failed for {rid}: {e}")
            
            # Wipe entire /timetable RTDB node for a clean slate
            try:
                if _firebase_db_enabled:
                    firebase_db.reference("/timetable").delete()
            except Exception as e:
                print(f"[WARNING] Failed to wipe RTDB timetable node: {e}")
        except Exception as e:
            print(f"[WARNING] Sync errors occurred: {e}")
        
        return jsonify(success=True, message=f"Archived {len(ids)} timetable entries (can be restored)")
    except Exception as e:
        print(f"[ERROR] delete_all_timetable failed: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify(success=False, error=str(e)), 500


@app.route("/api/timetable/delete-by-faculty", methods=["DELETE"])
def delete_timetable_by_faculty():
    """Archive all timetable entries for a faculty (soft delete, can be restored).
    Synced to Supabase, RTDB, and Firestore."""
    if not sb:
        return jsonify(success=False, error="Supabase not configured"), 500
    d = request.json or {}
    faculty_username = (d.get("faculty_username") or "").strip()
    faculty_id       = (d.get("faculty_id")       or "").strip()
    if not faculty_username and not faculty_id:
        return jsonify(success=False, error="faculty_username or faculty_id required"), 400
    
    print(f"[DELETE-FACULTY] Deleting timetable for faculty: {faculty_username or faculty_id}")
    
    try:
        archived_ids = []
        
        # Fetch entries to archive
        entries_to_archive = []
        if faculty_username:
            rows = sb.table("timetable").select("*").eq("faculty_username", faculty_username).execute().data or []
            entries_to_archive.extend(rows)
        if faculty_id and faculty_id != faculty_username:
            rows2 = sb.table("timetable").select("*").eq("faculty_id", faculty_id).execute().data or []
            entries_to_archive.extend(rows2)
        
        print(f"[DELETE-FACULTY] Found {len(entries_to_archive)} entries to archive")
        
        # Archive entries
        archived_count = 0
        for entry in entries_to_archive:
            try:
                archive_record = {
                    "original_id": entry.get("id"),
                    "faculty_id": entry.get("faculty_id"),
                    "faculty_name": entry.get("faculty_name"),
                    "faculty_username": entry.get("faculty_username"),
                    "batch": entry.get("batch"),
                    "session_type": entry.get("session_type"),
                    "subject": entry.get("subject"),
                    "day_of_week": entry.get("day_of_week"),
                    "hour_number": entry.get("hour_number"),
                    "start_time": entry.get("start_time"),
                    "end_time": entry.get("end_time"),
                    "room_number": entry.get("room_number"),
                    "academic_year": entry.get("academic_year"),
                    "semester": entry.get("semester"),
                    "mode": entry.get("mode"),
                    "deleted_at": datetime.utcnow().isoformat(),
                    "deletion_reason": "Faculty timetable archival"
                }
                sb.table("timetable_archive").insert(archive_record).execute()
                archived_ids.append(entry.get("id"))
                archived_count += 1
            except Exception as e:
                print(f"[WARNING] Failed to archive entry {entry.get('id')}: {e}")
        
        print(f"[DELETE-FACULTY] Successfully archived {archived_count} entries")
        
        # Delete from main table in batches
        deleted_count = 0
        BATCH = 100
        if faculty_username:
            try:
                sb.table("timetable").delete().eq("faculty_username", faculty_username).execute()
            except Exception as e:
                print(f"[ERROR] Failed to delete faculty timetable by username: {e}")
                raise
        if faculty_id and faculty_id != faculty_username:
            try:
                sb.table("timetable").delete().eq("faculty_id", faculty_id).execute()
            except Exception as e:
                print(f"[ERROR] Failed to delete faculty timetable by ID: {e}")
                raise
        
        # Sync deletion to RTDB and Firestore (skip on error to not block)
        try:
            for rid in archived_ids:
                try:
                    rtdb_delete(f"/timetable/{rid}")
                except Exception as e:
                    print(f"[WARNING] RTDB delete failed for {rid}: {e}")
                try:
                    fstore_delete("timetable", rid)
                except Exception as e:
                    print(f"[WARNING] Firestore delete failed for {rid}: {e}")
        except Exception as e:
            print(f"[WARNING] Sync errors occurred: {e}")
        
        print(f"[SUCCESS] Moved {len(archived_ids)} faculty timetable entries to archive")
        return jsonify(success=True, message="Faculty timetable entries archived (can be restored)", archived=len(archived_ids))
    except Exception as e:
        print(f"[ERROR] Delete_timetable_by_faculty failed: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify(success=False, error=str(e)), 500


@app.route("/api/timetable/delete-bulk", methods=["DELETE"])
def delete_timetable_bulk():
    """Soft-delete timetable entries (move to archive table instead of hard deleting).
    Body: { ids: [...] }
    """
    if not sb:
        return jsonify(success=False, error="Supabase not configured"), 500
    d = request.json or {}
    ids = [str(i).strip() for i in (d.get("ids") or []) if str(i).strip()]
    if not ids:
        return jsonify(success=False, error="ids array is required"), 400
    
    print(f"[DELETE-TIMETABLE] Attempting to delete {len(ids)} timetable entries")
    
    try:
        # 1. Fetch all records to be deleted
        timetable_records = sb.table("timetable").select("*").in_("id", ids).execute().data or []
        print(f"[DELETE-TIMETABLE] Found {len(timetable_records)} entries to archive")
        
        archived_count = 0
        # 2. Move to archive table
        for record in timetable_records:
            try:
                archive_record = {
                    "original_id": record.get("id"),
                    "faculty_id": record.get("faculty_id"),
                    "faculty_name": record.get("faculty_name"),
                    "faculty_username": record.get("faculty_username"),
                    "batch": record.get("batch"),
                    "session_type": record.get("session_type"),
                    "subject": record.get("subject"),
                    "day_of_week": record.get("day_of_week"),
                    "hour_number": record.get("hour_number"),
                    "start_time": record.get("start_time"),
                    "end_time": record.get("end_time"),
                    "room_number": record.get("room_number"),
                    "academic_year": record.get("academic_year"),
                    "semester": record.get("semester"),
                    "mode": record.get("mode"),
                    "deleted_at": datetime.utcnow().isoformat(),
                    "deletion_reason": d.get("reason", "Admin deletion")
                }
                sb.table("timetable_archive").insert(archive_record).execute()
                archived_count += 1
            except Exception as e:
                print(f"[WARNING] Failed to archive entry {record.get('id')}: {e}")
        
        print(f"[DELETE-TIMETABLE] Successfully archived {archived_count} entries")
        
        # 3. Hard delete from main table in batches
        deleted_count = 0
        BATCH = 100
        for i in range(0, len(ids), BATCH):
            batch = ids[i:i + BATCH]
            try:
                sb.table("timetable").delete().in_("id", batch).execute()
                deleted_count += len(batch)
            except Exception as e:
                print(f"[ERROR] Failed to delete batch of timetable entries: {e}")
                raise
        
        print(f"[DELETE-TIMETABLE] Deleted {deleted_count} entries from main table")
        
        # 4. Sync deletion to RTDB and Firestore (skip on error to not block)
        try:
            for rid in ids:
                try:
                    rtdb_delete(f"/timetable/{rid}")
                except Exception as e:
                    print(f"[WARNING] RTDB delete failed for {rid}: {e}")
                try:
                    fstore_delete("timetable", rid)
                except Exception as e:
                    print(f"[WARNING] Firestore delete failed for {rid}: {e}")
        except Exception as e:
            print(f"[WARNING] Sync errors occurred: {e}")
        
        print(f"[SUCCESS] Moved {len(ids)} timetable entries to archive")
        return jsonify(success=True, archived=len(ids), message=f"Archived {archived_count} timetable entries")
    except Exception as e:
        print(f"[ERROR] Timetable soft-delete failed: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify(success=False, error=str(e)), 500


@app.route("/api/users/delete-bulk", methods=["DELETE"])
def delete_users_bulk():
    """Soft-delete users (move to archive table instead of hard deleting).
    Optimized for batch sizes up to 4000+ users.
    Body: { user_ids: [...], reason: "optional reason" }
    Timeout: 60 seconds max
    """
    if not sb:
        return jsonify(success=False, error="Supabase not configured"), 500
    
    try:
        d = request.json or {}
        ids = [str(i).strip() for i in (d.get("user_ids") or d.get("ids") or []) if str(i).strip()]
        if not ids:
            print(f"[DELETE-USERS] Error: No user_ids provided")
            return jsonify(success=False, error="user_ids array is required"), 400
        
        print(f"[DELETE-USERS] Starting bulk delete for {len(ids)} users")
        start_time = datetime.utcnow()
        
        # Set a 60-second timeout for the operation
        import signal
        
        def timeout_handler(signum, frame):
            raise TimeoutError("Delete operation exceeded 60 seconds")
        
        # Only set timeout on Unix systems
        if hasattr(signal, 'SIGALRM'):
            signal.signal(signal.SIGALRM, timeout_handler)
            signal.alarm(60)
        
        try:
            # 1. Fetch ALL users in ONE query (much faster than one-by-one)
            user_records = []
            archived_count = 0
            deleted_count = 0
            roll_nos_to_delete = set()
            
            try:
                # Build query to fetch all users at once
                print(f"[DELETE-USERS] Fetching {len(ids)} users from database...")
                result = sb.table("users").select("*").in_("id", ids).execute()
                user_records = result.data if result.data else []
                print(f"[DELETE-USERS] Fetched {len(user_records)} users for archiving")
                
                # Collect roll_nos for face encoding cleanup
                for user in user_records:
                    if user.get("roll_no"):
                        roll_nos_to_delete.add(user.get("roll_no"))
            except Exception as e:
                print(f"[DELETE-USERS] Fetch error: {e}")
                # Fallback: try smaller chunks
                fetch_chunk_size = 50
                for chunk_idx in range(0, len(ids), fetch_chunk_size):
                    chunk = ids[chunk_idx:chunk_idx + fetch_chunk_size]
                    try:
                        result = sb.table("users").select("*").in_("id", chunk).execute()
                        user_records.extend(result.data if result.data else [])
                    except Exception as e2:
                        print(f"[DELETE-USERS] Chunk fetch failed: {e2}")
                        continue
            
            print(f"[DELETE-USERS] Found {len(user_records)} users to archive")
            
            # 2. Archive users in batches
            archive_batch_size = 100
            for batch_idx in range(0, len(user_records), archive_batch_size):
                batch = user_records[batch_idx:batch_idx + archive_batch_size]
                try:
                    archive_records = []
                    for user in batch:
                        archive_user = {
                            "original_id": user.get("id"),
                            "username": user.get("username"),
                            "email": user.get("email"),
                            "password_hash": user.get("password_hash"),
                            "full_name": user.get("full_name"),
                            "role": user.get("role"),
                            "roll_no": user.get("roll_no"),
                            "program": user.get("program"),
                            "section": user.get("section"),
                            "year": user.get("year"),
                            "semester": user.get("semester"),
                            "employee_id": user.get("employee_id"),
                            "designation": user.get("designation"),
                            "subjects": user.get("subjects"),
                            "department": user.get("department"),
                            "phone": user.get("phone"),
                            "firebase_uid": user.get("firebase_uid"),
                            "is_active": user.get("is_active"),
                            "created_at": user.get("created_at"),
                            "updated_at": user.get("updated_at"),
                            "last_login": user.get("last_login"),
                            "deletion_reason": d.get("reason", "Admin deletion"),
                            "deleted_at": datetime.utcnow().isoformat()
                        }
                        archive_records.append(archive_user)
                    
                    if archive_records:
                        try:
                            sb.table("users_archive").insert(archive_records).execute()
                            archived_count += len(archive_records)
                        except Exception as e:
                            print(f"[DELETE-USERS] Warning: Archive batch failed: {e}")
                except Exception as e:
                    print(f"[DELETE-USERS] Warning: Archive batch {batch_idx//archive_batch_size} error: {e}")
            
            # 3. Delete from main table
            delete_batch_size = 100
            try:
                for batch_idx in range(0, len(ids), delete_batch_size):
                    batch = ids[batch_idx:batch_idx + delete_batch_size]
                    try:
                        sb.table("users").delete().in_("id", batch).execute()
                        deleted_count += len(batch)
                    except Exception as e:
                        print(f"[DELETE-USERS] Batch delete warning: {e}")
                        # Fallback: one-by-one
                        for user_id in batch:
                            try:
                                sb.table("users").delete().eq("id", user_id).execute()
                                deleted_count += 1
                            except Exception as e2:
                                print(f"[WARNING] Failed to delete user {user_id}: {e2}")
            except Exception as e:
                print(f"[DELETE-USERS] Delete phase error: {e}")
            
            print(f"[DELETE-USERS] Deleted {deleted_count} users in {(datetime.utcnow() - start_time).total_seconds():.2f}s")
            
            # 3b. Cleanup face encodings
            if roll_nos_to_delete:
                try:
                    print(f"[DELETE-USERS] Cleaning {len(roll_nos_to_delete)} face encoding records...")
                    for roll_no in roll_nos_to_delete:
                        try:
                            sb.table("face_encodings").delete().eq("roll_no", roll_no).execute()
                        except Exception as e:
                            print(f"[WARNING] Face encoding cleanup for {roll_no}: {e}")
                except Exception as e:
                    print(f"[DELETE-USERS] Face cleanup error: {e}")
            
            # Return success immediately
            response = jsonify(
                success=True, 
                archived=archived_count,
                deleted=deleted_count,
                message=f"Archived {archived_count}, deleted {deleted_count} users"
            )
            
            # Async Firebase sync
            try:
                import threading
                def async_firebase_sync():
                    try:
                        for uid in ids:
                            try:
                                rtdb_delete(f"/users/{uid}")
                            except: pass
                            try:
                                fstore_delete("users", uid)
                            except: pass
                        print(f"[ASYNC] Firebase sync completed")
                    except Exception as e:
                        print(f"[ASYNC] Firebase sync error: {e}")
                
                sync_thread = threading.Thread(target=async_firebase_sync, daemon=True)
                sync_thread.start()
            except Exception as e:
                print(f"[WARNING] Async Firebase sync failed: {e}")
            
            return response
        
        finally:
            # Cancel alarm
            if hasattr(signal, 'SIGALRM'):
                signal.alarm(0)
    
    except TimeoutError as e:
        print(f"[DELETE-USERS] TIMEOUT: {str(e)}")
        return jsonify(success=False, error="Delete operation timed out. Try with fewer users.", details=str(e)), 504
    except Exception as e:
        print(f"[ERROR] User soft-delete failed: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify(success=False, error="Delete failed", details=str(e)), 500


# ────────────────────────────────────────────────────────────────────
# ARCHIVE MANAGEMENT ENDPOINTS
# ────────────────────────────────────────────────────────────────────

@app.route("/api/archive/users", methods=["GET"])
def get_archived_users():
    """Retrieve all archived users.
    ⚠️ ADMIN ONLY: Only Super Admin can access this endpoint.
    Query params: limit, offset, admin_role (required)
    """
    if not sb:
        return jsonify(success=False, error="Supabase not configured"), 500
    
    # ── ACCESS CONTROL: Only Super Admin can see archived users ────
    admin_role = request.args.get("admin_role", "").strip()
    if admin_role != "super_admin":
        print(f"[AUDIT] Unauthorized archive access attempt with role: {admin_role}")
        return jsonify(success=False, error="Only Super Admin can access archived users"), 403
    
    try:
        limit = int(request.args.get("limit", 100))
        offset = int(request.args.get("offset", 0))
        
        archived = sb.table("users_archive").select("*").order("deleted_at", desc=True).range(offset, offset + limit - 1).execute().data or []
        total_count = sb.table("users_archive").select("id", count="exact").execute()
        total = total_count.count if hasattr(total_count, 'count') else len(archived)
        
        print(f"[AUDIT] Super Admin accessed {len(archived)} archived users")
        return jsonify(success=True, archived_users=archived, total=total, limit=limit, offset=offset)
    except Exception as e:
        return jsonify(success=False, error=str(e)), 500


@app.route("/api/archive/timetable", methods=["GET"])
def get_archived_timetable():
    """Retrieve all archived timetable entries.
    ⚠️ ADMIN ONLY: Only Super Admin can access this endpoint.
    Query params: limit, offset, admin_role (required)
    """
    if not sb:
        return jsonify(success=False, error="Supabase not configured"), 500
    
    # ── ACCESS CONTROL: Only Super Admin can see archived timetable ────
    admin_role = request.args.get("admin_role", "").strip()
    if admin_role != "super_admin":
        print(f"[AUDIT] Unauthorized timetable archive access attempt with role: {admin_role}")
        return jsonify(success=False, error="Only Super Admin can access archived timetable"), 403
    
    try:
        limit = int(request.args.get("limit", 100))
        offset = int(request.args.get("offset", 0))
        
        archived = sb.table("timetable_archive").select("*").order("deleted_at", desc=True).range(offset, offset + limit - 1).execute().data or []
        total_count = sb.table("timetable_archive").select("id", count="exact").execute()
        total = total_count.count if hasattr(total_count, 'count') else len(archived)
        
        print(f"[AUDIT] Super Admin accessed {len(archived)} archived timetable entries")
        return jsonify(success=True, archived_timetable=archived, total=total, limit=limit, offset=offset)
    except Exception as e:
        return jsonify(success=False, error=str(e)), 500


@app.route("/api/users/archive-bulk", methods=["POST"])
def archive_users_bulk():
    """Archive multiple users (move to users_archive table).
    Optimized for batches up to 4000+ users.
    Body: { user_ids: [...] }
    """
    if not sb:
        return jsonify(success=False, error="Supabase not configured"), 500
    d = request.json or {}
    ids = [str(i).strip() for i in (d.get("user_ids") or d.get("ids") or []) if str(i).strip()]
    if not ids:
        print(f"[ARCHIVE-BULK] Error: No user_ids provided")
        return jsonify(success=False, error="user_ids array is required"), 400
    
    print(f"[ARCHIVE-BULK] Starting archive for {len(ids)} users")
    start_time = datetime.utcnow()
    
    try:
        archived_count = 0
        deleted_count = 0
        user_records = []
        
        # 1. Fetch all users - do in chunks
        fetch_chunk_size = 100
        for chunk_idx in range(0, len(ids), fetch_chunk_size):
            chunk = ids[chunk_idx:chunk_idx + fetch_chunk_size]
            for user_id in chunk:
                try:
                    result = sb.table("users").select("*").eq("id", user_id).execute()
                    if result.data:
                        user_records.extend(result.data)
                except Exception as e:
                    print(f"[ARCHIVE-BULK] Warning: Could not fetch user {user_id}: {e}")
        
        print(f"[ARCHIVE-BULK] Found {len(user_records)} users to archive")
        
        # 2. Archive in batches (batch inserts are faster)
        archive_batch_size = 100
        for batch_idx in range(0, len(user_records), archive_batch_size):
            batch = user_records[batch_idx:batch_idx + archive_batch_size]
            
            try:
                archive_records = []
                for user in batch:
                    archive_user = {
                        "original_id": user.get("id"),
                        "username": user.get("username"),
                        "email": user.get("email"),
                        "password_hash": user.get("password_hash"),
                        "full_name": user.get("full_name"),
                        "role": user.get("role"),
                        "roll_no": user.get("roll_no"),
                        "program": user.get("program"),
                        "section": user.get("section"),
                        "year": user.get("year"),
                        "semester": user.get("semester"),
                        "employee_id": user.get("employee_id"),
                        "designation": user.get("designation"),
                        "subjects": user.get("subjects"),
                        "department": user.get("department"),
                        "phone": user.get("phone"),
                        "firebase_uid": user.get("firebase_uid"),
                        "is_active": user.get("is_active"),
                        "created_at": user.get("created_at"),
                        "updated_at": user.get("updated_at"),
                        "last_login": user.get("last_login"),
                        "deleted_at": datetime.utcnow().isoformat()
                    }
                    archive_records.append(archive_user)
                
                # Batch insert to archive
                if archive_records:
                    try:
                        sb.table("users_archive").insert(archive_records).execute()
                        archived_count += len(archive_records)
                    except Exception as e:
                        print(f"[ARCHIVE-BULK] Warning: Batch insert failed: {e}")
                        # Fallback to one-by-one
                        for rec in archive_records:
                            try:
                                sb.table("users_archive").insert(rec).execute()
                                archived_count += 1
                            except Exception as e2:
                                print(f"[WARNING] Failed to archive single user: {e2}")
            except Exception as e:
                print(f"[ARCHIVE-BULK] Warning: Archive batch {batch_idx//archive_batch_size} failed: {e}")
        
        print(f"[ARCHIVE-BULK] Successfully archived {archived_count} users")
        
        # 3. Delete from main table in batches
        delete_batch_size = 100
        for batch_idx in range(0, len(ids), delete_batch_size):
            batch = ids[batch_idx:batch_idx + delete_batch_size]
            for user_id in batch:
                try:
                    sb.table("users").delete().eq("id", user_id).execute()
                    deleted_count += 1
                except Exception as e:
                    print(f"[ARCHIVE-BULK] Warning: Failed to delete user {user_id}: {e}")
        
        elapsed = (datetime.utcnow() - start_time).total_seconds()
        print(f"[ARCHIVE-BULK] Completed in {elapsed:.2f}s - Archived: {archived_count}, Deleted: {deleted_count}")
        
        return jsonify(
            success=True,
            archived=archived_count,
            deleted=deleted_count,
            message=f"Successfully archived {archived_count} users"
        )
    except Exception as e:
        print(f"[ERROR] Bulk archive failed: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify(success=False, error=str(e)), 500


@app.route("/api/archive/users/restore", methods=["POST"])
def restore_archived_user():
    """Restore an archived user back to active users table.
    ⚠️ ADMIN ONLY: Only Super Admin can restore users.
    Body: { archive_id: "uuid", admin_role: "super_admin" }
    """
    if not sb:
        return jsonify(success=False, error="Supabase not configured"), 500
    
    d = request.json or {}
    archive_id = d.get("archive_id", "").strip()
    admin_role = d.get("admin_role", "").strip()
    
    if not archive_id:
        return jsonify(success=False, error="archive_id is required"), 400
    
    # ── ACCESS CONTROL: Only Super Admin can restore users ────
    if admin_role != "super_admin":
        print(f"[AUDIT] Unauthorized user restore attempt with role: {admin_role}")
        return jsonify(success=False, error="Only Super Admin can restore archived users"), 403
    
    try:
        # Get archived user
        archived_user = sb.table("users_archive").select("*").eq("id", archive_id).execute().data
        if not archived_user:
            return jsonify(success=False, error="Archive record not found"), 404
        
        user = archived_user[0]
        
        # Create new user (with new ID since original_id might already exist)
        new_user = {
            "username": user.get("username"),
            "email": user.get("email"),
            "password_hash": user.get("password_hash"),
            "full_name": user.get("full_name"),
            "role": user.get("role"),
            "roll_no": user.get("roll_no"),
            "program": user.get("program"),
            "section": user.get("section"),
            "year": user.get("year"),
            "semester": user.get("semester"),
            "employee_id": user.get("employee_id"),
            "designation": user.get("designation"),
            "subjects": user.get("subjects"),
            "department": user.get("department"),
            "phone": user.get("phone"),
            "firebase_uid": user.get("firebase_uid"),
            "is_active": True
        }
        
        result = sb.table("users").insert(new_user).execute()
        
        if result.data:
            restored_id = result.data[0].get("id")
            rtdb_set(f"/users/{restored_id}", new_user)
            fstore_set("users", restored_id, new_user)
            
            # Remove from archive
            sb.table("users_archive").delete().eq("id", archive_id).execute()
            
            print(f"[AUDIT] Super Admin restored user: {user.get('username')} (new_id: {restored_id})")
            return jsonify(success=True, message="User restored successfully", new_id=restored_id)
        else:
            return jsonify(success=False, error="Failed to restore user"), 500
    except Exception as e:
        return jsonify(success=False, error=str(e)), 500


@app.route("/api/archive/users/purge", methods=["DELETE"])
def purge_archived_user():
    """Permanently delete an archived user (cannot be restored).
    ⚠️ ADMIN ONLY: Only Super Admin can permanently purge users.
    Body: { archive_id: "uuid", admin_role: "super_admin" }
    """
    if not sb:
        return jsonify(success=False, error="Supabase not configured"), 500
    
    d = request.json or {}
    archive_id = d.get("archive_id", "").strip()
    admin_role = d.get("admin_role", "").strip()
    
    if not archive_id:
        return jsonify(success=False, error="archive_id is required"), 400
    
    # ── ACCESS CONTROL: Only Super Admin can purge users ────
    if admin_role != "super_admin":
        print(f"[AUDIT] Unauthorized user purge attempt with role: {admin_role}")
        return jsonify(success=False, error="Only Super Admin can permanently delete archived users"), 403
    
    try:
        archived_user = sb.table("users_archive").select("username").eq("id", archive_id).execute().data
        sb.table("users_archive").delete().eq("id", archive_id).execute()
        
        username = archived_user[0].get("username") if archived_user else "unknown"
        print(f"[AUDIT] Super Admin permanently purged archived user: {username}")
        return jsonify(success=True, message="Archived user purged permanently")
    except Exception as e:
        return jsonify(success=False, error=str(e)), 500


@app.route("/api/archive/timetable/restore", methods=["POST"])
def restore_archived_timetable():
    """Restore an archived timetable entry back to active timetable.
    ⚠️ ADMIN ONLY: Only Super Admin can restore timetable entries.
    Body: { archive_id: "uuid", admin_role: "super_admin" }
    """
    if not sb:
        return jsonify(success=False, error="Supabase not configured"), 500
    
    d = request.json or {}
    archive_id = d.get("archive_id", "").strip()
    admin_role = d.get("admin_role", "").strip()
    
    if not archive_id:
        return jsonify(success=False, error="archive_id is required"), 400
    
    # ── ACCESS CONTROL: Only Super Admin can restore timetable ────
    if admin_role != "super_admin":
        print(f"[AUDIT] Unauthorized timetable restore attempt with role: {admin_role}")
        return jsonify(success=False, error="Only Super Admin can restore timetable entries"), 403
    
    try:
        # Get archived timetable entry
        archived_entry = sb.table("timetable_archive").select("*").eq("id", archive_id).execute().data
        if not archived_entry:
            return jsonify(success=False, error="Archive record not found"), 404
        
        entry = archived_entry[0]
        
        # Create new timetable entry (with new ID)
        new_entry = {
            "faculty_id": entry.get("faculty_id"),
            "faculty_name": entry.get("faculty_name"),
            "faculty_username": entry.get("faculty_username"),
            "batch": entry.get("batch"),
            "session_type": entry.get("session_type"),
            "subject": entry.get("subject"),
            "day_of_week": entry.get("day_of_week"),
            "hour_number": entry.get("hour_number"),
            "start_time": entry.get("start_time"),
            "end_time": entry.get("end_time"),
            "room_number": entry.get("room_number"),
            "academic_year": entry.get("academic_year"),
            "semester": entry.get("semester"),
            "mode": entry.get("mode")
        }
        
        result = sb.table("timetable").insert(new_entry).execute()
        
        if result.data:
            restored_id = result.data[0].get("id")
            rtdb_set(f"/timetable/{restored_id}", new_entry)
            fstore_set("timetable", restored_id, new_entry)
            
            # Remove from archive
            sb.table("timetable_archive").delete().eq("id", archive_id).execute()
            
            print(f"[AUDIT] Super Admin restored timetable entry: {entry.get('subject')} (new_id: {restored_id})")
            return jsonify(success=True, message="Timetable entry restored successfully", new_id=restored_id)
        else:
            return jsonify(success=False, error="Failed to restore timetable entry"), 500
    except Exception as e:
        return jsonify(success=False, error=str(e)), 500


@app.route("/api/archive/timetable/purge", methods=["DELETE"])
def purge_archived_timetable():
    """Permanently delete an archived timetable entry (cannot be restored).
    ⚠️ ADMIN ONLY: Only Super Admin can permanently purge timetable entries.
    Body: { archive_id: "uuid", admin_role: "super_admin" }
    """
    if not sb:
        return jsonify(success=False, error="Supabase not configured"), 500
    
    d = request.json or {}
    archive_id = d.get("archive_id", "").strip()
    admin_role = d.get("admin_role", "").strip()
    
    if not archive_id:
        return jsonify(success=False, error="archive_id is required"), 400
    
    # ── ACCESS CONTROL: Only Super Admin can purge timetable ────
    if admin_role != "super_admin":
        print(f"[AUDIT] Unauthorized timetable purge attempt with role: {admin_role}")
        return jsonify(success=False, error="Only Super Admin can permanently delete timetable entries"), 403
    
    try:
        archived_entry = sb.table("timetable_archive").select("subject").eq("id", archive_id).execute().data
        sb.table("timetable_archive").delete().eq("id", archive_id).execute()
        
        subject = archived_entry[0].get("subject") if archived_entry else "unknown"
        print(f"[AUDIT] Super Admin permanently purged archived timetable entry: {subject}")
        return jsonify(success=True, message="Archived timetable entry purged permanently")
    except Exception as e:
        return jsonify(success=False, error=str(e)), 500


@app.route("/api/timetable/bulk-upload", methods=["POST"])
def timetable_bulk_upload():
    """Validate and save many timetable slots at once.
    Body: { slots: [...], academic_year, semester, commit: bool }
    If commit=false: returns validation result + conflict report (dry-run).
    If commit=true:  saves all valid slots and creates course assignments.
    """
    if not sb:
        return jsonify(success=False, error="Supabase not configured"), 500
    d = request.json or {}
    slots       = d.get("slots", [])
    academic_yr = d.get("academic_year", "")
    semester    = d.get("semester", "")
    is_commit   = bool(d.get("commit", False))

    if not slots:
        return jsonify(success=False, error="slots array is required"), 400

    DAY_ORDER = ["Monday","Tuesday","Wednesday","Thursday","Friday","Saturday","Sunday"]

    # ── Build existing-timetable index for conflict detection ──
    try:
        existing_rows = sb.table("timetable").select(
            "id,faculty_id,faculty_name,batch,day_of_week,hour_number,start_time,end_time,room_number"
        ).execute().data or []
    except Exception as e:
        return jsonify(success=False, error=f"Could not read existing timetable: {e}"), 500

    # Index by (faculty_id, day, hour) and (room, day, hour) and (batch, day, hour)
    # Account for duration: theory=1 hour, lab=2 hours
    fac_slots  = {}   # (faculty_id, day, hour) -> existing entry
    room_slots = {}   # (room, day, hour) -> existing entry
    batch_slots = {}  # (batch, day, hour) -> existing entry
    for row in existing_rows:
        fid  = (row.get("faculty_id") or "").strip().lower()
        day  = row.get("day_of_week","")
        hr   = row.get("hour_number","")
        room = (row.get("room_number") or "").strip().lower()
        bat  = (row.get("batch") or "").strip().lower()
        sess_t = (row.get("session_type") or "lecture").lower()
        
        # Determine duration: theory=1 hour, lab/tutorial=2 hours
        duration = 2 if sess_t in ["lab", "tutorial"] else 1
        
        if fid and day and hr:
            hr_num = int(hr) if str(hr).isdigit() else 0
            # Index all hours this class occupies
            for i in range(duration):
                key = (fid, day, str(hr_num + i))
                fac_slots[key] = row
        
        if room and day and hr:
            hr_num = int(hr) if str(hr).isdigit() else 0
            for i in range(duration):
                key = (room.lower(), day, str(hr_num + i))
                room_slots[key] = row
        
        if bat and day and hr:
            hr_num = int(hr) if str(hr).isdigit() else 0
            for i in range(duration):
                key = (bat.lower(), day, str(hr_num + i))
                batch_slots[key] = row

    errors   = []
    warnings = []
    valid    = []

    for i, slot in enumerate(slots):
        row_num = i + 1
        errs = []

        day     = (slot.get("day_of_week") or "").strip().title()
        hr_raw  = slot.get("hour_number")
        subject = (slot.get("subject_name") or "").strip()
        fname   = (slot.get("faculty_name") or "").strip()
        fid     = (slot.get("faculty_id")   or "").strip()
        batch   = (slot.get("batch")        or "").strip()
        dept    = (slot.get("department")   or "").strip()
        room    = (slot.get("room_number")  or "").strip()
        start   = (slot.get("start_time")   or "").strip()
        end     = (slot.get("end_time")     or "").strip()
        sess_t  = (slot.get("session_type") or "lecture").strip().lower()
        lab_bat = (slot.get("lab_batch")    or "").strip()   # B1 / B2

        # Required field checks
        if day not in DAY_ORDER:
            errs.append(f"Row {row_num}: Invalid day '{day}' (must be Monday–Saturday)")
        if not subject:
            errs.append(f"Row {row_num}: subject_name is required")
        if not batch:
            errs.append(f"Row {row_num}: batch is required")
        if not start or not end:
            errs.append(f"Row {row_num}: start_time and end_time are required")

        try:
            hr = int(hr_raw) if hr_raw is not None else None
            if hr is None or not (1 <= hr <= 10):
                raise ValueError
        except (ValueError, TypeError):
            errs.append(f"Row {row_num}: hour_number must be 1–10")
            hr = None

        if errs:
            errors.extend(errs)
            continue

        hr_str = str(hr)
        
        # Determine duration for this new class: theory=1 hour, lab/tutorial=2 hours
        duration = 2 if sess_t in ["lab", "tutorial"] else 1

        # Conflict checks - check all hours this class will occupy
        if fid:
            for offset in range(duration):
                key = (fid.lower(), day, str(hr + offset))
                if key in fac_slots:
                    ex = fac_slots[key]
                    ex_hr = hr + offset
                    warnings.append(f"Row {row_num}: Faculty conflict — {fname or fid} already has a class at {day} Hour {ex_hr} ({ex.get('subject_name','?')})")
                    break  # Only warn once per faculty
        
        if room:
            for offset in range(duration):
                key = (room.lower(), day, str(hr + offset))
                if key in room_slots:
                    ex = room_slots[key]
                    ex_hr = hr + offset
                    warnings.append(f"Row {row_num}: Room conflict — {room} already booked at {day} Hour {ex_hr} ({ex.get('faculty_name','?')} / {ex.get('subject_name','?')})")
                    break  # Only warn once per room
        
        for offset in range(duration):
            bat_key = (batch.lower(), day, str(hr + offset))
            if bat_key in batch_slots:
                ex = batch_slots[bat_key]
                ex_hr = hr + offset
                warnings.append(f"Row {row_num}: Batch conflict — {batch} already has class at {day} Hour {ex_hr} ({ex.get('subject_name','?')})")
                break  # Only warn once per batch

        valid.append({
            "course_id":    slot.get("course_id"),
            "subject_name": subject,
            "subject_code": (slot.get("subject_code") or "").strip(),
            "faculty_id":   fid or None,
            "faculty_name": fname,
            "batch":        batch,
            "department":   dept,
            "hour_number":  hr,
            "day_of_week":  day,
            "start_time":   start,
            "end_time":     end,
            "room_number":  room,
            "session_type": sess_t,
            "lab_batch":    lab_bat or None,
            "academic_year": academic_yr,
            "semester":     semester,
            "created_at":   datetime.utcnow().isoformat(),
        })

    if not is_commit:
        # Dry-run: return preview
        return jsonify(
            success=True,
            dry_run=True,
            total=len(slots),
            valid_count=len(valid),
            error_count=len(errors),
            warning_count=len(warnings),
            errors=errors,
            warnings=warnings,
            preview=valid[:50],  # first 50 for preview table
        )

    # Commit phase
    if errors:
        return jsonify(
            success=False,
            error="Fix validation errors before committing.",
            errors=errors, warnings=warnings
        ), 400

    saved = []
    failed = []
    course_keys_created = set()

    for entry in valid:
        try:
            result = sb.table("timetable").insert(entry).execute()
            rec = result.data[0] if result.data else entry
            saved.append(rec)
            rtdb_set(f"/timetable/{rec.get('id', str(uuid.uuid4()))}", rec)

            # Auto-create course assignment if not already done
            ck = (entry.get("faculty_id") or "", entry["subject_name"], entry["batch"])
            if ck not in course_keys_created and entry.get("faculty_id"):
                try:
                    exists = sb.table("courses").select("id") \
                        .eq("course_name", entry["subject_name"]) \
                        .eq("faculty_id", entry["faculty_id"]) \
                        .eq("department", entry["department"]).limit(1).execute().data
                    if not exists:
                        course_payload = {
                            "course_code":   entry.get("subject_code") or entry["subject_name"][:8].upper().replace(" ",""),
                            "course_name":   entry["subject_name"],
                            "department":    entry["department"],
                            "faculty_id":    entry["faculty_id"],
                            "academic_year": academic_yr,
                            "semester":      int(semester) if str(semester).isdigit() else 1,
                            "credits":       3,
                            "created_at":    datetime.utcnow().isoformat(),
                        }
                        cr = sb.table("courses").insert(course_payload).execute()
                        if cr.data:
                            rtdb_set(f"/courses/{cr.data[0]['id']}", cr.data[0])
                    course_keys_created.add(ck)
                except Exception:
                    pass  # course creation is best-effort

        except Exception as e:
            failed.append({"slot": entry, "error": str(e)})

    return jsonify(
        success=True,
        committed=True,
        saved_count=len(saved),
        failed_count=len(failed),
        warnings=warnings,
        errors=[f["error"] for f in failed],
    )


# ── TIMETABLE CSV TEMPLATE ────────────────────────────────────
@app.route("/api/timetable/template", methods=["GET"])
def timetable_csv_template():
    """Return an Excel template for timetable upload (new MVP format)."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        import io as _io
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Timetable"

        headers = [
            "department","section","year","day_of_week","period_number",
            "start_time","end_time","subject_code","subject_name",
            "faculty_username","room_number","type","program","semester","academic_year",
        ]
        header_fill = PatternFill("solid", fgColor="1F4E79")
        header_font = Font(color="FFFFFF", bold=True)
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Hint row (prefixed with * — skipped by parser)
        hints = [
            "* Required","* Required","* 1–6","* Mon–Sat","* 1–10",
            "* HH:MM","* HH:MM","* Required","* Required",
            "* Required (username)","Optional","Theory/Lab/Tutorial","B.Tech","1","2025-26",
        ]
        hint_font = Font(color="7F7F7F", italic=True)
        for col, h in enumerate(hints, 1):
            cell = ws.cell(row=2, column=col, value=h)
            cell.font = hint_font

        # Sample rows
        samples = [
            ["CSE","A",1,"Monday",1,"09:00","10:00","CS301","Data Structures","faculty1","A101","Theory","B.Tech",1,"2025-26"],
            ["CSE","A",1,"Monday",2,"10:00","11:00","CS302","Algorithms","faculty2","A102","Theory","B.Tech",1,"2025-26"],
            ["CSE","B",1,"Tuesday",1,"09:00","10:00","CS301","Data Structures","faculty1","A103","Theory","B.Tech",1,"2025-26"],
            ["ECE","A",2,"Monday",1,"09:00","10:00","EC201","Digital Electronics","faculty3","B101","Theory","B.Tech",3,"2025-26"],
        ]
        for r_idx, row in enumerate(samples, 3):
            for c_idx, val in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=val)

        # Column widths
        widths = [15,10,6,12,8,10,10,12,25,18,10,10,10,10,12]
        for col, w in enumerate(widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

        buf = _io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        from flask import Response
        return Response(
            buf.read(),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=timetable_template.xlsx"},
        )
    except ImportError:
        # Fall back to CSV if openpyxl not available
        import csv, io
        headers = [
            "department","section","year","day_of_week","period_number",
            "start_time","end_time","subject_code","subject_name",
            "faculty_username","room_number","type","program","semester","academic_year",
        ]
        sample = ["CSE","A","1","Monday","1","09:00","10:00","CS301","Data Structures","faculty1","A101","Theory","B.Tech","1","2025-26"]
        buf = io.StringIO()
        w = csv.writer(buf)
        w.writerow(headers)
        w.writerow(sample)
        from flask import Response
        return Response(
            buf.getvalue(),
            mimetype="text/csv",
            headers={"Content-Disposition": "attachment; filename=timetable_template.csv"},
        )


# ── WORKLOAD ALLOCATION ───────────────────────────────────────
@app.route("/api/timetable/workload", methods=["GET"])
def get_workload():
    """Return faculty workload summary — hours per week per faculty for current timetable."""
    if not sb:
        return jsonify(success=True, workload=[])
    dept   = request.args.get("department","")
    ac_yr  = request.args.get("academic_year","")
    try:
        q = sb.table("timetable").select(
            "faculty_id,faculty_name,subject_name,batch,department,session_type,day_of_week,hour_number"
        )
        if dept:   q = q.eq("department", dept)
        if ac_yr:  q = q.eq("academic_year", ac_yr)
        rows = q.execute().data or []

        from collections import defaultdict
        fac_map = defaultdict(lambda: {
            "faculty_name": "", "department": "",
            "slots": [], "subjects": set(), "batches": set(),
            "hours_per_week": 0
        })
        for r in rows:
            fid = r.get("faculty_id") or r.get("faculty_name") or "unknown"
            fac_map[fid]["faculty_name"] = r.get("faculty_name","")
            fac_map[fid]["department"]   = r.get("department","")
            fac_map[fid]["hours_per_week"] += 1
            fac_map[fid]["subjects"].add(r.get("subject_name",""))
            fac_map[fid]["batches"].add(r.get("batch",""))
            fac_map[fid]["slots"].append({
                "day": r.get("day_of_week"), "hour": r.get("hour_number"),
                "subject": r.get("subject_name"), "batch": r.get("batch"),
                "session_type": r.get("session_type","lecture"),
            })

        result_list = []
        for fid, fw in fac_map.items():
            result_list.append({
                "faculty_id":     fid,
                "faculty_name":   fw["faculty_name"],
                "department":     fw["department"],
                "hours_per_week": fw["hours_per_week"],
                "subjects":       sorted(fw["subjects"]),
                "batches":        sorted(fw["batches"]),
                "slots":          fw["slots"],
                "overloaded":     fw["hours_per_week"] > 20,
                "underloaded":    fw["hours_per_week"] < 10,
            })
        result_list.sort(key=lambda x: -x["hours_per_week"])
        return jsonify(success=True, workload=result_list)
    except Exception as e:
        return jsonify(success=False, error=str(e)), 500


# ── TIMETABLE CONFLICT CHECK ──────────────────────────────────
@app.route("/api/timetable/conflicts", methods=["GET"])
def get_timetable_conflicts():
    """Scan entire timetable and return all conflicts."""
    if not sb:
        return jsonify(success=True, conflicts=[])
    try:
        rows = sb.table("timetable").select("*").execute().data or []
        from collections import defaultdict
        fac  = defaultdict(list)
        room = defaultdict(list)
        bat  = defaultdict(list)
        for r in rows:
            fid  = (r.get("faculty_id") or "").strip()
            rm   = (r.get("room_number") or "").strip().lower()
            b    = (r.get("batch") or "").strip().lower()
            day  = r.get("day_of_week","")
            hr   = str(r.get("hour_number",""))
            if fid: fac [(fid, day, hr)].append(r)
            if rm:  room[(rm,  day, hr)].append(r)
            if b:   bat [(b,   day, hr)].append(r)

        conflicts = []
        for key, entries in fac.items():
            if len(entries) > 1:
                subjects = ", ".join(e.get("subject_name","?") for e in entries)
                conflicts.append({"type": "faculty", "day": key[1], "hour": key[2],
                    "detail": f"Faculty '{entries[0].get('faculty_name',key[0])}' has {len(entries)} classes: {subjects}",
                    "entries": [e.get("id") for e in entries]})
        for key, entries in room.items():
            if len(entries) > 1:
                subjects = ", ".join(e.get("subject_name","?") for e in entries)
                conflicts.append({"type": "room", "day": key[1], "hour": key[2],
                    "detail": f"Room '{key[0]}' double-booked: {subjects}",
                    "entries": [e.get("id") for e in entries]})
        for key, entries in bat.items():
            if len(entries) > 1:
                subjects = ", ".join(e.get("subject_name","?") for e in entries)
                conflicts.append({"type": "batch", "day": key[1], "hour": key[2],
                    "detail": f"Batch '{key[0]}' double-booked: {subjects}",
                    "entries": [e.get("id") for e in entries]})

        return jsonify(success=True, conflicts=conflicts, count=len(conflicts))
    except Exception as e:
        return jsonify(success=False, error=str(e)), 500


# ── TIMETABLE SUBSTITUTION ────────────────────────────────────
@app.route("/api/timetable/<entry_id>/substitute", methods=["POST"])
def substitute_faculty(entry_id):
    """Assign a substitute faculty to a timetable slot (mid-semester change)."""
    if not sb:
        return jsonify(success=False, error="Supabase not configured"), 500
    d = request.json or {}
    sub_fid   = d.get("substitute_faculty_id","").strip()
    sub_fname = d.get("substitute_faculty_name","").strip()
    reason    = d.get("reason","").strip()
    if not sub_fname and not sub_fid:
        return jsonify(success=False, error="substitute_faculty_id or substitute_faculty_name required"), 400
    try:
        update = {
            "substitute_faculty_id":   sub_fid or None,
            "substitute_faculty_name": sub_fname,
            "substitute_reason":       reason,
            "substituted_at":          datetime.utcnow().isoformat(),
        }
        result = sb.table("timetable").update(update).eq("id", entry_id).execute()
        rec    = result.data[0] if result.data else update
        rtdb_update(f"/timetable/{entry_id}", update)
        fstore_update("timetable", entry_id, update)
        return jsonify(success=True, entry=rec)
    except Exception as e:
        return jsonify(success=False, error=str(e)), 500


# ── ONLINE CLASSES ────────────────────────────────────────────
@app.route("/api/online-classes", methods=["GET"])
def get_online_classes():
    if not sb:
        return jsonify(success=True, classes=[])
    try:
        q = sb.table("online_classes").select("*").order("scheduled_at", desc=True)
        if request.args.get("faculty_id"):
            q = q.eq("faculty_id", request.args["faculty_id"])
        if request.args.get("status"):
            q = q.eq("status", request.args["status"])
        result = q.execute()
        return jsonify(success=True, classes=result.data or [])
    except Exception as e:
        return jsonify(success=False, error=str(e)), 500

@app.route("/api/online-classes", methods=["POST"])
def create_online_class():
    if not sb:
        return jsonify(success=False, error="Supabase not configured"), 500
    d = request.json or {}
    if not d.get("title") or not d.get("scheduled_at"):
        return jsonify(success=False, error="title and scheduled_at required"), 400
    try:
        payload = {
            "course_id": d.get("course_id"),
            "faculty_id": d.get("faculty_id"),
            "title": d["title"],
            "scheduled_at": d["scheduled_at"],
            "duration_minutes": int(d.get("duration_minutes", 60)),
            "meeting_link": d.get("meeting_link", ""),
            "recording_link": d.get("recording_link", ""),
            "status": d.get("status", "scheduled"),
            "created_at": datetime.utcnow().isoformat(),
        }
        result = sb.table("online_classes").insert(payload).execute()
        record = result.data[0] if result.data else payload
        rtdb_set(f"/online_classes/{record.get('id', str(uuid.uuid4()))}", record)
        return jsonify(success=True, online_class=record)
    except Exception as e:
        return jsonify(success=False, error=str(e)), 500

@app.route("/api/online-classes/<oc_id>", methods=["PUT"])
def update_online_class(oc_id):
    if not sb:
        return jsonify(success=False, error="Supabase not configured"), 500
    d = request.json or {}
    try:
        allowed = ("title", "scheduled_at", "duration_minutes", "meeting_link", "recording_link", "status")
        update_fields = {k: v for k, v in d.items() if k in allowed}
        result = sb.table("online_classes").update(update_fields).eq("id", oc_id).execute()
        rtdb_update(f"/online_classes/{oc_id}", update_fields)
        return jsonify(success=True, online_class=result.data[0] if result.data else {})
    except Exception as e:
        return jsonify(success=False, error=str(e)), 500

@app.route("/api/online-classes/<oc_id>", methods=["DELETE"])
def delete_online_class(oc_id):
    if not sb:
        return jsonify(success=False, error="Supabase not configured"), 500
    try:
        sb.table("online_classes").delete().eq("id", oc_id).execute()
        rtdb_delete(f"/online_classes/{oc_id}")
        return jsonify(success=True)
    except Exception as e:
        return jsonify(success=False, error=str(e)), 500

# ── LEAVE APPLICATIONS ────────────────────────────────────────
@app.route("/api/leave-applications", methods=["GET"])
def get_leave_applications():
    if not sb:
        return jsonify(success=True, applications=[])
    try:
        q = sb.table("leave_applications").select("*").order("created_at", desc=True)
        if request.args.get("status"):
            q = q.eq("status", request.args["status"])
        if request.args.get("student_id"):
            q = q.eq("student_id", request.args["student_id"])
        result = q.execute()
        return jsonify(success=True, applications=result.data or [])
    except Exception as e:
        return jsonify(success=False, error=str(e)), 500

@app.route("/api/leave-applications", methods=["POST"])
def submit_leave():
    if not sb:
        return jsonify(success=False, error="Supabase not configured"), 500
    d = request.json or {}
    if not all([d.get("student_id"), d.get("leave_type"), d.get("from_date"), d.get("to_date"), d.get("reason")]):
        return jsonify(success=False, error="student_id, leave_type, from_date, to_date, reason required"), 400
    try:
        payload = {
            "student_id": d["student_id"],
            "leave_type": d["leave_type"],
            "from_date": d["from_date"],
            "to_date": d["to_date"],
            "reason": d["reason"],
            "status": "pending",
            "created_at": datetime.utcnow().isoformat(),
            "updated_at": datetime.utcnow().isoformat(),
        }
        result = sb.table("leave_applications").insert(payload).execute()
        record = result.data[0] if result.data else payload
        rtdb_set(f"/leave_applications/{record.get('id', str(uuid.uuid4()))}", record)
        return jsonify(success=True, application=record)
    except Exception as e:
        return jsonify(success=False, error=str(e)), 500

@app.route("/api/leave-applications/<leave_id>", methods=["PUT"])
def update_leave(leave_id):
    if not sb:
        return jsonify(success=False, error="Supabase not configured"), 500
    d = request.json or {}
    try:
        update_fields = {k: v for k, v in d.items() if k in ("status", "approved_by", "response")}
        update_fields["updated_at"] = datetime.utcnow().isoformat()
        result = sb.table("leave_applications").update(update_fields).eq("id", leave_id).execute()
        rtdb_update(f"/leave_applications/{leave_id}", update_fields)
        return jsonify(success=True, application=result.data[0] if result.data else {})
    except Exception as e:
        return jsonify(success=False, error=str(e)), 500

# ── FEES ──────────────────────────────────────────────────────
@app.route("/api/fees", methods=["GET"])
def get_fees():
    if not sb:
        return jsonify(success=True, fees=[])
    try:
        q = sb.table("fees").select("*").order("created_at", desc=True)
        if request.args.get("student_id"):
            q = q.eq("student_id", request.args["student_id"])
        if request.args.get("payment_status"):
            q = q.eq("payment_status", request.args["payment_status"])
        result = q.execute()
        return jsonify(success=True, fees=result.data or [])
    except Exception as e:
        return jsonify(success=False, error=str(e)), 500

@app.route("/api/fees", methods=["POST"])
def create_fee():
    if not sb:
        return jsonify(success=False, error="Supabase not configured"), 500
    d = request.json or {}
    if not all([d.get("student_id"), d.get("fee_type"), d.get("amount")]):
        return jsonify(success=False, error="student_id, fee_type, amount required"), 400
    try:
        payload = {
            "student_id": d["student_id"],
            "fee_type": d["fee_type"],
            "amount": float(d["amount"]),
            "due_date": d.get("due_date"),
            "payment_status": d.get("payment_status", "pending"),
            "payment_date": d.get("payment_date"),
            "payment_reference": d.get("payment_reference"),
            "created_at": datetime.utcnow().isoformat(),
        }
        result = sb.table("fees").insert(payload).execute()
        record = result.data[0] if result.data else payload
        rtdb_set(f"/fees/{record.get('id', str(uuid.uuid4()))}", record)
        return jsonify(success=True, fee=record)
    except Exception as e:
        return jsonify(success=False, error=str(e)), 500

@app.route("/api/fees/<fee_id>", methods=["PUT"])
def update_fee(fee_id):
    if not sb:
        return jsonify(success=False, error="Supabase not configured"), 500
    d = request.json or {}
    try:
        update_fields = {k: v for k, v in d.items() if k in ("payment_status", "payment_date", "payment_reference", "amount", "due_date")}
        result = sb.table("fees").update(update_fields).eq("id", fee_id).execute()
        rtdb_update(f"/fees/{fee_id}", update_fields)
        return jsonify(success=True, fee=result.data[0] if result.data else {})
    except Exception as e:
        return jsonify(success=False, error=str(e)), 500

# ── COURSES ───────────────────────────────────────────────────
@app.route("/api/courses", methods=["GET"])
def get_courses():
    if not sb:
        return jsonify(success=True, courses=[])
    try:
        q = sb.table("courses").select("*").order("created_at", desc=True)
        if request.args.get("department"):
            q = q.eq("department", request.args["department"])
        if request.args.get("faculty_id"):
            q = q.eq("faculty_id", request.args["faculty_id"])
        result = q.execute()
        return jsonify(success=True, courses=result.data or [])
    except Exception as e:
        return jsonify(success=False, error=str(e)), 500

@app.route("/api/courses", methods=["POST"])
def create_course():
    if not sb:
        return jsonify(success=False, error="Supabase not configured"), 500
    d = request.json or {}
    if not all([d.get("course_code"), d.get("course_name")]):
        return jsonify(success=False, error="course_code and course_name required"), 400
    try:
        payload = {
            "course_code": d["course_code"],
            "course_name": d["course_name"],
            "credits": int(d.get("credits", 3)),
            "department": d.get("department", ""),
            "semester": int(d.get("semester", 1)),
            "academic_year": d.get("academic_year", "2024-25"),
            "faculty_id": d.get("faculty_id"),
            "created_at": datetime.utcnow().isoformat(),
        }
        result = sb.table("courses").insert(payload).execute()
        record = result.data[0] if result.data else payload
        rtdb_set(f"/courses/{record.get('id', str(uuid.uuid4()))}", record)
        return jsonify(success=True, course=record)
    except Exception as e:
        return jsonify(success=False, error=str(e)), 500

@app.route("/api/courses/upsert", methods=["POST"])
def upsert_course():
    """Create or update a course assignment.
    Matches on (course_code, faculty_id, department) — updates if exists, inserts if not.
    """
    if not sb:
        return jsonify(success=False, error="Supabase not configured"), 500
    d = request.json or {}
    course_code = (d.get("course_code") or "").strip().upper()
    course_name = (d.get("course_name") or "").strip()
    faculty_id  = (d.get("faculty_id")  or "").strip() or None
    if not course_code or not course_name:
        return jsonify(success=False, error="course_code and course_name required"), 400
    try:
        dept    = d.get("department", "")
        sem     = int(d.get("semester", 1)) if d.get("semester") else 1
        ac_yr   = d.get("academic_year", "2025-26")
        credits = int(d.get("credits", 3)) if d.get("credits") else 3
        # Check if already exists
        q = sb.table("courses").select("id").eq("course_code", course_code)
        if faculty_id: q = q.eq("faculty_id", faculty_id)
        if dept:       q = q.eq("department", dept)
        existing = q.execute().data or []
        payload = {
            "course_code":   course_code,
            "course_name":   course_name,
            "credits":       credits,
            "department":    dept,
            "semester":      sem,
            "academic_year": ac_yr,
            "faculty_id":    faculty_id,
        }
        if existing:
            cid = existing[0]["id"]
            result = sb.table("courses").update(payload).eq("id", cid).execute()
            record = result.data[0] if result.data else {**payload, "id": cid}
            rtdb_set(f"/courses/{cid}", record)
            fstore_set("courses", cid, record)
            return jsonify(success=True, course=record, action="updated")
        else:
            payload["created_at"] = datetime.utcnow().isoformat()
            result = sb.table("courses").insert(payload).execute()
            record = result.data[0] if result.data else payload
            cid = str(record.get("id", ""))
            if cid:
                rtdb_set(f"/courses/{cid}", record)
                fstore_set("courses", cid, record)
            return jsonify(success=True, course=record, action="created")
    except Exception as e:
        return jsonify(success=False, error=str(e)), 500


@app.route("/api/courses/<course_id>", methods=["DELETE"])
def delete_course(course_id):
    if not sb:
        return jsonify(success=False, error="Supabase not configured"), 500
    try:
        sb.table("courses").delete().eq("id", course_id).execute()
        rtdb_delete(f"/courses/{course_id}")
        return jsonify(success=True)
    except Exception as e:
        return jsonify(success=False, error=str(e)), 500

@app.route("/api/faculty/<faculty_id>/previous-details", methods=["GET"])
def faculty_previous_details(faculty_id):
    """Return courses assigned to this faculty, optionally filtered."""
    if not sb:
        return jsonify(success=True, records=[])
    try:
        q = sb.table("courses").select("*").eq("faculty_id", faculty_id).order("created_at", desc=True)
        if request.args.get("department"):
            q = q.eq("department", request.args["department"])
        if request.args.get("semester"):
            try:
                q = q.eq("semester", int(request.args["semester"]))
            except ValueError:
                pass
        if request.args.get("academic_year"):
            q = q.eq("academic_year", request.args["academic_year"])
        result = q.execute()
        records = result.data or []
        # Apply subject_code filter in Python (partial match)
        sc_filter = request.args.get("subject_code", "").strip().upper()
        if sc_filter:
            records = [r for r in records if sc_filter in (r.get("course_code") or "").upper()]
        return jsonify(success=True, records=records)
    except Exception as e:
        return jsonify(success=False, error=str(e)), 500

# ── USER SUSPEND / ACTIVATE / DELETE ─────────────────────────
@app.route("/api/users/<user_id>/suspend", methods=["POST"])
def suspend_user(user_id):
    """Suspend a user — synced to Supabase, RTDB, and Firestore."""
    if not sb:
        return jsonify(success=False, error="Supabase not configured"), 500
    try:
        now = datetime.utcnow().isoformat()
        payload = {"is_active": False, "suspended_at": now}
        sb.table("users").update(payload).eq("id", user_id).execute()
        rtdb_update(f"/users/{user_id}", payload)
        fstore_update("users", user_id, payload)
        return jsonify(success=True, message="User suspended")
    except Exception as e:
        return jsonify(success=False, error=str(e)), 500

@app.route("/api/users/<user_id>/activate", methods=["POST"])
def activate_user(user_id):
    """Reactivate a suspended user — synced to Supabase, RTDB, and Firestore."""
    if not sb:
        return jsonify(success=False, error="Supabase not configured"), 500
    try:
        payload = {"is_active": True, "suspended_at": None}
        sb.table("users").update(payload).eq("id", user_id).execute()
        rtdb_update(f"/users/{user_id}", {"is_active": True, "suspended_at": ""})
        fstore_update("users", user_id, {"is_active": True, "suspended_at": ""})
        return jsonify(success=True, message="User activated")
    except Exception as e:
        return jsonify(success=False, error=str(e)), 500

@app.route("/api/users/<user_id>", methods=["DELETE"])
def delete_user(user_id):
    """Permanently delete a user — synced across Supabase, RTDB, and Firestore."""
    if not sb:
        return jsonify(success=False, error="Supabase not configured"), 500
    try:
        # face_encodings is keyed by roll_no, not user_id — fetch it first
        user_rows = sb.table("users").select("roll_no").eq("id", user_id).execute().data
        if user_rows and user_rows[0].get("roll_no"):
            sb.table("face_encodings").delete().eq("roll_no", user_rows[0]["roll_no"]).execute()
        sb.table("users").delete().eq("id", user_id).execute()
        rtdb_delete(f"/users/{user_id}")
        fstore_delete("users", user_id)
        print(f"[SYNC] Deleted user {user_id} from Supabase + RTDB + Firestore")
        return jsonify(success=True, message="User deleted")
    except Exception as e:
        return jsonify(success=False, error=str(e)), 500

@app.route("/api/users/<user_id>", methods=["PUT"])
def update_user(user_id):
    """Update user fields — synced to Supabase, RTDB, and Firestore."""
    if not sb:
        return jsonify(success=False, error="Supabase not configured"), 500
    d = request.json or {}
    allowed = ("full_name", "email", "department", "program", "section", "role",
               "employee_id", "roll_no", "designation", "subjects")
    update_fields = {k: v for k, v in d.items() if k in allowed}
    if not update_fields:
        return jsonify(success=False, error="No updatable fields provided"), 400
    try:
        result = sb.table("users").update(update_fields).eq("id", user_id).execute()
        rtdb_update(f"/users/{user_id}", update_fields)
        fstore_update("users", user_id, update_fields)
        print(f"[SYNC] Updated user {user_id} in Supabase + RTDB + Firestore")
        return jsonify(success=True, user=result.data[0] if result.data else {})
    except Exception as e:
        return jsonify(success=False, error=str(e)), 500

# ── DEPARTMENTS ───────────────────────────────────────────────
# Departments are stored in Supabase table `departments`.
# Schema: id uuid, name text, code text (3-letter), parent_code text,
#   programs jsonb (array of program objects  {name,code,batches:[...],semesters:int}), created_at
# If the table doesn't exist yet, fall back to an in-memory default list so the
# UI never breaks.
# ALL codes are strictly 3 letters.  parent_code links specializations to parent dept.

DEPT_DEFAULTS = [
    # ── CSE (General) ───────────────────────────────────────
    {"id":"d-cse", "name":"Computer Science & Engineering", "code":"CSE", "parent_code":None, "programs":[
        {"name":"B.Tech CSE","code":"CSE","batches":["CSE-A","CSE-B","CSE-C"],"semesters":8},
    ]},
    # ── CSE Specializations ─────────────────────────────────
    {"id":"d-aim", "name":"Artificial Intelligence & Machine Learning", "code":"AIM", "parent_code":"CSE", "programs":[
        {"name":"B.Tech AIML","code":"AIM","batches":["AIM-A","AIM-B"],"semesters":8},
    ]},
    {"id":"d-ads", "name":"Data Science", "code":"ADS", "parent_code":"CSE", "programs":[
        {"name":"B.Tech DS","code":"ADS","batches":["ADS-A","ADS-B"],"semesters":8},
    ]},
    {"id":"d-cbs", "name":"Cyber Security", "code":"CBS", "parent_code":"CSE", "programs":[
        {"name":"B.Tech CBS","code":"CBS","batches":["CBS-A","CBS-B"],"semesters":8},
    ]},
    {"id":"d-iot", "name":"Internet of Things", "code":"IOT", "parent_code":"CSE", "programs":[
        {"name":"B.Tech IoT","code":"IOT","batches":["IOT-A","IOT-B"],"semesters":8},
    ]},
    {"id":"d-clc", "name":"Cloud Computing", "code":"CLC", "parent_code":"CSE", "programs":[
        {"name":"B.Tech CC","code":"CLC","batches":["CLC-A","CLC-B"],"semesters":8},
    ]},
    {"id":"d-fsd", "name":"Full Stack Development", "code":"FSD", "parent_code":"CSE", "programs":[
        {"name":"B.Tech FSD","code":"FSD","batches":["FSD-A","FSD-B"],"semesters":8},
    ]},
    {"id":"d-bct", "name":"Blockchain Technology", "code":"BCT", "parent_code":"CSE", "programs":[
        {"name":"B.Tech BCT","code":"BCT","batches":["BCT-A"],"semesters":8},
    ]},
    {"id":"d-rat", "name":"Robotics & Automation", "code":"RAT", "parent_code":"CSE", "programs":[
        {"name":"B.Tech RA","code":"RAT","batches":["RAT-A"],"semesters":8},
    ]},
    {"id":"d-bda", "name":"Big Data Analytics", "code":"BDA", "parent_code":"CSE", "programs":[
        {"name":"B.Tech BDA","code":"BDA","batches":["BDA-A"],"semesters":8},
    ]},
    {"id":"d-dvo", "name":"DevOps Engineering", "code":"DVO", "parent_code":"CSE", "programs":[
        {"name":"B.Tech DevOps","code":"DVO","batches":["DVO-A"],"semesters":8},
    ]},
    # ── ECE (General) ───────────────────────────────────────
    {"id":"d-ece", "name":"Electronics & Communication Engineering", "code":"ECE", "parent_code":None, "programs":[
        {"name":"B.Tech ECE","code":"ECE","batches":["ECE-A","ECE-B","ECE-C"],"semesters":8},
    ]},
    # ── ECE Specializations ─────────────────────────────────
    {"id":"d-vls", "name":"VLSI Design", "code":"VLS", "parent_code":"ECE", "programs":[
        {"name":"B.Tech VLSI","code":"VLS","batches":["VLS-A"],"semesters":8},
    ]},
    {"id":"d-ebs", "name":"Embedded Systems", "code":"EBS", "parent_code":"ECE", "programs":[
        {"name":"B.Tech ES","code":"EBS","batches":["EBS-A"],"semesters":8},
    ]},
    {"id":"d-sgp", "name":"Signal Processing", "code":"SGP", "parent_code":"ECE", "programs":[
        {"name":"B.Tech SP","code":"SGP","batches":["SGP-A"],"semesters":8},
    ]},
    {"id":"d-wlc", "name":"Wireless Communication", "code":"WLC", "parent_code":"ECE", "programs":[
        {"name":"B.Tech WC","code":"WLC","batches":["WLC-A"],"semesters":8},
    ]},
    {"id":"d-rbe", "name":"Robotics (ECE)", "code":"RBE", "parent_code":"ECE", "programs":[
        {"name":"B.Tech Robotics","code":"RBE","batches":["RBE-A"],"semesters":8},
    ]},
    # ── EEE (General) ───────────────────────────────────────
    {"id":"d-eee", "name":"Electrical & Electronics Engineering", "code":"EEE", "parent_code":None, "programs":[
        {"name":"B.Tech EEE","code":"EEE","batches":["EEE-A","EEE-B"],"semesters":8},
    ]},
    # ── EEE Specializations ─────────────────────────────────
    {"id":"d-pws", "name":"Power Systems", "code":"PWS", "parent_code":"EEE", "programs":[
        {"name":"B.Tech PS","code":"PWS","batches":["PWS-A"],"semesters":8},
    ]},
    {"id":"d-cts", "name":"Control Systems", "code":"CTS", "parent_code":"EEE", "programs":[
        {"name":"B.Tech CS","code":"CTS","batches":["CTS-A"],"semesters":8},
    ]},
    {"id":"d-evh", "name":"Electric Vehicles", "code":"EVH", "parent_code":"EEE", "programs":[
        {"name":"B.Tech EV","code":"EVH","batches":["EVH-A"],"semesters":8},
    ]},
    {"id":"d-res", "name":"Renewable Energy Systems", "code":"RES", "parent_code":"EEE", "programs":[
        {"name":"B.Tech RES","code":"RES","batches":["RES-A"],"semesters":8},
    ]},
    {"id":"d-sgt", "name":"Smart Grid Technology", "code":"SGT", "parent_code":"EEE", "programs":[
        {"name":"B.Tech SGT","code":"SGT","batches":["SGT-A"],"semesters":8},
    ]},
    # ── Designing (each stream its own dept) ─────────────────
    {"id":"d-grd", "name":"Graphic Design", "code":"GRD", "parent_code":None, "programs":[
        {"name":"B.Des Graphic Design","code":"GRD","batches":["GRD-A"],"semesters":8},
    ]},
    {"id":"d-uix", "name":"UI/UX Design", "code":"UIX", "parent_code":None, "programs":[
        {"name":"B.Des UI/UX Design","code":"UIX","batches":["UIX-A"],"semesters":8},
    ]},
    {"id":"d-avx", "name":"Animation & VFX", "code":"AVX", "parent_code":None, "programs":[
        {"name":"B.Des Animation & VFX","code":"AVX","batches":["AVX-A"],"semesters":8},
    ]},
    {"id":"d-gmd", "name":"Game Design & Development", "code":"GMD", "parent_code":None, "programs":[
        {"name":"B.Des Game Design","code":"GMD","batches":["GMD-A"],"semesters":8},
    ]},
    {"id":"d-fdn", "name":"Fashion Design", "code":"FDN", "parent_code":None, "programs":[
        {"name":"B.Des Fashion Design","code":"FDN","batches":["FDN-A"],"semesters":8},
    ]},
    {"id":"d-itd", "name":"Interior Design", "code":"ITD", "parent_code":None, "programs":[
        {"name":"B.Des Interior Design","code":"ITD","batches":["ITD-A"],"semesters":8},
    ]},
    {"id":"d-txd", "name":"Textile Design", "code":"TXD", "parent_code":None, "programs":[
        {"name":"B.Des Textile Design","code":"TXD","batches":["TXD-A"],"semesters":8},
    ]},
    {"id":"d-ftv", "name":"Film & Television Production", "code":"FTV", "parent_code":None, "programs":[
        {"name":"B.Des Film & TV Production","code":"FTV","batches":["FTV-A"],"semesters":8},
    ]},
    # ── MBA (each specialization its own dept) ───────────────
    {"id":"d-mba", "name":"MBA General", "code":"MBA", "parent_code":None, "programs":[
        {"name":"MBA General","code":"MBA","batches":["MBA-A","MBA-B"],"semesters":4},
    ]},
    {"id":"d-mbf", "name":"MBA – Finance", "code":"MBF", "parent_code":"MBA", "programs":[
        {"name":"MBA Finance","code":"MBF","batches":["MBF-A"],"semesters":4},
    ]},
    {"id":"d-mbh", "name":"MBA – Human Resource Management", "code":"MBH", "parent_code":"MBA", "programs":[
        {"name":"MBA HRM","code":"MBH","batches":["MBH-A"],"semesters":4},
    ]},
    {"id":"d-mbm", "name":"MBA – Marketing", "code":"MBM", "parent_code":"MBA", "programs":[
        {"name":"MBA Marketing","code":"MBM","batches":["MBM-A"],"semesters":4},
    ]},
    {"id":"d-mbb", "name":"MBA – Business Analytics", "code":"MBB", "parent_code":"MBA", "programs":[
        {"name":"MBA Analytics","code":"MBB","batches":["MBB-A"],"semesters":4},
    ]},
    {"id":"d-mbo", "name":"MBA – Operations Management", "code":"MBO", "parent_code":"MBA", "programs":[
        {"name":"MBA Operations","code":"MBO","batches":["MBO-A"],"semesters":4},
    ]},
    {"id":"d-mbi", "name":"MBA – International Business", "code":"MBI", "parent_code":"MBA", "programs":[
        {"name":"MBA IB","code":"MBI","batches":["MBI-A"],"semesters":4},
    ]},
    {"id":"d-mbe", "name":"MBA – Entrepreneurship & Innovation", "code":"MBE", "parent_code":"MBA", "programs":[
        {"name":"MBA E&I","code":"MBE","batches":["MBE-A"],"semesters":4},
    ]},
    # ── BBA (each specialization its own dept) ───────────────
    {"id":"d-bba", "name":"BBA General", "code":"BBA", "parent_code":None, "programs":[
        {"name":"BBA General","code":"BBA","batches":["BBA-A","BBA-B"],"semesters":6},
    ]},
    {"id":"d-bbf", "name":"BBA – Finance", "code":"BBF", "parent_code":"BBA", "programs":[
        {"name":"BBA Finance","code":"BBF","batches":["BBF-A"],"semesters":6},
    ]},
    {"id":"d-bbm", "name":"BBA – Marketing", "code":"BBM", "parent_code":"BBA", "programs":[
        {"name":"BBA Marketing","code":"BBM","batches":["BBM-A"],"semesters":6},
    ]},
    {"id":"d-bbh", "name":"BBA – Human Resource Management", "code":"BBH", "parent_code":"BBA", "programs":[
        {"name":"BBA HRM","code":"BBH","batches":["BBH-A"],"semesters":6},
    ]},
    {"id":"d-bbi", "name":"BBA – International Business", "code":"BBI", "parent_code":"BBA", "programs":[
        {"name":"BBA IB","code":"BBI","batches":["BBI-A"],"semesters":6},
    ]},
    {"id":"d-bbe", "name":"BBA – Entrepreneurship", "code":"BBE", "parent_code":"BBA", "programs":[
        {"name":"BBA Entrepreneurship","code":"BBE","batches":["BBE-A"],"semesters":6},
    ]},
]

@app.route("/api/departments", methods=["GET"])
def get_departments():
    if not sb:
        return jsonify(success=True, departments=DEPT_DEFAULTS)
    try:
        result = sb.table("departments").select("*").order("name").execute()
        db_depts = result.data or []
        if not db_depts:
            return jsonify(success=True, departments=DEPT_DEFAULTS)
        # Merge: build default map by code, overlay with DB data,
        # and fill in missing programs from defaults
        default_map = {d["code"]: d for d in DEPT_DEFAULTS}
        db_map = {d["code"]: d for d in db_depts}
        merged = {}
        for code, dd in default_map.items():
            if code in db_map:
                row = dict(db_map[code])
                # If DB programs list is empty/missing, use defaults
                if not row.get("programs"):
                    row["programs"] = dd["programs"]
                merged[code] = row
            else:
                merged[code] = dd
        # Also include any DB depts not in defaults (custom depts)
        for code, row in db_map.items():
            if code not in merged:
                merged[code] = row
        return jsonify(success=True, departments=sorted(merged.values(), key=lambda x: x.get("name","")))
    except Exception:
        return jsonify(success=True, departments=DEPT_DEFAULTS)

@app.route("/api/departments", methods=["POST"])
def create_department():
    if not sb:
        return jsonify(success=False, error="Supabase not configured"), 500
    d = request.json or {}
    if not d.get("name") or not d.get("code"):
        return jsonify(success=False, error="name and code required"), 400
    code = d["code"].strip().upper()
    if len(code) != 3:
        return jsonify(success=False, error="Department code must be exactly 3 characters"), 400
    try:
        payload = {
            "name": d["name"].strip(),
            "code": code,
            "parent_code": (d.get("parent_code") or "").strip().upper() or None,
            "programs": d.get("programs", []),
            "created_at": datetime.utcnow().isoformat(),
        }
        result = sb.table("departments").insert(payload).execute()
        saved = result.data[0] if result.data else payload
        fstore_set("departments", saved.get("id", saved.get("code")), saved)
        return jsonify(success=True, department=saved)
    except Exception as e:
        return jsonify(success=False, error=str(e)), 500

@app.route("/api/departments/<dept_id>", methods=["PUT"])
def update_department(dept_id):
    if not sb:
        return jsonify(success=False, error="Supabase not configured"), 500
    d = request.json or {}
    allowed = ("name", "code", "parent_code", "programs")
    fields = {k: v for k, v in d.items() if k in allowed}
    if "code" in fields:
        fields["code"] = fields["code"].strip().upper()
    if "parent_code" in fields:
        fields["parent_code"] = (fields["parent_code"] or "").strip().upper() or None
    try:
        result = sb.table("departments").update(fields).eq("id", dept_id).execute()
        saved = result.data[0] if result.data else fields
        fstore_update("departments", dept_id, saved)
        return jsonify(success=True, department=saved)
    except Exception as e:
        return jsonify(success=False, error=str(e)), 500

@app.route("/api/departments/<dept_id>", methods=["DELETE"])
def delete_department(dept_id):
    if not sb:
        return jsonify(success=False, error="Supabase not configured"), 500
    try:
        sb.table("departments").delete().eq("id", dept_id).execute()
        fstore_delete("departments", dept_id)
        return jsonify(success=True)
    except Exception as e:
        return jsonify(success=False, error=str(e)), 500

# ── BULK USER IMPORT ──────────────────────────────────────────
# POST /api/users/bulk-import
# Optimized for HIGH PAYLOADS (1000+ records)
# Body: { users: [ {role, full_name, username, email, password, ...} ] }
# FLEXIBLE: Auto-maps common column name variations
@app.route("/api/users/bulk-import", methods=["POST"])
def bulk_import_users():
    if not sb:
        return jsonify(success=False, error="Supabase not configured"), 500
    
    data = request.json or {}
    users = data.get("users", [])
    if not users:
        return jsonify(success=False, error="No users provided"), 400

    logger.info(f"[BULK_IMPORT] Starting optimized import of {len(users)} users")
    
    import hashlib, re
    created, failed = [], []
    to_insert = []  # Batch insert queue
    BATCH_SIZE = 500  # Supabase batch insert size
    
    def _normalize_field(user_dict, *keys):
        """Try multiple key variations to extract a field value"""
        for key in keys:
            val = user_dict.get(key, "")
            if val and isinstance(val, str):
                return val.strip()
        return ""
    
    def _is_valid_email(email):
        """Check if string looks like email"""
        return "@" in email and "." in email.split("@")[-1]
    
    def _extract_username_from_email(email):
        """Extract username part from email"""
        if "@" in email:
            return email.split("@")[0]
        return email
    
    def _flush_batch_insert(payloads):
        """Batch insert optimized for Supabase"""
        if not payloads:
            return 0
        try:
            result = sb.table("users").insert(payloads).execute()
            return len(payloads)
        except Exception as e:
            logger.error(f"[BULK_IMPORT] Batch insert failed: {str(e)[:100]}")
            # Try individual inserts as fallback
            success = 0
            for p in payloads:
                try:
                    sb.table("users").insert(p).execute()
                    success += 1
                except:
                    pass
            return success
    
    # ── PHASE 1: Prepare all records (no DB queries yet) ──
    logger.info(f"[BULK_IMPORT] Phase 1: Validating {len(users)} records...")
    
    for idx, u in enumerate(users):
        try:
            role = _normalize_field(u, "role", "user_role", "account_type").lower().strip()
            if not role or role not in ("student", "faculty", "admin"):
                failed.append({"username": _normalize_field(u, "username", "email") or f"row_{idx}", "error": "Invalid role"})
                continue
            
            username = _normalize_field(u, "username", "user_id", "login_id")
            email = _normalize_field(u, "email", "email_id", "email_address")
            
            if not username and email and _is_valid_email(email):
                username = _extract_username_from_email(email)
            
            if not username:
                username = _normalize_field(u, "roll_no", "roll_number", "employee_id", "emp_id")
            
            if not username or username in ("Mr.", "Ms.", "Dr.", "Prof.", "faculty", ""):
                failed.append({"username": email or f"row_{idx}", "error": "Missing valid username"})
                continue
            
            # Build payload
            password = _normalize_field(u, "password") or (username + "@123")
            pwd_hash = _hash_password_secure(password)
            
            semester = None
            if role == "student":
                try:
                    semester = int(_normalize_field(u, "semester", "sem", "year") or "1")
                except:
                    semester = 1
            
            payload = {
                "username": username,
                "password_hash": pwd_hash,
                "role": role,
                "full_name": _normalize_field(u, "full_name", "name", "employee_name") or "",
                "email": email or "",
                "department": _normalize_field(u, "department", "dept") or "",
                "program": _normalize_field(u, "program", "course") or "",
                "section": _normalize_field(u, "section", "batch") or "",
                "semester": semester,
                "roll_no": _normalize_field(u, "roll_no", "roll_number") if role == "student" else None,
                "employee_id": _normalize_field(u, "employee_id", "emp_id") if role == "faculty" else None,
                "designation": _normalize_field(u, "designation", "position") or "",
                "subjects": _normalize_field(u, "subjects", "courses") or "",
                "is_active": True,
                "created_at": datetime.utcnow().isoformat(),
            }
            
            to_insert.append(payload)
            
        except Exception as e:
            failed.append({"username": _normalize_field(u, "username") or f"row_{idx}", "error": str(e)[:50]})
    
    logger.info(f"[BULK_IMPORT] Phase 1 complete: {len(to_insert)} valid, {len(failed)} invalid")
    
    # ── PHASE 2: Batch insert all prepared records ──
    logger.info(f"[BULK_IMPORT] Phase 2: Inserting {len(to_insert)} records in batches of {BATCH_SIZE}...")
    
    inserted = 0
    for i in range(0, len(to_insert), BATCH_SIZE):
        batch = to_insert[i:i+BATCH_SIZE]
        batch_num = (i // BATCH_SIZE) + 1
        total_batches = (len(to_insert) + BATCH_SIZE - 1) // BATCH_SIZE
        
        result = _flush_batch_insert(batch)
        inserted += result
        logger.info(f"[BULK_IMPORT] Batch {batch_num}/{total_batches}: +{result} records")
    
    created = [{"inserted": True}] * inserted  # Simplified for speed
    
    logger.info(f"[BULK_IMPORT] COMPLETE: {inserted} inserted, {len(failed)} failed")
    
    return jsonify(success=True,
                   created=inserted,
                   failed=len(failed),
                   errors=failed[:50])

# ── TIMETABLE BULK IMPORT ────────────────────────────────────
# POST /api/timetable/bulk-import
# Optimized for HIGH PAYLOADS (2000+ slots)
# Body: { slots: [ {DEPARTMENT, PROGRAM, SEMESTER, CLASS, DAY, START_TIME, END_TIME, ...} ] }
@app.route("/api/timetable/bulk-import", methods=["POST"])
def bulk_import_timetable():
    if not sb:
        return jsonify(success=False, error="Supabase not configured"), 500
    
    data = request.json or {}
    slots = data.get("slots", [])
    if not slots:
        return jsonify(success=False, error="No slots provided"), 400
    
    logger.info(f"[TIMETABLE_IMPORT] Starting optimized import of {len(slots)} slots")
    
    to_insert = []
    failed = []
    BATCH_SIZE = 500
    
    # ── PHASE 1: Validate and prepare all slots ──
    logger.info(f"[TIMETABLE_IMPORT] Phase 1: Validating {len(slots)} slots...")
    
    for idx, slot_data in enumerate(slots):
        try:
            # Normalize field names
            payload = {
                "SLOT_ID": str(slot_data.get("SLOT_ID", slot_data.get("slot_id", f"SLOT_{idx+1:04d}"))),
                "DEPARTMENT": str(slot_data.get("DEPARTMENT", slot_data.get("department", ""))).upper(),
                "PROGRAM": str(slot_data.get("PROGRAM", slot_data.get("program", ""))).upper(),
                "SEMESTER": str(slot_data.get("SEMESTER", slot_data.get("semester", "1"))),
                "CLASS": str(slot_data.get("CLASS", slot_data.get("class", "A"))).upper(),
                "DAY": str(slot_data.get("DAY", slot_data.get("day", ""))).title(),
                "START_TIME": str(slot_data.get("START_TIME", slot_data.get("start_time", ""))),
                "END_TIME": str(slot_data.get("END_TIME", slot_data.get("end_time", ""))),
                "DURATION_HOURS": float(slot_data.get("DURATION_HOURS", slot_data.get("duration_hours", 1))),
                "SLOT_TYPE": str(slot_data.get("SLOT_TYPE", slot_data.get("type", "THEORY"))).upper(),
                "COURSE": str(slot_data.get("COURSE", slot_data.get("course", ""))),
                "FACULTY_ID": str(slot_data.get("FACULTY_ID", slot_data.get("faculty_id", ""))),
                "FACULTY_NAME": str(slot_data.get("FACULTY_NAME", slot_data.get("faculty_name", ""))),
                "ROOM": str(slot_data.get("ROOM", slot_data.get("room", ""))),
                "created_at": datetime.utcnow().isoformat(),
            }
            
            # Validate required fields
            if not payload["DEPARTMENT"] or not payload["DAY"] or not payload["START_TIME"]:
                failed.append({
                    "slot_id": payload["SLOT_ID"],
                    "error": "Missing required fields: DEPARTMENT, DAY, START_TIME"
                })
                continue
            
            to_insert.append(payload)
            
        except Exception as e:
            failed.append({
                "slot_id": str(slot_data.get("SLOT_ID", f"row_{idx}")),
                "error": str(e)[:50]
            })
    
    logger.info(f"[TIMETABLE_IMPORT] Phase 1 complete: {len(to_insert)} valid, {len(failed)} invalid")
    
    # ── PHASE 2: Batch insert all slots ──
    logger.info(f"[TIMETABLE_IMPORT] Phase 2: Inserting {len(to_insert)} slots in batches of {BATCH_SIZE}...")
    
    inserted = 0
    try:
        for i in range(0, len(to_insert), BATCH_SIZE):
            batch = to_insert[i:i+BATCH_SIZE]
            batch_num = (i // BATCH_SIZE) + 1
            total_batches = (len(to_insert) + BATCH_SIZE - 1) // BATCH_SIZE
            
            try:
                result = sb.table("timetable").insert(batch).execute()
                inserted += len(batch)
                logger.info(f"[TIMETABLE_IMPORT] Batch {batch_num}/{total_batches}: +{len(batch)} slots")
            except Exception as batch_err:
                logger.error(f"[TIMETABLE_IMPORT] Batch error: {str(batch_err)[:100]}")
                # Try individual inserts
                for slot in batch:
                    try:
                        sb.table("timetable").insert(slot).execute()
                        inserted += 1
                    except:
                        failed.append({"slot_id": slot["SLOT_ID"], "error": "Insert failed"})
    except Exception as e:
        logger.error(f"[TIMETABLE_IMPORT] Import failed: {e}")
    
    logger.info(f"[TIMETABLE_IMPORT] COMPLETE: {inserted} inserted, {len(failed)} failed")
    
    return jsonify(success=True,
                   created=inserted,
                   failed=len(failed),
                   errors=failed[:50])

# ── GRIEVANCES ────────────────────────────────────────────────
@app.route("/api/grievances", methods=["GET"])
def get_grievances():
    if not sb:
        return jsonify(success=True, grievances=[])
    try:
        q = sb.table("grievances").select("*").order("created_at", desc=True)
        if request.args.get("status"):
            q = q.eq("status", request.args["status"])
        if request.args.get("student_id"):
            q = q.eq("student_id", request.args["student_id"])
        result = q.execute()
        return jsonify(success=True, grievances=result.data or [])
    except Exception as e:
        return jsonify(success=False, error=str(e)), 500

@app.route("/api/grievances", methods=["POST"])
def submit_grievance():
    if not sb:
        return jsonify(success=False, error="Supabase not configured"), 500
    d = request.json or {}
    if not all([d.get("student_id"), d.get("subject"), d.get("description")]):
        return jsonify(success=False, error="student_id, subject, description required"), 400
    try:
        payload = {
            "student_id": d["student_id"],
            "category": d.get("category", "general"),
            "subject": d["subject"],
            "description": d["description"],
            "anonymous": bool(d.get("anonymous", False)),
            "status": "open",
            "created_at": datetime.utcnow().isoformat(),
        }
        result = sb.table("grievances").insert(payload).execute()
        record = result.data[0] if result.data else payload
        rtdb_set(f"/grievances/{record.get('id', str(uuid.uuid4()))}", record)
        return jsonify(success=True, grievance=record)
    except Exception as e:
        return jsonify(success=False, error=str(e)), 500

@app.route("/api/grievances/<grv_id>", methods=["PUT"])
def update_grievance(grv_id):
    if not sb:
        return jsonify(success=False, error="Supabase not configured"), 500
    d = request.json or {}
    try:
        update_fields = {k: v for k, v in d.items() if k in ("status", "response")}
        result = sb.table("grievances").update(update_fields).eq("id", grv_id).execute()
        rtdb_update(f"/grievances/{grv_id}", update_fields)
        return jsonify(success=True, grievance=result.data[0] if result.data else {})
    except Exception as e:
        return jsonify(success=False, error=str(e)), 500

# ── MESSAGES ──────────────────────────────────────────────────
@app.route("/api/messages", methods=["GET"])
def get_messages():
    if not sb:
        return jsonify(success=True, messages=[])
    try:
        q = sb.table("messages").select("*").order("created_at", desc=True).limit(100)
        if request.args.get("recipient_id"):
            q = q.eq("recipient_id", request.args["recipient_id"])
        if request.args.get("sender_id"):
            q = q.eq("sender_id", request.args["sender_id"])
        result = q.execute()
        return jsonify(success=True, messages=result.data or [])
    except Exception as e:
        return jsonify(success=False, error=str(e)), 500

@app.route("/api/messages", methods=["POST"])
def send_message():
    if not sb:
        return jsonify(success=False, error="Supabase not configured"), 500
    d = request.json or {}
    if not all([d.get("sender_id"), d.get("recipient_id"), d.get("body")]):
        return jsonify(success=False, error="sender_id, recipient_id, body required"), 400
    try:
        payload = {
            "sender_id": d["sender_id"],
            "recipient_id": d["recipient_id"],
            "subject": d.get("subject", ""),
            "body": d["body"],
            "is_read": False,
            "created_at": datetime.utcnow().isoformat(),
        }
        result = sb.table("messages").insert(payload).execute()
        record = result.data[0] if result.data else payload
        msg_id = record.get("id", str(uuid.uuid4()))
        rtdb_set(f"/messages/{record.get('recipient_id', 'all')}/{msg_id}", record)
        return jsonify(success=True, message=record)
    except Exception as e:
        return jsonify(success=False, error=str(e)), 500

# ── ATTENDANCE SESSIONS ────────────────────────────────────────
@app.route("/api/attendance-sessions", methods=["GET"])
def get_attendance_sessions():
    if not sb:
        return jsonify(success=True, sessions=[])
    try:
        q = sb.table("attendance_sessions").select("*").order("created_at", desc=True)
        if request.args.get("date"):
            q = q.eq("date", request.args["date"])
        if request.args.get("faculty_id"):
            q = q.eq("faculty_id", request.args["faculty_id"])
        result = q.execute()
        return jsonify(success=True, sessions=result.data or [])
    except Exception as e:
        return jsonify(success=False, error=str(e)), 500

@app.route("/api/attendance-sessions", methods=["POST"])
def create_attendance_session():
    if not sb:
        return jsonify(success=False, error="Supabase not configured"), 500
    d = request.json or {}
    if not d.get("course") and not d.get("subject_name") and not d.get("timetable_id"):
        return jsonify(success=False, error="course or subject_name required"), 400
    today = datetime.utcnow().date().isoformat()

    # ── Past-date lock ────────────────────────────────────────
    sess_date = d.get("date", today)
    if sess_date < today and not d.get("admin_override"):
        return jsonify(
            success=False,
            error=f"Cannot open a session for past date {sess_date}. Sessions can only be created for today.",
            locked=True
        ), 403
    try:
        payload = {
            "course":        d.get("course", d.get("subject_name", "")),
            "session_id":    d.get("session_id", str(uuid.uuid4())),
            "timetable_id": d.get("timetable_id", ""),
            "subject_name": d.get("subject_name", d.get("course", "")),
            "faculty_id":   d.get("faculty_id", ""),
            "faculty_name": d.get("faculty_name", ""),
            "batch":         d.get("batch", ""),
            "department":    d.get("department", ""),
            "section":       d.get("section", ""),
            "semester":      d.get("semester", ""),
            "method":        d.get("method", "manual"),
            "session_type":  d.get("session_type", "lecture"),
            "status":        "open",
            "date":          d.get("date", today),
            "created_at":    datetime.utcnow().isoformat(),
        }
        result = sb.table("attendance_sessions").insert(payload).execute()
        record = result.data[0] if result.data else payload
        rtdb_set(f"/attendance_sessions/{record.get('id', str(uuid.uuid4()))}", record)
        return jsonify(success=True, session=record)
    except Exception as e:
        return jsonify(success=False, error=str(e)), 500

@app.route("/api/attendance-sessions/<session_id>", methods=["PUT"])
def update_attendance_session(session_id):
    if not sb:
        return jsonify(success=False, error="Supabase not configured"), 500
    d = request.json or {}
    try:
        allowed = ("status", "method", "course", "subject_name", "batch", "faculty_id", "faculty_name", "session_type", "department", "section", "semester")
        fields = {k: v for k, v in d.items() if k in allowed}
        result = sb.table("attendance_sessions").update(fields).eq("id", session_id).execute()
        record = result.data[0] if result.data else fields
        rtdb_update(f"/attendance_sessions/{session_id}", fields)
        return jsonify(success=True, session=record)
    except Exception as e:
        return jsonify(success=False, error=str(e)), 500

# ── ASSESSMENTS ───────────────────────────────────────────────
# ── ASSESSMENTS — FULL CRUD + TIMER + SUBMISSIONS ─────────────
@app.route("/api/assessments", methods=["GET"])
def get_assessments():
    if not sb:
        return jsonify(success=True, assessments=[])
    try:
        q = sb.table("assessments").select("*").order("created_at", desc=True)
        if request.args.get("department"):
            q = q.eq("department", request.args["department"])
        if request.args.get("section"):
            q = q.eq("section", request.args["section"])
        if request.args.get("year"):
            try:
                q = q.eq("year", int(request.args["year"]))
            except ValueError:
                pass
        if request.args.get("faculty_id"):
            q = q.eq("faculty_id", request.args["faculty_id"])
        if request.args.get("status"):
            q = q.eq("status", request.args["status"])
        if request.args.get("type"):
            q = q.eq("type", request.args["type"])
        result = q.execute()
        assessments = result.data or []
        # Auto-close expired assessments
        now = datetime.utcnow()
        for a in assessments:
            end = a.get("end_time")
            if end and a.get("status") == "active":
                try:
                    end_dt = datetime.fromisoformat(end.replace("Z", "+00:00").replace("+00:00", ""))
                    if now > end_dt:
                        sb.table("assessments").update({"status": "closed", "updated_at": now.isoformat()}).eq("id", a["id"]).execute()
                        a["status"] = "closed"
                except Exception:
                    pass
        return jsonify(success=True, assessments=assessments)
    except Exception as e:
        return jsonify(success=False, error=str(e)), 500


@app.route("/api/assessments", methods=["POST"])
def create_assessment():
    if not sb:
        return jsonify(success=False, error="DB not connected"), 500
    d = request.json or {}
    if not d.get("title"):
        return jsonify(success=False, error="title is required"), 400
    try:
        payload = {
            "title":            d["title"].strip(),
            "description":      d.get("description", ""),
            "type":             d.get("type", "quiz"),
            "course_code":      d.get("course_code", ""),
            "course_name":      d.get("course_name", ""),
            "department":       d.get("department", ""),
            "section":          d.get("section", ""),
            "year":             int(d.get("year") or 1),
            "semester":         int(d.get("semester") or 1),
            "academic_year":    d.get("academic_year", "2025-26"),
            "total_marks":      int(d.get("total_marks") or 100),
            "pass_marks":       int(d.get("pass_marks") or 40),
            "duration_mins":    int(d.get("duration_mins") or 0),
            "start_time":       d.get("start_time"),
            "end_time":         d.get("end_time"),
            "status":           d.get("status", "draft"),
            "created_by":       d.get("created_by"),
            "created_by_role":  d.get("created_by_role", "admin"),
            "faculty_id":       d.get("faculty_id"),
            "allow_late":       bool(d.get("allow_late", False)),
            "shuffle_questions": bool(d.get("shuffle_questions", False)),
            "questions":        d.get("questions", []),
            "answer_key":       d.get("answer_key", []),
            "source_file":      d.get("source_file", ""),
            "marks_format":     d.get("marks_format", []),
            "attachments":      d.get("attachments", []),
            "created_at":       datetime.utcnow().isoformat(),
            "updated_at":       datetime.utcnow().isoformat(),
        }
        result = sb.table("assessments").insert(payload).execute()
        rec = result.data[0] if result.data else payload
        rtdb_set(f"/assessments/{rec.get('id', '')}", rec)

        # ── Send notification to targeted students ─────────
        dept = payload["department"]
        section = payload["section"]
        notif_title = f"📋 New {payload['type'].title()}: {payload['title']}"
        notif_msg = f"{payload['course_code']} — {payload['title']}"
        if payload.get("end_time"):
            notif_msg += f" | Due: {payload['end_time'][:16].replace('T',' ')}"
        try:
            # Get students matching department + section
            sq = sb.table("users").select("id").eq("role", "student").eq("is_active", True)
            if dept:
                sq = sq.eq("department", dept)
            if section:
                sq = sq.eq("section", section)
            students = sq.execute().data or []
            notif_rows = []
            for s in students:
                notif_rows.append({
                    "user_id":           s["id"],
                    "target_role":       "student",
                    "title":             notif_title,
                    "message":           notif_msg,
                    "notification_type": "assignment",
                    "is_read":           False,
                    "sent_by":           payload.get("created_by_role", "admin"),
                })
            if notif_rows:
                sb.table("notifications").insert(notif_rows).execute()
        except Exception:
            pass  # notifications are best-effort

        return jsonify(success=True, assessment=rec)
    except Exception as e:
        return jsonify(success=False, error=str(e)), 500


@app.route("/api/assessments/<asmt_id>", methods=["GET"])
def get_assessment_detail(asmt_id):
    if not sb:
        return jsonify(success=False, error="DB not connected"), 500
    try:
        rows = sb.table("assessments").select("*").eq("id", asmt_id).execute().data or []
        if not rows:
            return jsonify(success=False, error="Assessment not found"), 404
        # Auto-close if expired
        a = rows[0]
        now = datetime.utcnow()
        end = a.get("end_time")
        if end and a.get("status") == "active":
            try:
                end_dt = datetime.fromisoformat(end.replace("Z", "+00:00").replace("+00:00", ""))
                if now > end_dt:
                    sb.table("assessments").update({"status": "closed"}).eq("id", asmt_id).execute()
                    a["status"] = "closed"
            except Exception:
                pass
        # Get submission stats
        subs = sb.table("assessment_submissions").select("id,status,score").eq("assessment_id", asmt_id).execute().data or []
        a["submission_count"] = len(subs)
        a["submitted_count"] = len([s for s in subs if s.get("status") in ("submitted", "graded")])
        a["graded_count"] = len([s for s in subs if s.get("status") == "graded"])
        return jsonify(success=True, assessment=a)
    except Exception as e:
        return jsonify(success=False, error=str(e)), 500


@app.route("/api/assessments/<asmt_id>", methods=["PUT"])
def update_assessment(asmt_id):
    if not sb:
        return jsonify(success=False, error="DB not connected"), 500
    d = request.json or {}
    try:
        allowed = ("title", "description", "type", "course_code", "course_name",
                   "department", "section", "year", "semester", "academic_year",
                   "total_marks", "pass_marks", "duration_mins", "start_time", "end_time",
                   "status", "faculty_id", "allow_late", "shuffle_questions",
                   "questions", "answer_key", "source_file", "marks_format", "attachments")
        fields = {k: v for k, v in d.items() if k in allowed}
        for int_key in ("year", "semester", "total_marks", "pass_marks", "duration_mins"):
            if int_key in fields and fields[int_key] is not None:
                fields[int_key] = int(fields[int_key])
        fields["updated_at"] = datetime.utcnow().isoformat()
        result = sb.table("assessments").update(fields).eq("id", asmt_id).execute()
        rec = result.data[0] if result.data else fields
        rtdb_update(f"/assessments/{asmt_id}", fields)
        return jsonify(success=True, assessment=rec)
    except Exception as e:
        return jsonify(success=False, error=str(e)), 500


@app.route("/api/assessments/<asmt_id>", methods=["DELETE"])
def delete_assessment(asmt_id):
    if not sb:
        return jsonify(success=False, error="DB not connected"), 500
    try:
        sb.table("assessments").delete().eq("id", asmt_id).execute()
        rtdb_delete(f"/assessments/{asmt_id}")
        return jsonify(success=True)
    except Exception as e:
        return jsonify(success=False, error=str(e)), 500


# ── ASSESSMENT SUBMISSIONS ─────────────────────────────────────
@app.route("/api/assessments/<asmt_id>/submissions", methods=["GET"])
def get_submissions(asmt_id):
    if not sb:
        return jsonify(success=True, submissions=[])
    try:
        q = sb.table("assessment_submissions").select("*").eq("assessment_id", asmt_id).order("submitted_at", desc=True)
        result = q.execute()
        return jsonify(success=True, submissions=result.data or [])
    except Exception as e:
        return jsonify(success=False, error=str(e)), 500


@app.route("/api/assessments/<asmt_id>/submit", methods=["POST"])
def submit_assessment(asmt_id):
    """Student starts or submits an assessment."""
    if not sb:
        return jsonify(success=False, error="DB not connected"), 500
    d = request.json or {}
    student_id = d.get("student_id", "")
    if not student_id:
        return jsonify(success=False, error="student_id required"), 400
    try:
        # Check assessment exists and is active
        asmt_rows = sb.table("assessments").select("*").eq("id", asmt_id).execute().data or []
        if not asmt_rows:
            return jsonify(success=False, error="Assessment not found"), 404
        asmt = asmt_rows[0]

        # Auto-close if expired
        now = datetime.utcnow()
        end = asmt.get("end_time")
        if end and asmt.get("status") == "active":
            try:
                end_dt = datetime.fromisoformat(end.replace("Z", "+00:00").replace("+00:00", ""))
                if now > end_dt:
                    sb.table("assessments").update({"status": "closed"}).eq("id", asmt_id).execute()
                    asmt["status"] = "closed"
            except Exception:
                pass

        if asmt["status"] != "active":
            return jsonify(success=False, error=f"Assessment is {asmt['status']}. Cannot submit."), 403

        # Check section match
        stu_rows = sb.table("users").select("department,section,roll_no,full_name").eq("id", student_id).execute().data or []
        stu = stu_rows[0] if stu_rows else {}
        if asmt.get("section") and stu.get("section") and stu["section"] != asmt["section"]:
            return jsonify(success=False, error="You are not in the targeted section for this assessment."), 403
        if asmt.get("department") and stu.get("department") and stu["department"] != asmt["department"]:
            return jsonify(success=False, error="You are not in the targeted department for this assessment."), 403

        # Check for existing submission
        existing = sb.table("assessment_submissions").select("id,status") \
            .eq("assessment_id", asmt_id).eq("student_id", student_id).execute().data or []

        is_final = d.get("final", False)
        answers = d.get("answers", [])

        if existing:
            sub = existing[0]
            if sub["status"] in ("submitted", "graded"):
                return jsonify(success=False, error="You have already submitted this assessment."), 409
            # Update in-progress submission
            upd = {"answers": answers, "updated_at": now.isoformat()}
            if is_final:
                # Auto-grade quiz
                score = _auto_grade(asmt.get("questions", []), answers)
                upd.update({"status": "submitted", "submitted_at": now.isoformat(),
                            "score": score, "total_marks": asmt.get("total_marks", 0)})
            result = sb.table("assessment_submissions").update(upd).eq("id", sub["id"]).execute()
            rec = result.data[0] if result.data else upd
            return jsonify(success=True, submission=rec, updated=True)
        else:
            # Create new submission
            payload = {
                "assessment_id": asmt_id,
                "student_id":    student_id,
                "roll_no":       stu.get("roll_no", d.get("roll_no", "")),
                "student_name":  stu.get("full_name", d.get("student_name", "")),
                "department":    stu.get("department", ""),
                "section":       stu.get("section", ""),
                "answers":       answers,
                "status":        "submitted" if is_final else "in_progress",
                "started_at":    now.isoformat(),
                "created_at":    now.isoformat(),
            }
            if is_final:
                score = _auto_grade(asmt.get("questions", []), answers)
                payload.update({"submitted_at": now.isoformat(), "score": score,
                                "total_marks": asmt.get("total_marks", 0)})
            result = sb.table("assessment_submissions").insert(payload).execute()
            rec = result.data[0] if result.data else payload
            return jsonify(success=True, submission=rec, created=True)
    except Exception as e:
        return jsonify(success=False, error=str(e)), 500


@app.route("/api/assessments/<asmt_id>/grade", methods=["PUT"])
def grade_submission(asmt_id):
    """Faculty grades a submission."""
    if not sb:
        return jsonify(success=False, error="DB not connected"), 500
    d = request.json or {}
    sub_id = d.get("submission_id")
    if not sub_id:
        return jsonify(success=False, error="submission_id required"), 400
    try:
        fields = {
            "score":     float(d.get("score", 0)),
            "feedback":  d.get("feedback", ""),
            "status":    "graded",
            "graded_at": datetime.utcnow().isoformat(),
            "graded_by": d.get("graded_by"),
        }
        result = sb.table("assessment_submissions").update(fields).eq("id", sub_id).execute()
        rec = result.data[0] if result.data else fields
        return jsonify(success=True, submission=rec)
    except Exception as e:
        return jsonify(success=False, error=str(e)), 500


def _auto_grade(questions, answers):
    """Auto-grade MCQ/quiz answers against question list."""
    if not questions or not answers:
        return 0
    score = 0
    for ans in answers:
        qi = ans.get("question_index")
        if qi is not None and qi < len(questions):
            q = questions[qi]
            if q.get("type") in ("mcq", "true_false"):
                if str(ans.get("answer", "")).strip().lower() == str(q.get("correct", "")).strip().lower():
                    score += int(q.get("marks", 1))
    return score


@app.route("/api/assessments/<asmt_id>/grade-detailed", methods=["PUT"])
def grade_submission_detailed(asmt_id):
    """Faculty grades each question individually for a submission."""
    if not sb:
        return jsonify(success=False, error="DB not connected"), 500
    d = request.json or {}
    sub_id = d.get("submission_id")
    if not sub_id:
        return jsonify(success=False, error="submission_id required"), 400
    try:
        question_scores = d.get("question_scores", [])
        total_score = sum(float(qs.get("score", 0)) for qs in question_scores)
        fields = {
            "question_scores": question_scores,
            "score":           total_score,
            "feedback":        d.get("feedback", ""),
            "status":          "graded",
            "graded_at":       datetime.utcnow().isoformat(),
            "graded_by":       d.get("graded_by"),
        }
        result = sb.table("assessment_submissions").update(fields).eq("id", sub_id).execute()
        rec = result.data[0] if result.data else fields
        return jsonify(success=True, submission=rec)
    except Exception as e:
        return jsonify(success=False, error=str(e)), 500


@app.route("/api/assessments/<asmt_id>/download-marks", methods=["GET"])
def download_marks(asmt_id):
    """Download marks as CSV for an assessment."""
    if not sb:
        return jsonify(success=False, error="DB not connected"), 500
    try:
        asmt_rows = sb.table("assessments").select("title,total_marks,questions").eq("id", asmt_id).execute().data or []
        if not asmt_rows:
            return jsonify(success=False, error="Assessment not found"), 404
        asmt = asmt_rows[0]
        questions = asmt.get("questions") or []
        if isinstance(questions, str):
            questions = json.loads(questions)

        subs = sb.table("assessment_submissions").select("*").eq("assessment_id", asmt_id).order("roll_no").execute().data or []

        import io, csv
        output = io.StringIO()
        writer = csv.writer(output)
        # Header
        header = ["Roll No", "Student Name", "Department", "Section"]
        for i, q in enumerate(questions):
            header.append(f"Q{i+1} ({q.get('marks', '')}m)")
        header.extend(["Total Score", f"Out of {asmt.get('total_marks', '')}", "Status", "Feedback"])
        writer.writerow(header)
        # Rows
        for s in subs:
            row = [s.get("roll_no", ""), s.get("student_name", ""), s.get("department", ""), s.get("section", "")]
            q_scores = s.get("question_scores") or []
            if isinstance(q_scores, str):
                q_scores = json.loads(q_scores)
            score_map = {qs.get("question_index", i): qs.get("score", "") for i, qs in enumerate(q_scores)}
            for i in range(len(questions)):
                row.append(score_map.get(i, ""))
            row.extend([s.get("score", ""), asmt.get("total_marks", ""), s.get("status", ""), s.get("feedback", "")])
            writer.writerow(row)

        csv_content = output.getvalue()
        output.close()
        return app.response_class(
            csv_content,
            mimetype="text/csv",
            headers={"Content-Disposition": f"attachment; filename=marks_{asmt_id[:8]}.csv"}
        )
    except Exception as e:
        return jsonify(success=False, error=str(e)), 500


# ── EXAM MARKS FORMAT (admin-configurable) ──────────────────
@app.route("/api/exam-marks-format", methods=["GET"])
def get_exam_marks_format():
    if not sb:
        return jsonify(success=True, formats=[])
    try:
        q = sb.table("exam_marks_format").select("*").order("created_at", desc=True)
        if request.args.get("exam_type"):
            q = q.eq("exam_type", request.args["exam_type"])
        if request.args.get("department"):
            q = q.eq("department", request.args["department"])
        if request.args.get("is_active"):
            q = q.eq("is_active", request.args["is_active"] == "true")
        result = q.execute()
        return jsonify(success=True, formats=result.data or [])
    except Exception as e:
        return jsonify(success=False, error=str(e)), 500


@app.route("/api/exam-marks-format", methods=["POST"])
def create_exam_marks_format():
    if not sb:
        return jsonify(success=False, error="DB not connected"), 500
    d = request.json or {}
    if not d.get("title"):
        return jsonify(success=False, error="title required"), 400
    try:
        record = {
            "exam_type":   d.get("exam_type", "midterm"),
            "title":       d["title"],
            "course_code": d.get("course_code", ""),
            "department":  d.get("department", ""),
            "sections":    d.get("sections", []),
            "total_marks": int(d.get("total_marks", 100)),
            "created_by":  d.get("created_by"),
            "is_active":   d.get("is_active", True),
        }
        result = sb.table("exam_marks_format").insert(record).execute()
        row = result.data[0] if result.data else record
        return jsonify(success=True, format=row)
    except Exception as e:
        return jsonify(success=False, error=str(e)), 500


@app.route("/api/exam-marks-format/<fmt_id>", methods=["PUT"])
def update_exam_marks_format(fmt_id):
    if not sb:
        return jsonify(success=False, error="DB not connected"), 500
    d = request.json or {}
    try:
        allowed = {"exam_type", "title", "course_code", "department", "sections", "total_marks", "is_active"}
        fields = {k: v for k, v in d.items() if k in allowed}
        fields["updated_at"] = datetime.utcnow().isoformat()
        result = sb.table("exam_marks_format").update(fields).eq("id", fmt_id).execute()
        row = result.data[0] if result.data else fields
        return jsonify(success=True, format=row)
    except Exception as e:
        return jsonify(success=False, error=str(e)), 500


@app.route("/api/exam-marks-format/<fmt_id>", methods=["DELETE"])
def delete_exam_marks_format(fmt_id):
    if not sb:
        return jsonify(success=False, error="DB not connected"), 500
    try:
        sb.table("exam_marks_format").delete().eq("id", fmt_id).execute()
        return jsonify(success=True)
    except Exception as e:
        return jsonify(success=False, error=str(e)), 500


# ── PLACEMENT OPPORTUNITIES ─────────────────────────────────
@app.route("/api/placements", methods=["GET"])
def get_placements():
    if not sb:
        return jsonify(success=True, placements=[])
    try:
        q = sb.table("placement_opportunities").select("*").order("created_at", desc=True)
        if request.args.get("active_only") != "false":
            q = q.eq("is_active", True)
        result = q.execute()
        return jsonify(success=True, placements=result.data or [])
    except Exception as e:
        return jsonify(success=False, error=str(e)), 500

@app.route("/api/placements", methods=["POST"])
def create_placement():
    if not sb:
        return jsonify(success=False, error="DB not connected"), 500
    d = request.get_json() or {}
    if not d.get("company_name") or not d.get("role"):
        return jsonify(success=False, error="company_name and role required"), 400
    try:
        record = {
            "company_name": d["company_name"], "role": d["role"],
            "package": d.get("package", ""), "deadline": d.get("deadline"),
            "eligibility_criteria": d.get("eligibility_criteria", ""),
            "description": d.get("description", ""), "apply_link": d.get("apply_link", ""),
            "is_active": d.get("is_active", True), "posted_by": d.get("posted_by", "admin"),
        }
        result = sb.table("placement_opportunities").insert(record).execute()
        row = result.data[0] if result.data else record
        rtdb_set(f"/placement_opportunities/{row.get('id', 'new')}", row)
        return jsonify(success=True, placement=row)
    except Exception as e:
        return jsonify(success=False, error=str(e)), 500

@app.route("/api/placements/<pid>", methods=["PUT"])
def update_placement(pid):
    if not sb:
        return jsonify(success=False, error="DB not connected"), 500
    d = request.get_json() or {}
    try:
        result = sb.table("placement_opportunities").update(d).eq("id", pid).execute()
        row = result.data[0] if result.data else {}
        rtdb_update(f"/placement_opportunities/{pid}", d)
        return jsonify(success=True, placement=row)
    except Exception as e:
        return jsonify(success=False, error=str(e)), 500

@app.route("/api/placements/<pid>", methods=["DELETE"])
def delete_placement(pid):
    if not sb:
        return jsonify(success=False, error="DB not connected"), 500
    try:
        sb.table("placement_opportunities").delete().eq("id", pid).execute()
        rtdb_delete(f"/placement_opportunities/{pid}")
        return jsonify(success=True)
    except Exception as e:
        return jsonify(success=False, error=str(e)), 500


# ── EXAM REGISTRATIONS ───────────────────────────────────────
@app.route("/api/exam-registrations", methods=["GET"])
def get_exam_registrations():
    if not sb:
        return jsonify(success=True, registrations=[])
    try:
        q = sb.table("exam_registrations").select("*").order("created_at", desc=True)
        if request.args.get("student_id"):
            q = q.eq("student_id", request.args["student_id"])
        result = q.execute()
        return jsonify(success=True, registrations=result.data or [])
    except Exception as e:
        return jsonify(success=False, error=str(e)), 500

@app.route("/api/exam-registrations", methods=["POST"])
def create_exam_registration():
    if not sb:
        return jsonify(success=False, error="DB not connected"), 500
    d = request.get_json() or {}
    if not d.get("student_id"):
        return jsonify(success=False, error="student_id required"), 400
    try:
        record = {
            "student_id": d["student_id"], "semester": d.get("semester", ""),
            "academic_year": d.get("academic_year", ""),
            "subjects": d.get("subjects", []), "status": "pending",
        }
        result = sb.table("exam_registrations").insert(record).execute()
        row = result.data[0] if result.data else record
        rtdb_set(f"/exam_registrations/{row.get('id', 'new')}", row)
        return jsonify(success=True, registration=row)
    except Exception as e:
        return jsonify(success=False, error=str(e)), 500


# ── LIBRARY RESOURCES ────────────────────────────────────────
@app.route("/api/library", methods=["GET"])
def get_library():
    if not sb:
        return jsonify(success=True, resources=[])
    try:
        q = sb.table("library_resources").select("*").order("created_at", desc=True)
        if request.args.get("category"):
            q = q.eq("category", request.args["category"])
        if request.args.get("resource_type"):
            q = q.eq("resource_type", request.args["resource_type"])
        result = q.execute()
        return jsonify(success=True, resources=result.data or [])
    except Exception as e:
        return jsonify(success=False, error=str(e)), 500

@app.route("/api/library", methods=["POST"])
def create_library_resource():
    if not sb:
        return jsonify(success=False, error="DB not connected"), 500
    d = request.get_json() or {}
    if not d.get("title"):
        return jsonify(success=False, error="title required"), 400
    try:
        record = {
            "title": d["title"], "author": d.get("author", ""),
            "subject": d.get("subject", ""), "resource_type": d.get("resource_type", "E-Book"),
            "pdf_link": d.get("pdf_link", ""), "available_copies": d.get("available_copies", 1),
            "category": d.get("category", ""), "description": d.get("description", ""),
            "added_by": d.get("added_by", "admin"),
        }
        result = sb.table("library_resources").insert(record).execute()
        row = result.data[0] if result.data else record
        rtdb_set(f"/library_resources/{row.get('id', 'new')}", row)
        return jsonify(success=True, resource=row)
    except Exception as e:
        return jsonify(success=False, error=str(e)), 500

@app.route("/api/library/<rid>", methods=["DELETE"])
def delete_library_resource(rid):
    if not sb:
        return jsonify(success=False, error="DB not connected"), 500
    try:
        sb.table("library_resources").delete().eq("id", rid).execute()
        rtdb_delete(f"/library_resources/{rid}")
        return jsonify(success=True)
    except Exception as e:
        return jsonify(success=False, error=str(e)), 500


# ── COURSE MATERIALS ─────────────────────────────────────────
@app.route("/api/course-materials", methods=["GET"])
def get_course_materials():
    if not sb:
        return jsonify(success=True, materials=[])
    try:
        q = sb.table("course_materials").select("*").order("created_at", desc=True)
        if request.args.get("subject"):
            q = q.eq("subject", request.args["subject"])
        if request.args.get("uploaded_by"):
            q = q.eq("uploaded_by", request.args["uploaded_by"])
        if request.args.get("course_code"):
            q = q.eq("course_code", request.args["course_code"])
        if request.args.get("public_only") == "true":
            q = q.eq("is_public", True)
        result = q.execute()
        return jsonify(success=True, materials=result.data or [])
    except Exception as e:
        return jsonify(success=False, error=str(e)), 500

@app.route("/api/course-materials", methods=["POST"])
def create_course_material():
    if not sb:
        return jsonify(success=False, error="DB not connected"), 500
    d = request.get_json() or {}
    if not d.get("file_name"):
        return jsonify(success=False, error="file_name required"), 400
    try:
        record = {
            "subject": d.get("subject", ""), "file_name": d["file_name"],
            "file_url": d.get("file_url", ""), "material_type": d.get("material_type", "Notes"),
            "is_public": d.get("is_public", True), "uploaded_by": d.get("uploaded_by", ""),
            "topic": d.get("topic", ""),
            "course_code": d.get("course_code", ""),
            "module_number": d.get("module_number", 0),
            "unit_name": d.get("unit_name", ""),
            "description": d.get("description", ""),
            "department": d.get("department", ""),
        }
        result = sb.table("course_materials").insert(record).execute()
        row = result.data[0] if result.data else record
        rtdb_set(f"/course_materials/{row.get('id', 'new')}", row)
        return jsonify(success=True, material=row)
    except Exception as e:
        return jsonify(success=False, error=str(e)), 500

@app.route("/api/course-materials/<mid>", methods=["PUT"])
def update_course_material(mid):
    if not sb:
        return jsonify(success=False, error="DB not connected"), 500
    d = request.get_json() or {}
    try:
        result = sb.table("course_materials").update(d).eq("id", mid).execute()
        row = result.data[0] if result.data else {}
        rtdb_update(f"/course_materials/{mid}", d)
        return jsonify(success=True, material=row)
    except Exception as e:
        return jsonify(success=False, error=str(e)), 500

@app.route("/api/course-materials/<mid>", methods=["DELETE"])
def delete_course_material(mid):
    if not sb:
        return jsonify(success=False, error="DB not connected"), 500
    try:
        sb.table("course_materials").delete().eq("id", mid).execute()
        rtdb_delete(f"/course_materials/{mid}")
        return jsonify(success=True)
    except Exception as e:
        return jsonify(success=False, error=str(e)), 500


# ── QUESTION PAPERS ──────────────────────────────────────────
@app.route("/api/question-papers", methods=["GET"])
def get_question_papers():
    if not sb:
        return jsonify(success=True, papers=[])
    try:
        q = sb.table("question_papers").select("*").order("created_at", desc=True)
        if request.args.get("course_code"):
            q = q.eq("course_code", request.args["course_code"])
        if request.args.get("generated_by"):
            q = q.eq("generated_by", request.args["generated_by"])
        if request.args.get("department"):
            q = q.eq("department", request.args["department"])
        if request.args.get("status"):
            q = q.eq("status", request.args["status"])
        if request.args.get("is_selected"):
            q = q.eq("is_selected", request.args["is_selected"] == "true")
        result = q.execute()
        return jsonify(success=True, papers=result.data or [])
    except Exception as e:
        return jsonify(success=False, error=str(e)), 500


@app.route("/api/question-papers", methods=["POST"])
def create_question_paper():
    """Create / save a question paper."""
    if not sb:
        return jsonify(success=False, error="DB not connected"), 500
    d = request.get_json() or {}
    if not d.get("title"):
        return jsonify(success=False, error="title required"), 400
    try:
        record = {
            "title": d["title"],
            "course_code": d.get("course_code", ""),
            "course_name": d.get("course_name", ""),
            "department": d.get("department", ""),
            "semester": d.get("semester", 1),
            "year": d.get("year", 1),
            "academic_year": d.get("academic_year", ""),
            "exam_type": d.get("exam_type", "internal"),
            "total_marks": d.get("total_marks", 100),
            "duration_mins": d.get("duration_mins", 180),
            "sections": json.dumps(d.get("sections", [])),
            "modules_used": json.dumps(d.get("modules_used", [])),
            "instructions": d.get("instructions", ""),
            "generated_by": d.get("generated_by", ""),
            "faculty_name": d.get("faculty_name", ""),
            "subject_code": d.get("subject_code", ""),
            "status": d.get("status", "draft"),
            "is_selected": False,
        }
        result = sb.table("question_papers").insert(record).execute()
        row = result.data[0] if result.data else record
        rtdb_set(f"/question_papers/{row.get('id', 'new')}", row)
        return jsonify(success=True, paper=row)
    except Exception as e:
        return jsonify(success=False, error=str(e)), 500


@app.route("/api/question-papers/<qp_id>", methods=["PUT"])
def update_question_paper(qp_id):
    if not sb:
        return jsonify(success=False, error="DB not connected"), 500
    d = request.get_json() or {}
    try:
        allowed = {"title","course_code","course_name","department","semester","year",
                   "academic_year","exam_type","total_marks","duration_mins",
                   "sections","modules_used","instructions","status",
                   "faculty_name","subject_code","is_selected","selected_by_admin","selected_at"}
        updates = {k: v for k, v in d.items() if k in allowed}
        if "sections" in updates:
            updates["sections"] = json.dumps(updates["sections"])
        if "modules_used" in updates:
            updates["modules_used"] = json.dumps(updates["modules_used"])
        result = sb.table("question_papers").update(updates).eq("id", qp_id).execute()
        row = result.data[0] if result.data else {}
        rtdb_update(f"/question_papers/{qp_id}", updates)
        return jsonify(success=True, paper=row)
    except Exception as e:
        return jsonify(success=False, error=str(e)), 500


@app.route("/api/question-papers/<qp_id>", methods=["DELETE"])
def delete_question_paper(qp_id):
    if not sb:
        return jsonify(success=False, error="DB not connected"), 500
    try:
        sb.table("question_papers").delete().eq("id", qp_id).execute()
        rtdb_delete(f"/question_papers/{qp_id}")
        return jsonify(success=True)
    except Exception as e:
        return jsonify(success=False, error=str(e)), 500


@app.route("/api/question-papers/<qp_id>/select", methods=["PUT"])
def select_question_paper(qp_id):
    """Admin selects one question paper for a course — deselects others for same course+exam_type."""
    if not sb:
        return jsonify(success=False, error="DB not connected"), 500
    d = request.json or {}
    try:
        # Get the paper being selected
        rows = sb.table("question_papers").select("course_code,exam_type").eq("id", qp_id).execute().data or []
        if not rows:
            return jsonify(success=False, error="Paper not found"), 404
        paper = rows[0]
        course_code = paper.get("course_code", "")
        exam_type = paper.get("exam_type", "")

        # Deselect all other papers for the same course + exam type
        if course_code:
            others = sb.table("question_papers").select("id").eq("course_code", course_code).eq("exam_type", exam_type).neq("id", qp_id).execute().data or []
            for o in others:
                sb.table("question_papers").update({
                    "is_selected": False,
                    "selected_by_admin": None,
                    "selected_at": None,
                    "status": "submitted"
                }).eq("id", o["id"]).execute()

        # Select this paper
        now = datetime.utcnow().isoformat()
        updates = {
            "is_selected": True,
            "selected_by_admin": d.get("admin_id"),
            "selected_at": now,
            "status": "selected"
        }
        result = sb.table("question_papers").update(updates).eq("id", qp_id).execute()
        rec = result.data[0] if result.data else updates
        return jsonify(success=True, paper=rec)
    except Exception as e:
        return jsonify(success=False, error=str(e)), 500


@app.route("/api/question-papers/generate", methods=["POST"])
def generate_question_paper():
    """Auto-generate a question paper from uploaded course-material modules."""
    if not sb:
        return jsonify(success=False, error="DB not connected"), 500
    d = request.get_json() or {}
    course_code = d.get("course_code", "")
    if not course_code:
        return jsonify(success=False, error="course_code required"), 400
    modules_requested = d.get("modules", [])  # list of module numbers
    total_marks = d.get("total_marks", 100)
    exam_type = d.get("exam_type", "internal")
    try:
        # Fetch all materials for this course
        q = sb.table("course_materials").select("*").eq("course_code", course_code)
        result = q.execute()
        materials = result.data or []
        if not materials:
            return jsonify(success=False, error="No modules uploaded for this course. Upload materials first."), 404

        # Group by module number
        module_map = {}
        for m in materials:
            mn = m.get("module_number", 0)
            if modules_requested and mn not in modules_requested:
                continue
            module_map.setdefault(mn, []).append(m)

        if not module_map:
            return jsonify(success=False, error="No matching modules found"), 404

        num_modules = len(module_map)
        marks_per_module = total_marks // max(num_modules, 1)

        # Build sections — one section per module
        sections = []
        for mn in sorted(module_map.keys()):
            mats = module_map[mn]
            unit = mats[0].get("unit_name", "") or mats[0].get("topic", "") or f"Module {mn}"
            topics = list({m.get("topic", "") or m.get("unit_name", "") for m in mats if m.get("topic") or m.get("unit_name")})
            # Generate placeholder questions from topics
            questions = []
            if len(topics) == 0:
                topics = [f"Module {mn} Content"]
            q_marks = max(1, marks_per_module // max(len(topics), 1))
            for ti, topic in enumerate(topics):
                questions.append({
                    "q": f"Explain the key concepts of {topic}.",
                    "marks": q_marks,
                    "module": mn,
                    "type": "descriptive",
                    "topic": topic,
                })
            # Adjust last question marks to make up rounding difference
            if questions:
                used = sum(qq["marks"] for qq in questions)
                questions[-1]["marks"] += (marks_per_module - used)

            sections.append({
                "title": f"Module {mn}: {unit}",
                "marks": marks_per_module,
                "questions": questions,
            })

        # Adjust last section marks for rounding
        if sections:
            assigned = sum(s["marks"] for s in sections)
            sections[-1]["marks"] += (total_marks - assigned)

        return jsonify(success=True, sections=sections, modules_used=sorted(module_map.keys()),
                       total_marks=total_marks, course_code=course_code)
    except Exception as e:
        return jsonify(success=False, error=str(e)), 500


# ── CALENDAR EVENTS ──────────────────────────────────────────
@app.route("/api/calendar-events", methods=["GET"])
def get_calendar_events():
    if not sb:
        return jsonify(success=True, events=[])
    try:
        q = sb.table("calendar_events").select("*").order("event_date", desc=False)
        if request.args.get("from_date"):
            q = q.gte("event_date", request.args["from_date"])
        result = q.execute()
        return jsonify(success=True, events=result.data or [])
    except Exception as e:
        return jsonify(success=False, error=str(e)), 500

@app.route("/api/calendar-events", methods=["POST"])
def create_calendar_event():
    if not sb:
        return jsonify(success=False, error="DB not connected"), 500
    d = request.get_json() or {}
    if not d.get("title") or not d.get("event_date"):
        return jsonify(success=False, error="title and event_date required"), 400
    try:
        record = {
            "title": d["title"], "event_date": d["event_date"],
            "end_date": d.get("end_date"), "event_type": d.get("event_type", "event"),
            "description": d.get("description", ""),
            "target_audience": d.get("target_audience", "all"),
            "created_by": d.get("created_by", "admin"),
        }
        result = sb.table("calendar_events").insert(record).execute()
        row = result.data[0] if result.data else record
        rtdb_set(f"/calendar_events/{row.get('id', 'new')}", row)
        return jsonify(success=True, event=row)
    except Exception as e:
        return jsonify(success=False, error=str(e)), 500

@app.route("/api/calendar-events/<eid>", methods=["DELETE"])
def delete_calendar_event(eid):
    if not sb:
        return jsonify(success=False, error="DB not connected"), 500
    try:
        sb.table("calendar_events").delete().eq("id", eid).execute()
        rtdb_delete(f"/calendar_events/{eid}")
        return jsonify(success=True)
    except Exception as e:
        return jsonify(success=False, error=str(e)), 500


# ── COMMUNITIES ──────────────────────────────────────────────
@app.route("/api/communities", methods=["GET"])
def get_communities():
    if not sb:
        return jsonify(success=True, communities=[])
    try:
        result = sb.table("communities").select("*").order("created_at", desc=False).execute()
        return jsonify(success=True, communities=result.data or [])
    except Exception as e:
        return jsonify(success=False, error=str(e)), 500

@app.route("/api/communities", methods=["POST"])
def create_community():
    if not sb:
        return jsonify(success=False, error="DB not connected"), 500
    d = request.get_json() or {}
    if not d.get("name"):
        return jsonify(success=False, error="name required"), 400
    try:
        record = {
            "course_code": d.get("course_code", ""), "name": d["name"],
            "description": d.get("description", ""), "members_count": 0,
            "created_by": d.get("created_by", "admin"),
        }
        result = sb.table("communities").insert(record).execute()
        row = result.data[0] if result.data else record
        rtdb_set(f"/communities/{row.get('id', 'new')}", row)
        return jsonify(success=True, community=row)
    except Exception as e:
        return jsonify(success=False, error=str(e)), 500

@app.route("/api/communities/<cid>", methods=["DELETE"])
def delete_community(cid):
    if not sb:
        return jsonify(success=False, error="DB not connected"), 500
    try:
        sb.table("communities").delete().eq("id", cid).execute()
        rtdb_delete(f"/communities/{cid}")
        return jsonify(success=True)
    except Exception as e:
        return jsonify(success=False, error=str(e)), 500


# ── COMMUNITY POSTS ──────────────────────────────────────────
@app.route("/api/community-posts", methods=["GET"])
def get_community_posts():
    if not sb:
        return jsonify(success=True, posts=[])
    try:
        q = sb.table("community_posts").select("*").order("created_at", desc=True)
        if request.args.get("community_id"):
            q = q.eq("community_id", request.args["community_id"])
        result = q.execute()
        return jsonify(success=True, posts=result.data or [])
    except Exception as e:
        return jsonify(success=False, error=str(e)), 500

@app.route("/api/community-posts", methods=["POST"])
def create_community_post():
    if not sb:
        return jsonify(success=False, error="DB not connected"), 500
    d = request.get_json() or {}
    if not d.get("community_id") or not d.get("content"):
        return jsonify(success=False, error="community_id and content required"), 400
    try:
        record = {
            "community_id": d["community_id"], "content": d["content"],
            "author_id": d.get("author_id", ""), "author_name": d.get("author_name", "Anonymous"),
            "post_type": d.get("post_type", "discussion"), "likes": 0,
        }
        result = sb.table("community_posts").insert(record).execute()
        row = result.data[0] if result.data else record
        rtdb_set(f"/community_posts/{row.get('id', 'new')}", row)
        return jsonify(success=True, post=row)
    except Exception as e:
        return jsonify(success=False, error=str(e)), 500


# ── STAFF EVALUATIONS ────────────────────────────────────────
@app.route("/api/staff-evaluations", methods=["GET"])
def get_staff_evaluations():
    if not sb:
        return jsonify(success=True, evaluations=[])
    try:
        q = sb.table("staff_evaluations").select("*").order("submitted_at", desc=True)
        if request.args.get("student_id"):
            q = q.eq("student_id", request.args["student_id"])
        if request.args.get("faculty_id"):
            q = q.eq("faculty_id", request.args["faculty_id"])
        result = q.execute()
        return jsonify(success=True, evaluations=result.data or [])
    except Exception as e:
        return jsonify(success=False, error=str(e)), 500

@app.route("/api/staff-evaluations", methods=["POST"])
def create_staff_evaluation():
    if not sb:
        return jsonify(success=False, error="DB not connected"), 500
    d = request.get_json() or {}
    if not d.get("student_id") or not d.get("faculty_id"):
        return jsonify(success=False, error="student_id and faculty_id required"), 400
    try:
        record = {
            "student_id": d["student_id"], "faculty_id": d["faculty_id"],
            "faculty_name": d.get("faculty_name", ""), "subject": d.get("subject", ""),
            "teaching_clarity": d.get("teaching_clarity"), "subject_knowledge": d.get("subject_knowledge"),
            "overall": d.get("overall"), "comments": d.get("comments", ""),
            "semester": d.get("semester", ""), "academic_year": d.get("academic_year", ""),
        }
        result = sb.table("staff_evaluations").insert(record).execute()
        row = result.data[0] if result.data else record
        rtdb_set(f"/staff_evaluations/{row.get('id', 'new')}", row)
        return jsonify(success=True, evaluation=row)
    except Exception as e:
        return jsonify(success=False, error=str(e)), 500


# ── LESSON PLANS ─────────────────────────────────────────────
@app.route("/api/lesson-plans", methods=["GET"])
def get_lesson_plans():
    if not sb:
        return jsonify(success=True, plans=[])
    try:
        q = sb.table("lesson_plans").select("*").order("week_number", desc=False)
        if request.args.get("faculty_id"):
            q = q.eq("faculty_id", request.args["faculty_id"])
        if request.args.get("subject"):
            q = q.eq("subject", request.args["subject"])
        result = q.execute()
        return jsonify(success=True, plans=result.data or [])
    except Exception as e:
        return jsonify(success=False, error=str(e)), 500

@app.route("/api/lesson-plans", methods=["POST"])
def create_lesson_plan():
    if not sb:
        return jsonify(success=False, error="DB not connected"), 500
    d = request.get_json() or {}
    if not d.get("topic"):
        return jsonify(success=False, error="topic required"), 400
    try:
        record = {
            "faculty_id": d.get("faculty_id", ""), "subject": d.get("subject", ""),
            "week_number": d.get("week_number", 1), "topic": d["topic"],
            "planned_hours": d.get("planned_hours", 1), "actual_hours": d.get("actual_hours", 0),
            "status": d.get("status", "planned"), "notes": d.get("notes", ""),
            "planned_date": d.get("planned_date"),
        }
        result = sb.table("lesson_plans").insert(record).execute()
        row = result.data[0] if result.data else record
        rtdb_set(f"/lesson_plans/{row.get('id', 'new')}", row)
        return jsonify(success=True, plan=row)
    except Exception as e:
        return jsonify(success=False, error=str(e)), 500

@app.route("/api/lesson-plans/<lid>", methods=["PUT"])
def update_lesson_plan(lid):
    if not sb:
        return jsonify(success=False, error="DB not connected"), 500
    d = request.get_json() or {}
    try:
        result = sb.table("lesson_plans").update(d).eq("id", lid).execute()
        row = result.data[0] if result.data else {}
        rtdb_update(f"/lesson_plans/{lid}", d)
        return jsonify(success=True, plan=row)
    except Exception as e:
        return jsonify(success=False, error=str(e)), 500

@app.route("/api/lesson-plans/<lid>", methods=["DELETE"])
def delete_lesson_plan(lid):
    if not sb:
        return jsonify(success=False, error="DB not connected"), 500
    try:
        sb.table("lesson_plans").delete().eq("id", lid).execute()
        rtdb_delete(f"/lesson_plans/{lid}")
        return jsonify(success=True)
    except Exception as e:
        return jsonify(success=False, error=str(e)), 500


# ── NOTIFICATIONS ────────────────────────────────────────────
@app.route("/api/notifications", methods=["GET"])
def get_notifications():
    if not sb:
        return jsonify(success=True, notifications=[])
    try:
        q = sb.table("notifications").select("*").order("created_at", desc=True)
        uid = request.args.get("user_id")
        role = request.args.get("role")
        if uid and role:
            # Fetch notifications for this user OR broadcast to their role OR to 'all'
            q = sb.table("notifications").select("*").order("created_at", desc=True).or_(
                f"user_id.eq.{uid},target_role.eq.{role},target_role.eq.all"
            )
        elif uid:
            q = q.eq("user_id", uid)
        result = q.execute()
        return jsonify(success=True, notifications=result.data or [])
    except Exception as e:
        return jsonify(success=False, error=str(e)), 500

@app.route("/api/notifications", methods=["POST"])
def create_notification():
    if not sb:
        return jsonify(success=False, error="DB not connected"), 500
    d = request.get_json() or {}
    if not d.get("title") or not d.get("message"):
        return jsonify(success=False, error="title and message required"), 400
    try:
        record = {
            "user_id": d.get("user_id"), "target_role": d.get("target_role", "all"),
            "title": d["title"], "message": d["message"],
            "notification_type": d.get("notification_type", "info"),
            "is_read": False, "sent_by": d.get("sent_by", "admin"),
        }
        result = sb.table("notifications").insert(record).execute()
        row = result.data[0] if result.data else record
        rtdb_set(f"/notifications/{row.get('id', 'new')}", row)
        return jsonify(success=True, notification=row)
    except Exception as e:
        return jsonify(success=False, error=str(e)), 500

@app.route("/api/notifications/<nid>/read", methods=["PUT"])
def mark_notification_read(nid):
    if not sb:
        return jsonify(success=False, error="DB not connected"), 500
    try:
        sb.table("notifications").update({"is_read": True}).eq("id", nid).execute()
        rtdb_update(f"/notifications/{nid}", {"is_read": True})
        return jsonify(success=True)
    except Exception as e:
        return jsonify(success=False, error=str(e)), 500

@app.route("/api/notifications/<nid>", methods=["DELETE"])
def delete_notification(nid):
    if not sb:
        return jsonify(success=False, error="DB not connected"), 500
    try:
        sb.table("notifications").delete().eq("id", nid).execute()
        rtdb_delete(f"/notifications/{nid}")
        return jsonify(success=True)
    except Exception as e:
        return jsonify(success=False, error=str(e)), 500


# ── ATTENDANCE – MANUAL BULK MARK ────────────────────────────────────────────
@app.route("/api/attendance/manual", methods=["POST"])
def manual_attendance():
    """Bulk-mark attendance for a class session.
    Body: { session_id, date, subject_name, batch, faculty_id, marked_by,
            records: [{roll_no, student_name, status}] }
    Faculty can only mark for TODAY — past dates are locked.
    """
    if not sb:
        return jsonify(success=False, error="Supabase not configured"), 500
    d = request.json or {}
    records = d.get("records", [])
    if not records:
        return jsonify(success=False, error="records array required"), 400

    # ── Past-date lock ────────────────────────────────────────
    mark_date  = d.get("date", datetime.utcnow().date().isoformat())
    today_str  = datetime.utcnow().date().isoformat()
    if mark_date < today_str and not d.get("admin_override"):
        return jsonify(
            success=False,
            error=f"Attendance for {mark_date} is locked. Faculty can only mark today's ({today_str}) attendance.",
            locked=True
        ), 403

    # ── Faculty assignment enforcement ─────────────────────────
    faculty_id = d.get("faculty_id", "")
    subject_code_check = (d.get("subject_code") or "").strip().upper()
    section_check      = (d.get("section") or "").strip()
    if sb and faculty_id and subject_code_check and section_check and not d.get("admin_override"):
        try:
            asgn = sb.table("faculty_assignments").select("id") \
                .eq("faculty_id", faculty_id) \
                .eq("subject_code", subject_code_check) \
                .eq("section", section_check).limit(1).execute().data or []
            if not asgn:
                return jsonify(
                    success=False,
                    error="Access denied — you are not assigned to this subject/section.",
                    forbidden=True
                ), 403
        except Exception:
            pass  # if table not created yet, skip the check gracefully
    try:
        rows = []
        for r in records:
            rows.append({
                "session_id":    d.get("session_id", ""),
                "date":          d.get("date", datetime.utcnow().date().isoformat()),
                "subject_name":  d.get("subject_name", ""),
                "batch":         d.get("batch", ""),
                "department":    d.get("department", ""),
                "section":       d.get("section", ""),
                "semester":      d.get("semester", ""),
                "academic_year": d.get("academic_year", ""),
                "marked_by":     d.get("faculty_id", ""),
                "roll_no":       r.get("roll_no", ""),
                "name":          r.get("student_name", r.get("roll_no", "")),
                "student_name":  r.get("student_name", ""),
                "status":        r.get("status", "absent"),
                "session_type":  d.get("session_type", "lecture"),
                "method":        "manual",
                "timestamp":     datetime.utcnow().isoformat(),
                "verified":      "true",
            })
        result = sb.table("attendance").insert(rows).execute()
        saved = result.data or rows
        for row in saved:
            rtdb_set(f"/attendance/{row.get('session_id','x')}/{row.get('roll_no','x')}", row)
        return jsonify(success=True, count=len(saved), records=saved)
    except Exception as e:
        return jsonify(success=False, error=str(e)), 500


@app.route("/api/attendance/<att_id>", methods=["PUT"])
def update_attendance_record(att_id):
    """Update an attendance record.
    Faculty can ONLY edit today's records.
    Past records require admin_override=true in the request body.
    Face-marked records (method=face_recognition) are locked from faculty edit — admin only.
    """
    if not sb:
        return jsonify(success=False, error="Supabase not configured"), 500
    d = request.json or {}
    try:
        # Fetch the record to enforce lock rules
        existing = sb.table("attendance").select("date,method").eq("id", att_id).execute().data
        if existing:
            rec_date  = (existing[0].get("date") or "")[:10]
            today_str = datetime.utcnow().date().isoformat()
            rec_method = (existing[0].get("method") or "").lower()

            # Past-date lock
            if rec_date and rec_date < today_str and not d.get("admin_override"):
                return jsonify(
                    success=False,
                    error=f"Attendance record dated {rec_date} is locked. Only admin can modify past records.",
                    locked=True
                ), 403

            # Face-recognition lock — faculty cannot override face-marked records
            if rec_method in ("face_recognition", "face") and not d.get("admin_override"):
                return jsonify(
                    success=False,
                    error="Face-recognition attendance cannot be edited by faculty. Admin override required.",
                    face_locked=True
                ), 403

        allowed = ("status", "method", "name", "student_name", "roll_no",
                   "subject_name", "batch", "marked_by", "verified",
                   "department", "section", "semester")
        update_fields = {k: v for k, v in d.items() if k in allowed}
        result = sb.table("attendance").update(update_fields).eq("id", att_id).execute()
        record = result.data[0] if result.data else update_fields
        return jsonify(success=True, record=record)
    except Exception as e:
        return jsonify(success=False, error=str(e)), 500


# ══════════════════════════════════════════════════════════════
# MVP TIMETABLE SYSTEM — MASTER DATA + EXCEL UPLOAD
# ══════════════════════════════════════════════════════════════

# ── ROOMS ──────────────────────────────────────────────────────
@app.route("/api/rooms", methods=["GET"])
def get_rooms():
    if not sb:
        return jsonify(success=True, rooms=[])
    try:
        q = sb.table("rooms").select("*").order("room_number")
        if request.args.get("type"):
            q = q.eq("type", request.args["type"])
        if request.args.get("is_active"):
            q = q.eq("is_active", request.args["is_active"].lower() == "true")
        result = q.execute()
        return jsonify(success=True, rooms=result.data or [])
    except Exception as e:
        return jsonify(success=False, error=str(e)), 500

@app.route("/api/rooms", methods=["POST"])
def create_room():
    if not sb:
        return jsonify(success=False, error="Supabase not configured"), 500
    d = request.json or {}
    if not d.get("room_number"):
        return jsonify(success=False, error="room_number is required"), 400
    try:
        payload = {
            "room_number": d["room_number"].strip().upper(),
            "capacity":    int(d.get("capacity") or 60),
            "type":        d.get("type", "classroom").lower(),
            "building":    d.get("building", ""),
            "floor":       d.get("floor", ""),
            "is_active":   bool(d.get("is_active", True)),
            "created_at":  datetime.utcnow().isoformat(),
        }
        result = sb.table("rooms").insert(payload).execute()
        rec = result.data[0] if result.data else payload
        return jsonify(success=True, room=rec)
    except Exception as e:
        return jsonify(success=False, error=str(e)), 500

@app.route("/api/rooms/<room_id>", methods=["PUT"])
def update_room(room_id):
    if not sb:
        return jsonify(success=False, error="Supabase not configured"), 500
    d = request.json or {}
    try:
        allowed = ("room_number", "capacity", "type", "building", "floor", "is_active")
        upd = {k: v for k, v in d.items() if k in allowed}
        result = sb.table("rooms").update(upd).eq("id", room_id).execute()
        return jsonify(success=True, room=result.data[0] if result.data else upd)
    except Exception as e:
        return jsonify(success=False, error=str(e)), 500

@app.route("/api/rooms/<room_id>", methods=["DELETE"])
def delete_room(room_id):
    if not sb:
        return jsonify(success=False, error="Supabase not configured"), 500
    try:
        sb.table("rooms").delete().eq("id", room_id).execute()
        return jsonify(success=True)
    except Exception as e:
        return jsonify(success=False, error=str(e)), 500


# ── SUBJECTS CATALOGUE ─────────────────────────────────────────
@app.route("/api/subjects", methods=["GET"])
def get_subjects():
    if not sb:
        return jsonify(success=True, subjects=[])
    try:
        q = sb.table("subjects").select("*").order("subject_code")
        if request.args.get("department"):
            q = q.eq("department", request.args["department"])
        if request.args.get("semester"):
            try:
                q = q.eq("semester", int(request.args["semester"]))
            except ValueError:
                pass
        if request.args.get("program"):
            q = q.eq("program", request.args["program"])
        if request.args.get("type"):
            q = q.eq("type", request.args["type"])
        result = q.execute()
        return jsonify(success=True, subjects=result.data or [])
    except Exception as e:
        return jsonify(success=False, error=str(e)), 500

@app.route("/api/subjects", methods=["POST"])
def create_subject():
    if not sb:
        return jsonify(success=False, error="Supabase not configured"), 500
    d = request.json or {}
    if not d.get("subject_code") or not d.get("subject_name") or not d.get("department"):
        return jsonify(success=False, error="subject_code, subject_name, department required"), 400
    try:
        payload = {
            "subject_code": d["subject_code"].strip().upper(),
            "subject_name": d["subject_name"].strip(),
            "department":   d["department"].strip(),
            "program":      d.get("program", "B.Tech"),
            "semester":     int(d.get("semester") or 1),
            "weekly_hours": int(d.get("weekly_hours") or 3),
            "type":         d.get("type", "Theory"),
            "credits":      int(d.get("credits") or 3),
            "is_active":    bool(d.get("is_active", True)),
            "created_at":   datetime.utcnow().isoformat(),
        }
        result = sb.table("subjects").insert(payload).execute()
        rec = result.data[0] if result.data else payload
        return jsonify(success=True, subject=rec)
    except Exception as e:
        return jsonify(success=False, error=str(e)), 500

@app.route("/api/subjects/<subj_id>", methods=["PUT"])
def update_subject(subj_id):
    if not sb:
        return jsonify(success=False, error="Supabase not configured"), 500
    d = request.json or {}
    try:
        allowed = ("subject_code", "subject_name", "department", "program",
                   "semester", "weekly_hours", "type", "credits", "is_active")
        upd = {k: v for k, v in d.items() if k in allowed}
        result = sb.table("subjects").update(upd).eq("id", subj_id).execute()
        return jsonify(success=True, subject=result.data[0] if result.data else upd)
    except Exception as e:
        return jsonify(success=False, error=str(e)), 500

@app.route("/api/subjects/<subj_id>", methods=["DELETE"])
def delete_subject(subj_id):
    if not sb:
        return jsonify(success=False, error="Supabase not configured"), 500
    try:
        sb.table("subjects").delete().eq("id", subj_id).execute()
        return jsonify(success=True)
    except Exception as e:
        return jsonify(success=False, error=str(e)), 500


# ── FACULTY ASSIGNMENTS ────────────────────────────────────────
@app.route("/api/faculty-assignments", methods=["GET"])
def get_faculty_assignments():
    if not sb:
        return jsonify(success=True, assignments=[])
    try:
        q = sb.table("faculty_assignments").select("*").order("department").order("year")
        if request.args.get("faculty_id"):
            q = q.eq("faculty_id", request.args["faculty_id"])
        if request.args.get("faculty_username"):
            q = q.eq("faculty_username", request.args["faculty_username"])
        if request.args.get("department"):
            q = q.eq("department", request.args["department"])
        if request.args.get("section"):
            q = q.eq("section", request.args["section"])
        if request.args.get("year"):
            try:
                q = q.eq("year", int(request.args["year"]))
            except ValueError:
                pass
        if request.args.get("academic_year"):
            q = q.eq("academic_year", request.args["academic_year"])
        result = q.execute()
        return jsonify(success=True, assignments=result.data or [])
    except Exception as e:
        return jsonify(success=False, error=str(e)), 500

@app.route("/api/faculty-assignments", methods=["POST"])
def create_faculty_assignment():
    if not sb:
        return jsonify(success=False, error="Supabase not configured"), 500
    d = request.json or {}
    if not all([d.get("faculty_username"), d.get("subject_code"), d.get("section"), d.get("department")]):
        return jsonify(success=False, error="faculty_username, subject_code, section, department required"), 400
    try:
        payload = {
            "faculty_id":       d.get("faculty_id"),
            "faculty_username": (d.get("faculty_username") or "").strip(),
            "subject_code":     d["subject_code"].strip().upper(),
            "subject_name":     d.get("subject_name", ""),
            "course_code":      d.get("course_code", d["subject_code"].strip().upper()),
            "course_name":      d.get("course_name", d.get("subject_name", "")),
            "section":          d["section"].strip(),
            "department":       d["department"].strip(),
            "year":             int(d.get("year") or 1),
            "semester":         int(d.get("semester") or 1),
            "academic_year":    d.get("academic_year", "2025-26"),
            "is_active":        True,
            "created_at":       datetime.utcnow().isoformat(),
        }
        result = sb.table("faculty_assignments").insert(payload).execute()
        rec = result.data[0] if result.data else payload
        rtdb_set(f"/faculty_assignments/{rec.get('id', str(uuid.uuid4()))}", rec)
        return jsonify(success=True, assignment=rec)
    except Exception as e:
        return jsonify(success=False, error=str(e)), 500

@app.route("/api/faculty-assignments/<asgn_id>", methods=["DELETE"])
def delete_faculty_assignment(asgn_id):
    if not sb:
        return jsonify(success=False, error="Supabase not configured"), 500
    try:
        sb.table("faculty_assignments").delete().eq("id", asgn_id).execute()
        rtdb_delete(f"/faculty_assignments/{asgn_id}")
        return jsonify(success=True)
    except Exception as e:
        return jsonify(success=False, error=str(e)), 500


# ── EXCEL TIMETABLE TEMPLATE ───────────────────────────────────
@app.route("/api/timetable/excel-template", methods=["GET"])
def timetable_excel_template():
    """Generate and return a formatted Excel template with a reference sheet."""
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.utils import get_column_letter
        import io

        wb = openpyxl.Workbook()

        # ── Sheet 1: Template ─────────────────────────────────
        ws = wb.active
        ws.title = "Timetable Template"

        HEADERS = [
            "Department", "Program", "Year", "Semester", "Section",
            "Day", "Period Number", "Start Time", "End Time",
            "Subject Code", "Subject Name", "Faculty Username",
            "Room Number", "Type",
        ]

        hdr_fill = PatternFill(start_color="1F6FEB", end_color="1F6FEB", fill_type="solid")
        hdr_font = Font(bold=True, color="FFFFFF", size=11)
        center    = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for col, hdr in enumerate(HEADERS, 1):
            c = ws.cell(row=1, column=col, value=hdr)
            c.fill, c.font, c.alignment = hdr_fill, hdr_font, center
        ws.row_dimensions[1].height = 28

        # Sample rows
        SAMPLES = [
            ["Computer Science", "B.Tech", 2, 3, "A",
             "Monday", 1, "09:00", "10:00", "CS301", "Data Structures", "john.faculty", "A101", "Theory"],
            ["Computer Science", "B.Tech", 2, 3, "A",
             "Monday", 2, "10:00", "11:00", "CS302", "Operating Systems", "jane.faculty", "A102", "Theory"],
            ["Computer Science", "B.Tech", 2, 3, "A",
             "Tuesday", 1, "09:00", "11:00", "CS301L", "Data Structures Lab", "john.faculty", "LAB1", "Lab"],
        ]
        light_fill = PatternFill(start_color="EBF4FF", end_color="EBF4FF", fill_type="solid")
        for row_data in SAMPLES:
            ws.append(row_data)

        for row in ws.iter_rows(min_row=2, max_row=1 + len(SAMPLES)):
            for cell in row:
                cell.fill = light_fill

        # Hint row
        HINTS = ["*Required", "*Required", "1-4", "1-8", "A/B/C",
                 "Monday-Saturday", "1-8", "HH:MM", "HH:MM",
                 "*Required", "*Required", "*Required",
                 "e.g. A101", "Theory/Lab/Tutorial"]
        ws.append(HINTS)
        hint_row = ws.max_row
        hint_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
        hint_font = Font(italic=True, color="888888", size=9)
        for cell in ws[hint_row]:
            cell.fill, cell.font = hint_fill, hint_font

        COL_WIDTHS = [20, 12, 6, 10, 8, 12, 14, 12, 10, 14, 24, 18, 14, 12]
        for i, w in enumerate(COL_WIDTHS, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        # ── Sheet 2: Reference Data ────────────────────────────
        ws2 = wb.create_sheet(title="Reference Data")
        ref_hdr_fill = PatternFill(start_color="2EA043", end_color="2EA043", fill_type="solid")
        ref_hdr_font = Font(bold=True, color="FFFFFF", size=10)

        def _ref_header(ws, row, col, text):
            c = ws.cell(row=row, column=col, value=text)
            c.fill, c.font = ref_hdr_fill, ref_hdr_font

        # Column A-C: Faculty
        _ref_header(ws2, 1, 1, "Faculty Username")
        _ref_header(ws2, 1, 2, "Full Name")
        _ref_header(ws2, 1, 3, "Department")
        try:
            if sb:
                faculty_rows = sb.table("users").select("username,full_name,department") \
                    .eq("role", "faculty").eq("is_active", True).execute().data or []
                for i, f in enumerate(faculty_rows):
                    ws2.cell(row=2+i, column=1, value=f.get("username", ""))
                    ws2.cell(row=2+i, column=2, value=f.get("full_name", ""))
                    ws2.cell(row=2+i, column=3, value=f.get("department", ""))
        except Exception:
            pass

        # Column E-G: Rooms
        _ref_header(ws2, 1, 5, "Room Number")
        _ref_header(ws2, 1, 6, "Type")
        _ref_header(ws2, 1, 7, "Capacity")
        try:
            if sb:
                room_rows = sb.table("rooms").select("room_number,type,capacity").execute().data or []
                for i, r in enumerate(room_rows):
                    ws2.cell(row=2+i, column=5, value=r.get("room_number", ""))
                    ws2.cell(row=2+i, column=6, value=r.get("type", ""))
                    ws2.cell(row=2+i, column=7, value=r.get("capacity", ""))
        except Exception:
            pass

        # Column I-L: Subject Codes
        _ref_header(ws2, 1, 9,  "Subject Code")
        _ref_header(ws2, 1, 10, "Subject Name")
        _ref_header(ws2, 1, 11, "Department")
        _ref_header(ws2, 1, 12, "Semester")
        try:
            if sb:
                subj_rows = sb.table("subjects").select(
                    "subject_code,subject_name,department,semester"
                ).execute().data or []
                for i, s in enumerate(subj_rows):
                    ws2.cell(row=2+i, column=9,  value=s.get("subject_code", ""))
                    ws2.cell(row=2+i, column=10, value=s.get("subject_name", ""))
                    ws2.cell(row=2+i, column=11, value=s.get("department", ""))
                    ws2.cell(row=2+i, column=12, value=s.get("semester", ""))
        except Exception:
            pass

        # Column N: Valid Types   Column P: Valid Days
        _ref_header(ws2, 1, 14, "Valid Types")
        for i, t in enumerate(["Theory", "Lab", "Tutorial"]):
            ws2.cell(row=2+i, column=14, value=t)

        _ref_header(ws2, 1, 16, "Valid Days")
        for i, day in enumerate(["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"]):
            ws2.cell(row=2+i, column=16, value=day)

        for col_letter, width in [("A",20),("B",25),("C",20),("E",14),("F",14),("G",10),
                                    ("I",14),("J",25),("K",20),("L",10)]:
            ws2.column_dimensions[col_letter].width = width

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        from flask import Response
        return Response(
            buf.read(),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=timetable_template.xlsx"},
        )
    except ImportError:
        return jsonify(success=False, error="openpyxl not installed. Run: pip install openpyxl"), 500
    except Exception as e:
        return jsonify(success=False, error=str(e)), 500


# ── EXCEL TIMETABLE UPLOAD + VALIDATE + COMMIT ─────────────────
@app.route("/api/timetable/excel-upload", methods=["POST"])
def timetable_excel_upload():
    """Upload an Excel or CSV timetable file.
    Form fields: file (xlsx/xls/csv), commit=true|false, academic_year, semester
    dry-run (commit=false): validate + return preview with errors/warnings.
    commit (commit=true):   save rows + create faculty_assignments.
    """
    import io as _io
    if "file" not in request.files:
        return jsonify(success=False, error="No file uploaded — use multipart/form-data field 'file'"), 400

    f           = request.files["file"]
    is_commit   = request.form.get("commit", "false").lower() == "true"
    academic_yr = request.form.get("academic_year", "2025-26")
    semester_raw = request.form.get("semester", "")
    fname       = (f.filename or "").lower()

    if not (fname.endswith(".xlsx") or fname.endswith(".xls") or fname.endswith(".csv")):
        return jsonify(success=False, error="File must be .xlsx, .xls, or .csv"), 400

    try:
        file_bytes = f.read()
        raw_slots = []

        if fname.endswith(".csv"):
            import csv
            text = file_bytes.decode("utf-8-sig", errors="replace")
            reader = csv.DictReader(_io.StringIO(text))
            for row in reader:
                raw_slots.append({k.strip().lower().replace(" ", "_"): (v or "").strip() for k, v in row.items() if k})
        else:
            import openpyxl
            wb = openpyxl.load_workbook(_io.BytesIO(file_bytes), read_only=True, data_only=True)
            ws = wb.active
            headers = None
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:
                    headers = [str(c or "").strip().lower().replace(" ", "_") for c in row]
                    continue
                if not headers or all(v is None for v in row):
                    continue
                slot = {}
                for j, val in enumerate(row):
                    if j < len(headers) and headers[j]:
                        slot[headers[j]] = str(val).strip() if val is not None else ""
                dept_val = slot.get("department", "")
                if not dept_val or dept_val.startswith("*"):
                    continue
                raw_slots.append(slot)
            wb.close()

        if not raw_slots:
            return jsonify(success=False, error="No data rows found in the file"), 400

        # ── Normalise column names ─────────────────────────────
        ALIAS = {
            "period_number": "period_number", "period number": "period_number",
            "hour": "period_number", "hour_number": "period_number",
            "day": "day", "day_of_week": "day", "day of week": "day", "weekday": "day",
            "faculty_username": "faculty_username", "faculty username": "faculty_username",
            "faculty": "faculty_username", "faculty_id": "faculty_username", "employee_id": "faculty_username",
            "faculty_user": "faculty_username",
            "start_time": "start_time", "start time": "start_time",
            "end_time": "end_time", "end time": "end_time",
            "subject_code": "subject_code", "subject code": "subject_code",
            "subject_name": "subject_name", "subject name": "subject_name",
            "room_number": "room_number", "room number": "room_number", "room": "room_number",
        }
        slots = []
        for s in raw_slots:
            ns = {}
            for k, v in s.items():
                ns[ALIAS.get(k, k)] = v
            slots.append(ns)

        # ── Fetch reference data for validation ────────────────
        valid_usernames  = set()
        username_to_info = {}
        valid_rooms      = set()
        valid_subj_codes = set()

        if sb:
            try:
                fu_rows = sb.table("users").select("id,username,employee_id,full_name").eq("role", "faculty").execute().data or []
                for u in fu_rows:
                    uname = (u.get("username") or "").strip().lower()
                    empid = (u.get("employee_id") or "").strip().lower()
                    info = {
                        "id": u["id"],
                        "full_name": u.get("full_name", ""),
                        "username": u.get("username", "")
                    }
                    if uname:
                        valid_usernames.add(uname)
                        username_to_info[uname] = info
                    if empid:
                        valid_usernames.add(empid)
                        username_to_info[empid] = info
            except Exception:
                pass
            try:
                rm_rows = sb.table("rooms").select("room_number").execute().data or []
                valid_rooms = {(r.get("room_number") or "").strip().lower() for r in rm_rows}
            except Exception:
                pass
            try:
                sc_rows = sb.table("subjects").select("subject_code").execute().data or []
                valid_subj_codes = {(s.get("subject_code") or "").strip().upper() for s in sc_rows}
            except Exception:
                pass

        # ── Build conflict index from existing timetable ───────
        fac_idx  = {}
        room_idx = {}
        sect_idx = {}
        if sb:
            try:
                ex_rows = sb.table("timetable").select(
                    "faculty_username,room_number,section,department,year,day_of_week,period_number,hour_number"
                ).execute().data or []
                for r in ex_rows:
                    fu  = (r.get("faculty_username") or "").strip().lower()
                    rm  = (r.get("room_number") or "").strip().lower()
                    sec = (r.get("section") or "").strip().lower()
                    dep = (r.get("department") or "").strip().lower()
                    yr  = str(r.get("year") or "")
                    day = r.get("day_of_week", "")
                    per = str(r.get("period_number") or r.get("hour_number") or "")
                    if fu:  fac_idx[(fu, day, per)]             = True
                    if rm:  room_idx[(rm, day, per)]            = True
                    if sec: sect_idx[(sec, dep, yr, day, per)]  = True
            except Exception:
                pass

        DAY_CANON = {d.lower(): d for d in
                     ["Monday","Tuesday","Wednesday","Thursday","Friday","Saturday"]}
        VALID_TYPES = {"theory", "lab", "tutorial"}

        errors    = []
        warnings  = []
        valid_rows = []

        for i, s in enumerate(slots):
            rn     = i + 2
            errs   = []

            dept     = (s.get("department","")).strip()
            program  = (s.get("program","")).strip() or "B.Tech"
            year_raw = s.get("year","")
            sem_raw  = s.get("semester","")
            section  = (s.get("section","")).strip()
            day_raw  = (s.get("day","")).strip().lower()
            per_raw  = (s.get("period_number","")).strip()
            start    = (s.get("start_time","")).strip()
            end_t    = (s.get("end_time","")).strip()
            subcode  = (s.get("subject_code","")).strip().upper()
            subname  = (s.get("subject_name","")).strip()
            fuser    = (s.get("faculty_username","")).strip().lower()
            room     = (s.get("room_number","")).strip()
            typ_raw  = (s.get("type","theory")).strip().lower()

            if not dept:    errs.append("Department is required")
            if not section: errs.append("Section is required")
            if not subcode: errs.append("Subject Code is required")
            if not subname: errs.append("Subject Name is required")
            if not fuser:   errs.append("Faculty Username is required")
            if not start or not end_t: errs.append("Start Time and End Time are required")

            day = DAY_CANON.get(day_raw)
            if not day:
                errs.append(f"Invalid day '{s.get('day','')}' — use Monday…Saturday")

            period = None
            try:
                period = int(float(per_raw))
                if not (1 <= period <= 10):
                    errs.append(f"Period Number must be 1–10, got '{per_raw}'")
            except (ValueError, TypeError):
                errs.append(f"Period Number must be a number, got '{per_raw}'")

            year = None
            try:
                year = int(float(year_raw))
                if not (1 <= year <= 6):
                    errs.append(f"Year must be 1–6, got '{year_raw}'")
            except (ValueError, TypeError):
                errs.append(f"Year must be a number, got '{year_raw}'")

            typ = typ_raw if typ_raw in VALID_TYPES else "theory"
            if typ_raw and typ_raw not in VALID_TYPES:
                errs.append(f"Type must be Theory/Lab/Tutorial, got '{s.get('type','')}'")

            # Cross-reference checks
            if fuser and valid_usernames and fuser not in valid_usernames:
                warnings.append(f"Row {rn}: Faculty username '{fuser}' not found in the system (saved without faculty_id)")
            if room and valid_rooms and room.lower() not in valid_rooms:
                warnings.append(f"Row {rn}: Room '{room}' not in rooms catalogue (will be saved anyway)")
            if valid_subj_codes and subcode and subcode not in valid_subj_codes:
                warnings.append(f"Row {rn}: Subject code '{subcode}' not in subjects catalogue")

            if errs:
                for e in errs:
                    errors.append({"row": rn, "message": e, "data": s})
                continue

            per_str = str(period)
            # Conflict checks
            if fuser in valid_usernames or not valid_usernames:
                if (fuser, day, per_str) in fac_idx:
                    warnings.append(f"Row {rn}: Faculty '{fuser}' already has a class on {day} Period {period}")
            if room and (room.lower(), day, per_str) in room_idx:
                warnings.append(f"Row {rn}: Room '{room}' already booked on {day} Period {period}")
            if section and dept and year is not None:
                sk = (section.lower(), dept.lower(), str(year), day, per_str)
                if sk in sect_idx:
                    warnings.append(f"Row {rn}: Section {section} ({dept} Yr{year}) already has a class on {day} Period {period}")

            fac_info = username_to_info.get(fuser, {})
            try:
                sem_val = int(float(sem_raw or semester_raw or 1))
            except (ValueError, TypeError):
                sem_val = 1

            resolved_faculty_username = (fac_info.get("username") or fuser or "").strip().lower()

            valid_rows.append({
                "department":       dept,
                "program":          program,
                "year":             year,
                "semester":         sem_val,
                "section":          section,
                "day_of_week":      day,
                "period_number":    period,
                "hour_number":      period,
                "start_time":       start,
                "end_time":         end_t,
                "subject_code":     subcode,
                "subject_name":     subname,
                "faculty_username": resolved_faculty_username,
                "faculty_id":       fac_info.get("id") or None,
                "faculty_name":     fac_info.get("full_name", "") or fuser,
                "room_number":      room,
                "session_type":     typ,
                "type":             typ,
                "academic_year":    academic_yr,
                "active":           True,
            })

        if not is_commit:
            return jsonify(
                success=True, dry_run=True,
                total=len(slots), valid_count=len(valid_rows),
                error_count=len(errors), warning_count=len(warnings),
                errors=errors, warnings=warnings, preview=valid_rows[:100],
            )

        # ── Commit phase ───────────────────────────────────────
        if errors:
            return jsonify(
                success=False,
                error=f"Fix {len(errors)} validation error(s) before saving.",
                errors=errors, warnings=warnings,
            ), 400

        saved  = []
        failed = []
        assign_keys = set()

        for entry in valid_rows:
            try:
                payload = {k: v for k, v in entry.items() if k != "type"}
                payload["created_at"] = datetime.utcnow().isoformat()
                result = sb.table("timetable").insert(payload).execute()
                rec    = result.data[0] if result.data else payload
                rec_id = rec.get("id", "")
                if rec_id:
                    rtdb_set(f"/timetable/{rec_id}", rec)
                saved.append(rec)

                # Auto-create faculty_assignment
                fid   = entry.get("faculty_id") or ""
                fuser = entry.get("faculty_username", "")
                scode = entry.get("subject_code", "")
                sect  = entry.get("section", "")
                dep   = entry.get("department", "")
                yr    = entry.get("year", 1)
                sem   = entry.get("semester", 1)
                ayr   = entry.get("academic_year", "2025-26")
                akey  = (fid, scode, sect, dep, yr, sem, ayr)

                if fid and fuser and akey not in assign_keys:
                    try:
                        ex_asgn = sb.table("faculty_assignments").select("id,timetable_slot_ids") \
                            .eq("faculty_id", fid) \
                            .eq("subject_code", scode) \
                            .eq("section", sect) \
                            .eq("department", dep) \
                            .eq("year", yr) \
                            .eq("semester", sem) \
                            .eq("academic_year", ayr).execute().data or []

                        if ex_asgn:
                            # Faculty is already assigned to this subject in this section
                            warnings.append(f"Faculty '{fuser}' is already assigned to '{scode}' in {sect} — updating slot mapping")
                            aid      = ex_asgn[0]["id"]
                            slot_ids = list(ex_asgn[0].get("timetable_slot_ids") or [])
                            if rec_id and rec_id not in slot_ids:
                                slot_ids.append(rec_id)
                            sb.table("faculty_assignments").update(
                                {"timetable_slot_ids": slot_ids}
                            ).eq("id", aid).execute()
                        else:
                            ap = {
                                "faculty_id":         fid,
                                "faculty_username":   fuser,
                                "subject_code":       scode,
                                "subject_name":       entry.get("subject_name", ""),
                                "section":            sect,
                                "department":         dep,
                                "year":               yr,
                                "semester":           sem,
                                "academic_year":      ayr,
                                "timetable_slot_ids": [rec_id] if rec_id else [],
                                "created_at":         datetime.utcnow().isoformat(),
                            }
                            ar = sb.table("faculty_assignments").insert(ap).execute()
                            if ar.data:
                                rtdb_set(f"/faculty_assignments/{ar.data[0]['id']}", ar.data[0])
                        assign_keys.add(akey)
                    except Exception:
                        pass  # assignment creation is best-effort

                # Also upsert course record
                try:
                    if fid and scode:
                        ex_course = sb.table("courses").select("id") \
                            .eq("course_code", scode).eq("faculty_id", fid).execute().data or []
                        if not ex_course:
                            cp = {
                                "course_code": scode,
                                "course_name": entry.get("subject_name", ""),
                                "department":  dep,
                                "faculty_id":  fid,
                                "semester":    sem,
                                "academic_year": ayr,
                                "credits":     3,
                                "created_at":  datetime.utcnow().isoformat(),
                            }
                            cr = sb.table("courses").insert(cp).execute()
                            if cr.data:
                                rtdb_set(f"/courses/{cr.data[0]['id']}", cr.data[0])
                except Exception:
                    pass

            except Exception as ex:
                failed.append({"row": entry, "error": str(ex)})

        return jsonify(
            success=True, committed=True,
            saved_count=len(saved), failed_count=len(failed),
            assignments_created=len(assign_keys),
            warnings=warnings,
            errors=[f["error"] for f in failed],
        )

    except ImportError:
        return jsonify(success=False, error="openpyxl not installed. Run: pip install openpyxl"), 500
    except Exception as e:
        return jsonify(success=False, error=str(e)), 500


# ── TODAY'S SCHEDULE FOR FACULTY ───────────────────────────────
@app.route("/api/faculty/<faculty_id>/today-schedule", methods=["GET"])
def faculty_today_schedule(faculty_id):
    """Return today's timetable entries for a faculty, with attendance status."""
    if not sb:
        return jsonify(success=True, slots=[], today="")
    today = datetime.utcnow().strftime("%A")   # e.g. "Monday"
    try:
        rows = sb.table("timetable").select("*") \
            .eq("faculty_id", faculty_id) \
            .eq("day_of_week", today) \
            .order("period_number").execute().data or []

        if not rows:
            rows = sb.table("timetable").select("*") \
                .eq("day_of_week", today).execute().data or []
            rows = [r for r in rows if r.get("faculty_id") == faculty_id
                    or (r.get("faculty_username") or "").lower() == faculty_id.lower()]

        # Mark which slots already have attendance today
        today_date = datetime.utcnow().date().isoformat()
        for slot in rows:
            try:
                att = sb.table("attendance").select("id").eq("date", today_date) \
                    .eq("subject_name", slot.get("subject_name", "")) \
                    .eq("marked_by", faculty_id).limit(1).execute().data or []
                slot["attendance_marked"] = bool(att)
            except Exception:
                slot["attendance_marked"] = False

        return jsonify(success=True, slots=rows, today=today)
    except Exception as e:
        return jsonify(success=False, error=str(e)), 500


# ────────────────────────────────────────────────────────────────────
# TIMETABLE GENERATION & OPTIMIZATION ENDPOINTS
# ────────────────────────────────────────────────────────────────────

@app.route("/api/timetable/generate", methods=["POST"])
def generate_timetable():
    """Start timetable generation using optimization algorithm"""
    if not sb:
        return jsonify(success=False, error="Supabase not configured"), 500
    
    try:
        d = request.json or {}
        academic_year = d.get("academic_year", "2025-26")
        semester = d.get("semester", 1)
        algorithm = d.get("algorithm", "simulated_annealing")  # or "genetic"
        
        if algorithm not in ["simulated_annealing", "genetic"]:
            return jsonify(success=False, error="Invalid algorithm type"), 400
        
        # Check if generation already in progress
        try:
            existing = sb.table("timetable_generation_jobs").select("*") \
                .eq("academic_year", academic_year) \
                .eq("semester", semester) \
                .eq("status", "processing").limit(1).execute().data or []
            if existing:
                return jsonify(success=False, 
                    error="Generation already in progress"), 409
        except Exception:
            pass  # Table may not exist yet
        
        # Load faculty assignments to check if we have enough data
        try:
            assignments = sb.table("faculty_assignments").select("*") \
                .eq("academic_year", academic_year) \
                .is_("is_active", True) \
                .execute().data or []
            
            if not assignments:
                return jsonify(success=False, 
                    error="No faculty assignments found. Please add assignments first."), 400
        except Exception as e:
            return jsonify(success=False, error=f"Could not load assignments: {e}"), 400
        
        # Load rooms
        try:
            rooms = sb.table("room_capacity").select("*") \
                .eq("academic_year", academic_year) \
                .is_("is_available", True) \
                .execute().data or []
            
            if not rooms:
                return jsonify(success=False, 
                    error="No rooms configured. Please configure room availability first."), 400
        except Exception as e:
            return jsonify(success=False, error=f"Could not load rooms: {e}"), 400
        
        # Load subjects
        try:
            subjects = sb.table("subjects").select("*") \
                .is_("is_active", True) \
                .execute().data or []
        except Exception as e:
            return jsonify(success=False, error=f"Could not load subjects: {e}"), 400
        
        # Check constraints
        total_periods_week = 6 * 8  # 6 days * 8 periods
        total_slots_needed = sum(s.get("weekly_hours", 0) for s in subjects)
        
        if total_slots_needed > total_periods_week * 3:  # Very rough estimate
            return jsonify(success=False, 
                error=f"Insufficient room capacity. Need {total_slots_needed} slots, "
                      f"have ~{total_periods_week} available."), 400
        
        # Return success with generation parameters
        return jsonify(
            success=True,
            message="Timetable generation started",
            academic_year=academic_year,
            semester=semester,
            algorithm=algorithm,
            assignments_count=len(assignments),
            rooms_count=len(rooms),
            subjects_count=len(subjects),
            status="ready_to_generate"
        )
        
    except Exception as e:
        logger.exception("Error in generate_timetable")
        return jsonify(success=False, error=str(e)), 500


@app.route("/api/timetable/config", methods=["GET", "POST"])
def timetable_config():
    """Get or update timetable generation configuration"""
    if not sb:
        return jsonify(success=False, error="Supabase not configured"), 500
    
    try:
        if request.method == "GET":
            academic_year = request.args.get("academic_year", "2025-26")
            semester = request.args.get("semester", "1")
            
            try:
                semester = int(semester)
            except:
                semester = 1
            
            # Load config
            config = sb.table("timetable_config").select("*") \
                .eq("academic_year", academic_year) \
                .eq("semester", semester).limit(1).execute().data or []
            
            if config:
                return jsonify(success=True, config=config[0])
            
            # Return default config
            return jsonify(success=True, config={
                "academic_year": academic_year,
                "semester": semester,
                "total_days_per_week": 6,
                "total_periods_per_day": 8,
                "max_hours_per_faculty": 5,
                "generation_algorithm": "simulated_annealing",
                "algorithm_iterations": 10000
            })
            
        elif request.method == "POST":
            d = request.json or {}
            academic_year = d.get("academic_year", "2025-26")
            semester = d.get("semester", 1)
            
            config_data = {
                "academic_year": academic_year,
                "semester": semester,
                "total_days_per_week": d.get("total_days_per_week", 6),
                "total_periods_per_day": d.get("total_periods_per_day", 8),
                "max_hours_per_faculty": d.get("max_hours_per_faculty", 5),
                "generation_algorithm": d.get("generation_algorithm", "simulated_annealing"),
                "algorithm_iterations": d.get("algorithm_iterations", 10000),
                "updated_at": datetime.utcnow().isoformat()
            }
            
            try:
                # Upsert config
                existing = sb.table("timetable_config").select("id") \
                    .eq("academic_year", academic_year) \
                    .eq("semester", semester).limit(1).execute().data or []
                
                if existing:
                    result = sb.table("timetable_config").update(config_data) \
                        .eq("academic_year", academic_year) \
                        .eq("semester", semester).execute()
                else:
                    result = sb.table("timetable_config").insert([config_data]).execute()
                
                return jsonify(success=True, message="Configuration saved", config=config_data)
            except Exception as e:
                return jsonify(success=False, error=f"Could not save config: {e}"), 500
                
    except Exception as e:
        logger.exception("Error in timetable_config")
        return jsonify(success=False, error=str(e)), 500


@app.route("/api/timetable/constraints/faculty", methods=["GET", "POST"])
def faculty_constraints():
    """Get or set faculty constraints (free hours, max weekly classes)"""
    if not sb:
        return jsonify(success=False, error="Supabase not configured"), 500
    
    try:
        if request.method == "GET":
            academic_year = request.args.get("academic_year", "2025-26")
            semester = request.args.get("semester", "1")
            faculty_username = request.args.get("faculty_username")
            
            try:
                semester = int(semester)
            except:
                semester = 1
            
            query = sb.table("faculty_constraints").select("*") \
                .eq("academic_year", academic_year) \
                .eq("semester", semester)
            
            if faculty_username:
                query = query.eq("faculty_username", faculty_username)
            
            constraints = query.execute().data or []
            return jsonify(success=True, constraints=constraints)
            
        elif request.method == "POST":
            d = request.json or {}
            academic_year = d.get("academic_year", "2025-26")
            semester = d.get("semester", 1)
            faculty_username = d.get("faculty_username")
            
            if not faculty_username:
                return jsonify(success=False, error="faculty_username is required"), 400
            
            constraint_data = {
                "faculty_username": faculty_username,
                "academic_year": academic_year,
                "semester": semester,
                "max_classes_per_week": d.get("max_classes_per_week", 5),
                "max_consecutive_hours": d.get("max_consecutive_hours", 3),
                "free_days": d.get("free_days", []),
                "free_periods": json.dumps(d.get("free_periods", {})),
                "no_first_period": d.get("no_first_period", False),
                "updated_at": datetime.utcnow().isoformat()
            }
            
            try:
                # Get faculty department if available
                faculty = sb.table("users").select("*") \
                    .eq("username", faculty_username).limit(1).execute().data or []
                if faculty:
                    constraint_data["department"] = faculty[0].get("department", "")
                
                # Upsert constraint
                result = sb.table("faculty_constraints").upsert([constraint_data]).execute()
                return jsonify(success=True, message="Constraint saved", constraint=constraint_data)
            except Exception as e:
                return jsonify(success=False, error=f"Could not save constraint: {e}"), 500
                
    except Exception as e:
        logger.exception("Error in faculty_constraints")
        return jsonify(success=False, error=str(e)), 500


@app.route("/api/timetable/constraints/department", methods=["GET", "POST"])
def department_constraints():
    """Get or set department free hour constraints"""
    if not sb:
        return jsonify(success=False, error="Supabase not configured"), 500
    
    try:
        if request.method == "GET":
            academic_year = request.args.get("academic_year", "2025-26")
            semester = request.args.get("semester", "1")
            department = request.args.get("department")
            
            try:
                semester = int(semester)
            except:
                semester = 1
            
            query = sb.table("department_constraints").select("*") \
                .eq("academic_year", academic_year) \
                .eq("semester", semester)
            
            if department:
                query = query.eq("department", department)
            
            constraints = query.execute().data or []
            return jsonify(success=True, constraints=constraints)
            
        elif request.method == "POST":
            d = request.json or {}
            academic_year = d.get("academic_year", "2025-26")
            semester = d.get("semester", 1)
            department = d.get("department")
            free_day = d.get("free_day")  # e.g., "Friday"
            free_periods = d.get("free_periods", [])  # e.g., [9, 10]
            
            if not department or not free_day or not free_periods:
                return jsonify(success=False, 
                    error="department, free_day, and free_periods are required"), 400
            
            constraint_data = {
                "department": department,
                "academic_year": academic_year,
                "semester": semester,
                "free_day": free_day,
                "free_periods": free_periods,
                "is_mandatory": d.get("is_mandatory", True),
                "description": d.get("description", "")
            }
            
            try:
                result = sb.table("department_constraints").upsert([constraint_data]).execute()
                return jsonify(success=True, message="Department constraint saved", 
                             constraint=constraint_data)
            except Exception as e:
                return jsonify(success=False, error=f"Could not save constraint: {e}"), 500
                
    except Exception as e:
        logger.exception("Error in department_constraints")
        return jsonify(success=False, error=str(e)), 500


@app.route("/api/timetable/rooms", methods=["GET", "POST"])
def room_management():
    """Get or add room capacity information"""
    if not sb:
        return jsonify(success=False, error="Supabase not configured"), 500
    
    try:
        if request.method == "GET":
            academic_year = request.args.get("academic_year", "2025-26")
            
            rooms = sb.table("room_capacity").select("*") \
                .eq("academic_year", academic_year) \
                .execute().data or []
            
            total_rooms = len(rooms)
            available_rooms = sum(1 for r in rooms if r.get("is_available"))
            
            return jsonify(success=True, rooms=rooms, 
                         total_rooms=total_rooms, 
                         available_rooms=available_rooms)
            
        elif request.method == "POST":
            d = request.json or {}
            academic_year = d.get("academic_year", "2025-26")
            rooms_list = d.get("rooms", [])
            
            if not rooms_list:
                return jsonify(success=False, error="rooms list is required"), 400
            
            # Prepare room data
            room_data = []
            for room in rooms_list:
                room_data.append({
                    "room_number": room.get("room_number"),
                    "academic_year": academic_year,
                    "capacity": room.get("capacity", 60),
                    "room_type": room.get("room_type", "classroom"),
                    "has_projector": room.get("has_projector", False),
                    "has_computers": room.get("has_computers", False),
                    "is_available": room.get("is_available", True)
                })
            
            try:
                result = sb.table("room_capacity").upsert(room_data).execute()
                return jsonify(success=True, message=f"Added {len(room_data)} rooms", 
                             rooms=result.data)
            except Exception as e:
                return jsonify(success=False, error=f"Could not save rooms: {e}"), 500
                
    except Exception as e:
        logger.exception("Error in room_management")
        return jsonify(success=False, error=str(e)), 500


@app.route("/api/timetable/validate", methods=["POST"])
def validate_timetable():
    """Validate current timetable for conflicts"""
    if not sb:
        return jsonify(success=False, error="Supabase not configured"), 500
    
    try:
        d = request.json or {}
        academic_year = d.get("academic_year", "2025-26")
        semester = d.get("semester", 1)
        
        # Get current timetable
        timetable = sb.table("timetable").select("*") \
            .eq("academic_year", academic_year) \
            .eq("semester", semester) \
            .execute().data or []
        
        conflicts = []
        warnings = []
        
        # Check for room conflicts
        room_schedule = {}
        for slot in timetable:
            room = (slot.get("room_number") or "").lower()
            day = slot.get("day_of_week", "")
            hr = str(slot.get("hour_number", ""))
            
            if room and day and hr:
                key = (room, day, hr)
                if key in room_schedule:
                    conflicts.append({
                        "type": "room_conflict",
                        "room": room,
                        "day": day,
                        "hour": hr,
                        "slots": [slot, room_schedule[key]]
                    })
                room_schedule[key] = slot
        
        # Check for faculty conflicts
        faculty_schedule = {}
        for slot in timetable:
            faculty = (slot.get("faculty_username") or "").lower()
            day = slot.get("day_of_week", "")
            hr = str(slot.get("hour_number", ""))
            
            if faculty and day and hr:
                key = (faculty, day, hr)
                if key in faculty_schedule:
                    conflicts.append({
                        "type": "faculty_conflict",
                        "faculty": faculty,
                        "day": day,
                        "hour": hr
                    })
                faculty_schedule[key] = slot
        
        # Check faculty weekly load
        faculty_weekly = {}
        for slot in timetable:
            faculty = slot.get("faculty_username", "").lower()
            if faculty:
                faculty_weekly[faculty] = faculty_weekly.get(faculty, 0) + 1
        
        for faculty, count in faculty_weekly.items():
            if count > 5:
                warnings.append({
                    "type": "faculty_overload",
                    "faculty": faculty,
                    "classes_per_week": count,
                    "max_allowed": 5
                })
        
        return jsonify(success=True, 
                      total_slots=len(timetable),
                      conflicts=conflicts,
                      conflict_count=len(conflicts),
                      warnings=warnings,
                      warning_count=len(warnings))
        
    except Exception as e:
        logger.exception("Error in validate_timetable")
        return jsonify(success=False, error=str(e)), 500


@app.route("/api/timetable/breaks", methods=["GET", "POST"])
def manage_breaks():
    """Get or create break/lunch schedule for timetable generation"""
    if not sb:
        return jsonify(success=False, error="Supabase not configured"), 500
    
    try:
        if request.method == "GET":
            academic_year = request.args.get("academic_year", "2025-26")
            semester = request.args.get("semester", "1")
            day = request.args.get("day")
            
            try:
                semester = int(semester)
            except:
                semester = 1
            
            query = sb.table("break_schedule").select("*") \
                .eq("academic_year", academic_year) \
                .eq("semester", semester) \
                .is_("is_active", True)
            
            if day:
                query = query.eq("day_of_week", day)
            
            breaks = query.execute().data or []
            return jsonify(success=True, breaks=breaks)
        
        elif request.method == "POST":
            d = request.json or {}
            academic_year = d.get("academic_year", "2025-26")
            semester = d.get("semester", 1)
            breaks_list = d.get("breaks", [])
            
            if not breaks_list:
                return jsonify(success=False, error="breaks list is required"), 400
            
            # Prepare break data
            break_data = []
            for brk in breaks_list:
                break_data.append({
                    "academic_year": academic_year,
                    "semester": semester,
                    "day_of_week": brk.get("day_of_week"),
                    "start_time": brk.get("start_time"),
                    "end_time": brk.get("end_time"),
                    "break_type": brk.get("break_type", "break"),  # 'break' or 'lunch'
                    "description": brk.get("description", ""),
                    "is_active": True
                })
            
            try:
                # Delete existing breaks first
                sb.table("break_schedule").delete() \
                    .eq("academic_year", academic_year) \
                    .eq("semester", semester).execute()
                
                # Insert new breaks
                result = sb.table("break_schedule").insert(break_data).execute()
                return jsonify(success=True, message=f"Added {len(break_data)} breaks",
                             breaks=result.data)
            except Exception as e:
                return jsonify(success=False, error=f"Could not save breaks: {e}"), 500
    
    except Exception as e:
        logger.exception("Error in manage_breaks")
        return jsonify(success=False, error=str(e)), 500


@app.route("/api/timetable/generate-v2", methods=["POST"])
def generate_timetable_v2():
    """Enhanced timetable generation with break awareness and class distribution constraints"""
    if not sb:
        return jsonify(success=False, error="Supabase not configured"), 500
    
    try:
        d = request.json or {}
        academic_year = d.get("academic_year", "2025-26")
        semester = d.get("semester", 1)
        algorithm = d.get("algorithm", "simulated_annealing")
        
        logger.info(f"[Timetable v2] Starting generation for {academic_year} sem {semester}")
        
        # Load configuration with error handling
        try:
            breaks = sb.table("break_schedule").select("*") \
                .eq("academic_year", academic_year) \
                .eq("semester", semester) \
                .is_("is_active", True) \
                .execute().data or []
            logger.info(f"[Timetable v2] Loaded {len(breaks)} break schedules")
        except Exception as e:
            logger.warning(f"[Timetable v2] Could not load break schedules: {e}")
            breaks = []
        
        # Load faculty assignments
        try:
            assignments = sb.table("faculty_assignments").select("*") \
                .eq("academic_year", academic_year) \
                .eq("semester", semester) \
                .is_("is_active", True) \
                .execute().data or []
            logger.info(f"[Timetable v2] Loaded {len(assignments)} faculty assignments")
        except Exception as e:
            logger.error(f"[Timetable v2] Error loading faculty assignments: {e}")
            assignments = []
        
        if not assignments:
            return jsonify(success=False, error="No faculty assignments found for this semester"), 400
        
        # Load subjects
        try:
            subjects = sb.table("subjects").select("*") \
                .eq("semester", semester) \
                .is_("is_active", True) \
                .execute().data or []
            logger.info(f"[Timetable v2] Loaded {len(subjects)} subjects")
        except Exception as e:
            logger.warning(f"[Timetable v2] Could not load subjects: {e}")
            subjects = []
        
        # Load rooms with better fallback
        try:
            rooms = sb.table("room_capacity").select("*") \
                .eq("academic_year", academic_year) \
                .is_("is_available", True) \
                .execute().data or []
            
            # Fallback: if no available rooms, get all rooms
            if not rooms:
                logger.warning("[Timetable v2] No available rooms, loading all rooms")
                rooms = sb.table("room_capacity").select("*") \
                    .eq("academic_year", academic_year) \
                    .execute().data or []
            
            logger.info(f"[Timetable v2] Loaded {len(rooms)} rooms")
        except Exception as e:
            logger.error(f"[Timetable v2] Error loading rooms: {e}")
            rooms = []
        
        if not rooms:
            return jsonify(success=False, error="No rooms configured for this academic year"), 400
        
        # Validate assignment data
        valid_assignments = []
        for asgn in assignments:
            required_fields = ['section', 'year', 'subject_code', 'subject_name', 'faculty_username']
            if all(asgn.get(field) for field in required_fields):
                valid_assignments.append(asgn)
            else:
                logger.warning(f"[Timetable v2] Skipping invalid assignment: {asgn}")
        
        if not valid_assignments:
            return jsonify(success=False, error="No valid faculty assignments (missing required fields)"), 400
        
        # Generate timetable with constraints
        generated_slots = generate_timetable_with_contraints(
            assignments=valid_assignments,
            subjects=subjects,
            rooms=rooms,
            breaks=breaks,
            academic_year=academic_year,
            semester=semester,
            algorithm=algorithm
        )
        
        if not generated_slots:
            return jsonify(success=False, 
                error="Could not generate valid timetable. Check constraints and data."), 400
        
        logger.info(f"[Timetable v2] Successfully generated {len(generated_slots)} slots")
        
        return jsonify(
            success=True,
            message="Timetable generated successfully",
            slots_count=len(generated_slots),
            slots=generated_slots[:50]  # Return first 50 slots
        )
    
    except Exception as e:
        logger.exception(f"[Timetable v2] Unexpected error in generate_timetable_v2: {e}")
        import traceback
        return jsonify(success=False, error=f"Server error: {str(e)}", trace=traceback.format_exc()), 500


def generate_timetable_with_contraints(assignments, subjects, rooms, breaks, academic_year, semester, algorithm):
    """
    Production-ready timetable generation using constraint satisfaction.
    
    Algorithm:
    1. Uses three conflict trackers: facultyBusy, roomBusy, sectionBusy
    2. Iterates: Year → Section → Subject (subject-first approach)
    3. Theory slots: Use periods [1,2,3,5] (avoid period 4 lunch)
    4. Lab slots: Consecutive afternoon pairs [[6,7],[7,8],[5,6]]
    5. Faculty assignment: Round-robin rotation
    6. Room uniqueness: No faculty uses same room twice per day
    7. Attempts: 80 for theory, 50 for labs
    """
    
    import random
    
    logger.info("[Timetable] Starting generation with constraint satisfaction")
    
    # Period configuration
    PERIODS = {
        1: {'start': '09:00', 'end': '10:00', 'shift': 1},
        2: {'start': '10:00', 'end': '11:00', 'shift': 1},
        3: {'start': '11:15', 'end': '12:15', 'shift': 1},
        4: {'start': '12:15', 'end': '01:15', 'shift': 1},  # Lunch - excluded from theory
        5: {'start': '02:00', 'end': '03:00', 'shift': 2},
        6: {'start': '03:00', 'end': '04:00', 'shift': 2},
        7: {'start': '04:00', 'end': '05:00', 'shift': 2},
        8: {'start': '05:00', 'end': '06:00', 'shift': 2},
    }
    
    THEORY_PERIODS = [1, 2, 3, 5]  # Excluding lunch (4) and afternoon labs (6,7,8)
    LAB_PAIRS = [[6, 7], [7, 8], [5, 6]]  # Consecutive afternoon periods for 2-hour labs
    DAYS = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday']
    
    # ── CONFLICT TRACKERS ──
    faculty_busy = {}    # key: "faculty|day|period" → True if occupied
    room_busy = {}       # key: "room|day|period" → True if occupied
    section_busy = {}    # key: "section|day|period" → True if occupied
    
    def conflict_key_fac(fac, day, period):
        return f"{fac}|{day}|{period}"
    
    def conflict_key_room(room, day, period):
        return f"{room}|{day}|{period}"
    
    def conflict_key_sec(sec, day, period):
        return f"{sec}|{day}|{period}"
    
    def can_assign(fac, sec, day, period, room):
        """Check if slot is free in all three conflict trackers"""
        if not fac or not room:
            return False
        return (conflict_key_fac(fac, day, period) not in faculty_busy and
                conflict_key_room(room, day, period) not in room_busy and
                conflict_key_sec(sec, day, period) not in section_busy)
    
    def commit(fac, sec, day, period, room, code, name, slot_type, year, prog):
        """Mark slot as occupied in all trackers and add to results"""
        try:
            faculty_busy[conflict_key_fac(fac, day, period)] = True
            room_busy[conflict_key_room(room, day, period)] = True
            section_busy[conflict_key_sec(sec, day, period)] = True
            
            p = PERIODS.get(period, PERIODS[1])
            generated_slots.append({
                'academic_year': academic_year,
                'semester': semester,
                'section': sec,
                'year': year,
                'subject_code': code,
                'subject_name': name,
                'faculty_username': fac,
                'day_of_week': day,
                'period_number': period,
                'start_time': p['start'],
                'end_time': p['end'],
                'room_number': room,
                'type': slot_type,
                'program': prog,
                'shift': p['shift'],
                'is_active': True,
            })
        except Exception as e:
            logger.error(f"[Timetable] Error in commit: {e}")
            raise
    
    def get_rooms_used_today(fac, day):
        """Get set of rooms already used by faculty on this day"""
        used = set()
        for slot in generated_slots:
            if slot.get('faculty_username') == fac and slot.get('day_of_week') == day:
                used.add(slot.get('room_number', ''))
        return used
    
    # ── PREPARE DATA ──
    generated_slots = []
    conflicts_list = []
    
    # Group assignments by section and year
    section_year_subjects = {}  # (section, year) → [(code, name, faculty), ...]
    for asgn in assignments:
        try:
            sec = asgn.get('section', '').strip()
            yr = int(asgn.get('year', 1))
            code = asgn.get('subject_code', '').strip()
            name = asgn.get('subject_name', '').strip()
            fac = asgn.get('faculty_username', '').strip()
            dept = asgn.get('department', '').strip()
            prog = asgn.get('program', 'B.Tech').strip()
            
            if code and name and fac and sec:
                key = (sec, yr)
                if key not in section_year_subjects:
                    section_year_subjects[key] = []
                section_year_subjects[key].append({
                    'code': code,
                    'name': name,
                    'faculty': fac,
                    'dept': dept,
                    'prog': prog,
                })
        except Exception as e:
            logger.warning(f"[Timetable] Skipping malformed assignment: {asgn} - {e}")
            continue
    
    if not section_year_subjects:
        logger.error("[Timetable] No valid section-year-subject mappings found")
        return []
    
    room_list = [str(rm.get('room_number', '')).strip() for rm in rooms if rm.get('room_number')]
    room_list = [r for r in room_list if r]  # Remove empty strings
    
    if not room_list:
        logger.error("[Timetable] No valid rooms available")
        return []
    
    faculty_list = list(set(a.get('faculty_username', '').strip() for a in assignments if a.get('faculty_username')))
    faculty_list = [f for f in faculty_list if f]  # Remove empty strings
    
    if not faculty_list:
        logger.error("[Timetable] No valid faculty members found")
        return []
    
    logger.info(f"[Timetable] Prepared: {len(section_year_subjects)} section-years, {len(room_list)} rooms, {len(faculty_list)} faculty")
    
    # ── MAIN SCHEDULING LOOP ──
    # Iterate: Year → Section → Subject
    subject_count = 0
    for (section, year) in sorted(section_year_subjects.keys()):
        subj_list = section_year_subjects[(section, year)]
        fac_idx = 0
        
        for subj in subj_list:
            try:
                code = subj.get('code', 'UNKNOWN')
                name = subj.get('name', 'Unknown Subject')
                fac = faculty_list[fac_idx % len(faculty_list)] if faculty_list else subj.get('faculty', 'UNKNOWN')
                fac_idx += 1
                prog = subj.get('prog', 'B.Tech')
                subject_count += 1
                
                # ── PLACE 3 THEORY SLOTS ──
                theory_target = 3
                theory_placed = 0
                attempts = 0
                used_theory_days = set()
                
                while theory_placed < theory_target and attempts < 80:
                    attempts += 1
                    day = random.choice(DAYS)
                    
                    # Don't place same subject twice on same day
                    if day in used_theory_days:
                        continue
                    
                    period = random.choice(THEORY_PERIODS)
                    
                    # Get available rooms (not used by this faculty today)
                    rooms_used = get_rooms_used_today(fac, day)
                    avail_rooms = [r for r in room_list if r not in rooms_used]
                    room = random.choice(avail_rooms) if avail_rooms else room_list[0]
                    
                    if can_assign(fac, section, day, period, room):
                        commit(fac, section, day, period, room, code, name, 'Theory', year, prog)
                        used_theory_days.add(day)
                        theory_placed += 1
                
                if theory_placed < theory_target:
                    conflicts_list.append(f"{code} (Sec {section}): Only {theory_placed}/{theory_target} theory slots")
                
                # ── PLACE 2 LAB SESSIONS (2 consecutive hours each) ──
                lab_target = 2
                lab_placed = 0
                attempts = 0
                used_lab_days = set()
                
                while lab_placed < lab_target and attempts < 50:
                    attempts += 1
                    day = random.choice(DAYS)
                    
                    # Don't place same lab twice on same day
                    if day in used_lab_days:
                        continue
                    
                    lab_pairs = LAB_PAIRS[:]
                    random.shuffle(lab_pairs)
                    
                    for pair in lab_pairs:
                        p1, p2 = pair
                        
                        # Get lab rooms
                        rooms_used = get_rooms_used_today(fac, day)
                        avail_labs = [r for r in room_list if r not in rooms_used]
                        lab_room = random.choice(avail_labs) if avail_labs else room_list[0]
                        
                        if can_assign(fac, section, day, p1, lab_room) and can_assign(fac, section, day, p2, lab_room):
                            # Commit both periods as one 2-hour lab block
                            p1_cfg = PERIODS.get(p1, PERIODS[1])
                            p2_cfg = PERIODS.get(p2, PERIODS[1])
                            
                            faculty_busy[conflict_key_fac(fac, day, p1)] = True
                            faculty_busy[conflict_key_fac(fac, day, p2)] = True
                            room_busy[conflict_key_room(lab_room, day, p1)] = True
                            room_busy[conflict_key_room(lab_room, day, p2)] = True
                            section_busy[conflict_key_sec(section, day, p1)] = True
                            section_busy[conflict_key_sec(section, day, p2)] = True
                            
                            # Single entry for 2-hour lab block
                            generated_slots.append({
                                'academic_year': academic_year,
                                'semester': semester,
                                'section': section,
                                'year': year,
                                'subject_code': code + '_Lab',
                                'subject_name': name + ' (Lab)',
                                'faculty_username': fac,
                                'day_of_week': day,
                                'period_number': p1,
                                'start_time': p1_cfg['start'],
                                'end_time': p2_cfg['end'],  # 2-hour block
                                'room_number': lab_room,
                                'type': 'Lab',
                                'program': prog,
                                'shift': p1_cfg['shift'],
                                'lab_pair': [p1, p2],
                                'is_active': True,
                            })
                            
                            lab_placed += 1
                            used_lab_days.add(day)
                            break
                
                if lab_placed < lab_target:
                    conflicts_list.append(f"{code} Lab (Sec {section}): Only {lab_placed}/{lab_target} lab sessions")
                    
            except Exception as e:
                logger.error(f"[Timetable] Error scheduling subject {subj.get('code', 'UNKNOWN')}: {e}")
                conflicts_list.append(f"Error: {str(e)[:100]}")
                continue
    
    # Add conflicts to logs
    if conflicts_list:
        logger.warning(f"[Timetable] Generation warnings ({len(conflicts_list)}): {conflicts_list[:10]}")
    
    logger.info(f"[Timetable] Generated {len(generated_slots)} slots for {subject_count} subjects in {len(section_year_subjects)} section-years")
    
    return generated_slots


# ── FIRESTORE DATA RETRIEVAL & SYNC ENDPOINTS ─────────────────────────

@app.route("/api/firestore/attendance/<roll_no>", methods=["GET"])
def get_firestore_attendance(roll_no):
    """Retrieve all attendance records for a student from Firestore."""
    try:
        if not _fstore:
            return jsonify(success=False, error="Firestore not available"), 503
        
        # Query all attendance records for this roll_no
        docs = _fstore.collection("attendance").where("roll_no", "==", roll_no).stream()
        records = []
        for doc in docs:
            data = doc.to_dict()
            data['firestore_id'] = doc.id
            records.append(data)
        
        return jsonify(
            success=True,
            count=len(records),
            records=records
        ), 200
    
    except Exception as e:
        print(f"[FIRESTORE] Error retrieving attendance: {e}")
        return jsonify(success=False, error=str(e)), 500


@app.route("/api/firestore/user/<user_id>", methods=["GET"])
def get_firestore_user(user_id):
    """Retrieve user data from Firestore."""
    try:
        if not _fstore:
            return jsonify(success=False, error="Firestore not available"), 503
        
        doc = _fstore.collection("users").document(user_id).get()
        if doc.exists:
            data = doc.to_dict()
            return jsonify(success=True, user=data), 200
        else:
            return jsonify(success=False, error="User not found in Firestore"), 404
    
    except Exception as e:
        print(f"[FIRESTORE] Error retrieving user: {e}")
        return jsonify(success=False, error=str(e)), 500


@app.route("/api/firestore/sync-status", methods=["GET"])
def firestore_sync_status():
    """Check if Firestore is available and syncing."""
    try:
        if not _fstore:
            return jsonify(
                available=False,
                message="Firebase Admin SDK not initialized"
            ), 503
        
        # Test basic connection
        docs = _fstore.collection("_test").limit(1).stream()
        list(docs)  # Force fetch
        
        return jsonify(
            available=True,
            message="Firestore is operational",
            url=os.getenv('FIREBASE_CONFIG_URL', 'not configured')
        ), 200
    
    except Exception as e:
        print(f"[FIRESTORE] Sync status check failed: {e}")
        return jsonify(
            available=False,
            error=str(e)
        ), 503


@app.route("/api/firestore/batch-sync", methods=["POST"])
def firestore_batch_sync():
    """Admin endpoint: Sync batch of records to Firestore from Supabase."""
    try:
        data = request.json or {}
        collection = data.get("collection")  # "users", "attendance", etc.
        filter_query = data.get("filter")  # Optional SQL-like filter
        limit = data.get("limit", 100)
        
        if not collection:
            return jsonify(success=False, error="Collection name required"), 400
        
        if not _fstore or not sb:
            return jsonify(success=False, error="Firestore or Supabase not available"), 503
        
        # Fetch from Supabase
        query = sb.table(collection).select("*").limit(limit)
        if filter_query:
            # Simple filter support (e.g., "role=eq.student")
            for part in filter_query.split("&"):
                if "=" in part:
                    k, v = part.split("=", 1)
                    query = query.eq(k.strip(), v.strip())
        
        result = query.execute()
        records = result.data if result.data else []
        
        # Batch write to Firestore
        batch = _fstore.batch()
        for record in records:
            doc_id = str(record.get("id", uuid.uuid4()))
            ref = _fstore.collection(collection).document(doc_id)
            batch.set(ref, record, merge=True)
        
        batch.commit()
        
        return jsonify(
            success=True,
            message=f"Synced {len(records)} records from {collection} to Firestore",
            synced_count=len(records)
        ), 200
    
    except Exception as e:
        print(f"[FIRESTORE] Batch sync error: {e}")
        return jsonify(success=False, error=str(e)), 500


# ── SESSION VALIDATION (Production-ready for Firestore sessions) ──────────────
@app.route("/api/session/validate", methods=["POST"])
def validate_session():
    """Validate session exists in Firestore.
    
    Frontend calls this to verify its Firestore session is still valid.
    Returns user data if valid, 401 if not found/expired.
    """
    try:
        data = request.json or {}
        username = data.get("username")
        
        if not username:
            return jsonify(valid=False, error="No username provided"), 400
        
        # Check if Firestore is available
        if not _fstore:
            return jsonify(valid=False, error="Session backend unavailable"), 503
        
        # Read session from Firestore with timeout
        try:
            doc = _fstore.collection('sessions').document(username).get(timeout=3)
            if doc.exists:
                session_data = doc.to_dict()
                user_data = session_data.get('user', {})
                return jsonify(
                    valid=True,
                    user={
                        'id': user_data.get('id'),
                        'username': user_data.get('username'),
                        'role': user_data.get('role'),
                        'full_name': user_data.get('full_name'),
                        'email': user_data.get('email')
                    }
                ), 200
            else:
                return jsonify(valid=False, error="Session not found"), 401
        except Exception as e:
            print(f"[SESSION] Firestore read error: {str(e)}")
            return jsonify(valid=False, error="Session validation failed"), 500
            
    except Exception as e:
        print(f"[SESSION] Error validating session: {str(e)}")
        return jsonify(valid=False, error="Server error"), 500


# ── Register Enhanced Bulk Import Routes ────────────────────────────
if BULK_ROUTES_AVAILABLE:
    try:
        register_bulk_routes(app, None)  # Pass None for db (not needed for routes)
        print("[BULK_ROUTES] ✓ Enhanced bulk import routes registered")
    except Exception as e:
        print(f"[BULK_ROUTES] ⚠ Failed to register bulk routes: {e}")
else:
    print("[BULK_ROUTES] Bulk import routes not available")

# ── Register Analytics Endpoints (Linways-like Dashboard) ────────────
if ANALYTICS_AVAILABLE:
    try:
        register_analytics_endpoints(app, db)
        print("[ANALYTICS] ✓ Analytics dashboard endpoints registered")
    except Exception as e:
        print(f"[ANALYTICS] ⚠ Failed to register analytics endpoints: {e}")
else:
    print("[ANALYTICS] Analytics endpoints not available")


if __name__=="__main__":

    port = int(os.environ.get("PORT", 6001))
    is_production = "PORT" in os.environ
    
    if is_production:
        print(f"SmartAMS Backend — Cloud Run Mode (port {port})")
    else:
        print("SmartAMS Backend — http://localhost:6001")
    
    print("QR Security System — ENABLED")
    
    # ── ADMIN: Initialize first admin user ──────────────────────────
    @app.route("/api/init-admin", methods=["POST"])
    def init_admin():
        """Initialize the first admin user (one-time setup endpoint)"""
        try:
            data = request.json
            password = data.get('password', '')
            password_confirm = data.get('password_confirm', '')
            
            if not password or password != password_confirm:
                return jsonify(error="Passwords don't match"), 400
            
            if len(password) < 8:
                return jsonify(error="Password must be at least 8 characters"), 400
            
            # Check if admin_demo already exists
            result = sb.table("users").select("id").eq("username", "admin_demo").execute()
            
            if result.data:
                # Update existing admin_demo with new password
                pwd_hash = _hash_password_secure(password)
                update_result = sb.table("users").update(
                    {"password_hash": pwd_hash}
                ).eq("username", "admin_demo").execute()
                
                if update_result.data:
                    return jsonify(success=True, message="Admin password updated", username="admin_demo")
            
            return jsonify(error="Admin user not found"), 404
                
        except Exception as e:
            logger.error(f"[INIT] Admin init error: {e}")
            return jsonify(error=str(e)), 500
    
    @app.route("/api/init-admin-demo", methods=["GET", "POST"])
    def create_demo_admins():
        """Create both demo admin users (admin_demo and superadmin)"""
        try:
            if not sb:
                return jsonify(success=False, error="Database not configured"), 500
            
            results = {}
            
            # Create admin_demo
            try:
                pwd_hash_1 = _hash_password_secure("Admin@123")
                user_data_1 = {
                    "id": str(__import__('uuid').uuid4()),
                    "username": "admin_demo",
                    "email": "admin@smartams.demo",
                    "full_name": "Demo Admin",
                    "role": "admin",
                    "password_hash": pwd_hash_1,
                    "is_active": True,
                    "department": "Administration",
                }
                
                # Delete if exists
                try:
                    sb.table("users").delete().eq("username", "admin_demo").execute()
                except:
                    pass
                
                sb.table("users").insert(user_data_1).execute()
                results["admin_demo"] = {
                    "success": True,
                    "username": "admin_demo",
                    "password": "Admin@123",
                    "email": "admin@smartams.demo"
                }
            except Exception as e:
                results["admin_demo"] = {"success": False, "error": str(e)}
            
            # Create superadmin
            try:
                pwd_hash_2 = _hash_password_secure("SuperAdmin@456")
                user_data_2 = {
                    "id": str(__import__('uuid').uuid4()),
                    "username": "superadmin",
                    "email": "superadmin@smartams.demo",
                    "full_name": "Super Admin",
                    "role": "admin",
                    "password_hash": pwd_hash_2,
                    "is_active": True,
                    "department": "Administration",
                }
                
                # Delete if exists
                try:
                    sb.table("users").delete().eq("username", "superadmin").execute()
                except:
                    pass
                
                sb.table("users").insert(user_data_2).execute()
                results["superadmin"] = {
                    "success": True,
                    "username": "superadmin",
                    "password": "SuperAdmin@456",
                    "email": "superadmin@smartams.demo"
                }
            except Exception as e:
                results["superadmin"] = {"success": False, "error": str(e)}
            
            logger.info(f"[INIT] Demo admins created: {results}")
            return jsonify(success=True, admins=results), 200
            
        except Exception as e:
            logger.error(f"[INIT] Error creating demo admins: {e}")
            return jsonify(success=False, error=str(e)), 500
    
    # ── Register RBAC Analytics Routes ──────────────────────────────
    if RBAC_AVAILABLE:
        try:
            register_rbac_analytics_routes(app, sb=sb)
            logger.info("[APP] ✓ RBAC analytics routes registered")
        except Exception as e:
            logger.warning(f"[APP] Could not register RBAC routes: {e}")
    
    # Disable debug and reloader to prevent hanging requests
    # Debug mode with Flask's debugger can cause request timeouts
    app.run(debug=False, use_reloader=False, host="0.0.0.0", port=port, threaded=True)

